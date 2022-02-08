import datetime
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global FundNameList
  global FundNameListAfterRemove
  global ct
  global Exe
  global D1
  global D2
  global d1
  global d2
  global DollarDate
  global FundToOpen
  global TotalFundsLengh

  TestName = "test_BIDS630"
  description = "This test scenario is to verify working of BIDS630 User story"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_SprintUS"
  Exe="Yes"
  Directory = 'test_SprintUS/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  FundNameList=[]
  FundNameListAfterRemove=[]

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()

      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      today = datetime.date.today()
      D1=today.strftime("%Y-%m-%d")
      d1=D1
      DollarDate=datetime.datetime.strptime(d1, '%Y-%m-%d')
      DollarDate="$"+DollarDate.date().__str__()+"$"
      d1 = datetime.datetime.strptime(D1, "%Y-%m-%d")

  yield
  if Exe == "Yes":
      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        try:
            print()
            # ----------------------------------------------------------------------------
            #---------------------------Verify Funds page-----------------------------
            PageName="Quarterly NAV Close"
            Ptitle1="Quarterly NAV Close - BIDS"
            driver.find_element_by_xpath("//*[@title='"+PageName+"']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                PageTitle1 = driver.title
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(PageName + " page opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName +     " page not able to open")
                TestResultStatus.append("Fail")

            PageName = "Look Back Analysis"
            Ptitle1 = "NAV Actual to Estimate Analysis - BIDS"
            driver.find_element_by_xpath("//*[text() = '" + PageName + "']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")

            Inside="Estimate to Actual Summary"
            # # --------------------Clicking a Fund--------------
            # try:
            #     button = driver.find_element_by_xpath(
            #         "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[1]/td[2]/div/p/a")
            #     driver.execute_script("arguments[0].click();", button)
            #     TestResult.append("A Fund clicked successfully")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append("A Fund after click not able to open")
            #     TestResultStatus.append("Fail")

            # ------For the BEN Reporting Period Ended---------
            Text1 = "For the BEN Reporting Period Ended"
            Type="section header text"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div[2]/div/p/strong").text
                assert Text1 in Element1, Text1+" "+Type+" is not present inside "+Inside
                TestResult.append(Text1+" "+Type+" is present inside "+Inside)
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1+" "+Type+" is not present inside "+Inside)
                TestResultStatus.append("Fail")

            # ------Ben Reported NAV - US$---------
            Text1 = "Ben Reported NAV - US$"
            Type = "label text"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[1]/div/p").text
                assert Text1 in Element1, Text1 + " " + Type + " is not present inside " + Inside
                TestResult.append(Text1 + " " + Type + " is present inside " + Inside)
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " " + Type + " is not present inside " + Inside)
                TestResultStatus.append("Fail")

            # ------Updated NAV - US$---------
            Text1 = "Updated NAV - US$"
            Type = "label text"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[2]/div/div/table/tbody/tr[2]/td[1]/div/p").text
                assert Text1 in Element1, Text1 + " " + Type + " is not present inside " + Inside
                TestResult.append(Text1 + " " + Type + " is present inside " + Inside)
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " " + Type + " is not present inside " + Inside)
                TestResultStatus.append("Fail")

            # ------Variance - US$---------
            Text1 = "Variance - US$"
            Type = "label text"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[2]/div/div/table/tbody/tr[3]/td[1]/div/p").text
                assert Text1 in Element1, Text1 + " " + Type + " is not present inside " + Inside
                TestResult.append(Text1 + " " + Type + " is present inside " + Inside)
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " " + Type + " is not present inside " + Inside)
                TestResultStatus.append("Fail")

            # ------Variance - %---------
            Text1 = "Variance - %"
            Type = "label text"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[2]/div/div/table/tbody/tr[4]/td[1]/div/p").text
                assert Text1 in Element1, Text1 + " " + Type + " is not present inside " + Inside
                TestResult.append(Text1 + " " + Type + " is present inside " + Inside)
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " " + Type + " is not present inside " + Inside)
                TestResultStatus.append("Fail")

            # ------Checking Materiality Threshold % label---------
            Text1 = "Materiality Threshold %"
            Type = "label text"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/label").text
                assert Text1 in Element1, Text1 + " " + Type + " is not present"
                TestResult.append(Text1 + " " + Type + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " " + Type + " is not present")
                TestResultStatus.append("Fail")

            # ------Percentage value Checking Materiality Threshold %---------
            Text1 = "5%"
            Type = "Materiality Threshold % value"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[2]/div/div[2]/div/input").get_attribute('value')
                assert Text1 in Element1, Text1 + " " + Type + " is not present"
                TestResult.append(Text1 + " " + Type + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " " + Type + " is not present")
                TestResultStatus.append("Fail")

            # ------Filter By dropdown Values---------
            DropDownNumber=5
            for ii3 in range(DropDownNumber):
                print()
                print()
                print(ii3)

                if ii3>0:
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[1]/div/div[2]/div/div").click()
                    for ii4 in range(1):
                        time.sleep(1)
                        ActionChains(driver).key_down(Keys.DOWN).perform()
                        time.sleep(1)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                Value=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[1]/div/div[2]/div/div/span").text
                print(Value)

                try:
                    FundsNumber=driver.find_elements_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/tbody/tr")
                    print(len(FundsNumber))
                    print("FundsNumber Present " + str(len(FundsNumber) - 1))
                except Exception:
                    TestResult.append(
                        "For " + Value + " Fund list table not present")
                    TestResultStatus.append("Fail")
                    pass

                #--------vlaue check for -----------
                if ii3==0:
                    TestResult.append("For Filter By= " + Value + " total [ " + str(
                        len(FundsNumber) - 1) + " ] Funds available in the list")
                    TestResultStatus.append("Pass")

                if ii3==1:
                    try:
                        try:
                            FundAudited = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/tbody/tr[last()]/td[9]/div/p/strong").text
                            print("FundAudited is " + FundAudited)
                        except Exception:
                            FundAudited = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/tbody/tr[last()]/td[9]/div/p/span/strong").text
                            print("FundAudited is " + FundAudited)

                        TestResult.append("For Filter By= " + Value + " total [ " + str(
                        len(FundsNumber) - 1) + " ] Funds available in the list\nFund Audited value found [ " + FundAudited + " ]")
                        TestResultStatus.append("Pass")
                    except Exception:
                        TestResult.append(
                            "For " + Value + " \nFund Audited value data not present")
                        TestResultStatus.append("Fail")

                if ii3 == 2:
                    try:
                        try:
                            FundNOTAudited = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/tbody/tr[last()]/td[10]/div/p/strong").text
                            print("FundNOTAudited is " + FundNOTAudited)
                        except Exception:
                            FundNOTAudited = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/tbody/tr[last()]/td[10]/div/p/span/strong").text
                            print("FundNOTAudited is " + FundNOTAudited)

                        TestResult.append("For Filter By= " + Value + " total [ " + str(
                        len(FundsNumber) - 1) + " ] Funds available in the list\nFund NOT Audited value found [ " + FundNOTAudited + " ]")
                        TestResultStatus.append("Pass")
                    except Exception:
                        TestResult.append(
                            "For " + Value + " \nFund NOT Audited value data not present")
                        TestResultStatus.append("Fail")

                if ii3 == 3:
                    try:
                        try:
                            NoValueReported = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/tbody/tr[last()]/td[12]/div/p/strong").text
                            print("NoValueReported is " + NoValueReported)
                        except Exception:
                            NoValueReported = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/tbody/tr[last()]/td[12]/div/p/span/strong").text
                            print("NoValueReported is " + NoValueReported)

                        TestResult.append("For Filter By= " + Value + "  total [ " + str(
                        len(FundsNumber) - 1) + " ] Funds available in the list\nNo Value Reported value found [ " + NoValueReported + " ]")
                        TestResultStatus.append("Pass")
                    except Exception:
                        TestResult.append(
                            "For " + Value + " \nNo Value Reported value data not present")
                        TestResultStatus.append("Fail")

                if ii3 == 4:
                    try:
                        try:
                            BenCalculatedPartnerNAV = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/tbody/tr[last()]/td[11]/div/p/strong").text
                            print("BenCalculatedPartnerNAV is " + BenCalculatedPartnerNAV)
                        except Exception:
                            BenCalculatedPartnerNAV = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/tbody/tr[last()]/td[11]/div/p/span/strong").text
                            print("BenCalculatedPartnerNAV is " + BenCalculatedPartnerNAV)

                        TestResult.append("For Filter By= " + Value + " total [ " + str(
                        len(FundsNumber) - 1) + " ] Funds available in the list\nBen Calculated Partner NAV value found [ " + BenCalculatedPartnerNAV + " ]")
                        TestResultStatus.append("Pass")
                    except Exception:
                        TestResult.append(
                            "For " + Value + " \nBen Calculated Partner NAV value data not present")
                        TestResultStatus.append("Fail")

            # ------Sort Field---------
            Text1 = "Sort Field"
            Type = "label text"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[3]/div/div[1]/span").text
                print(Element1)
                assert Text1 in Element1, Text1 + " " + Type + " is not present"
                TestResult.append(Text1 + " " + Type + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " " + Type + " is not present")
                TestResultStatus.append("Fail")

            # ------Sort Field dropdown Values---------
            DropDownNumber = 5
            for ii5 in range(DropDownNumber):
                print()
                print()
                print(ii5)
                if ii5>0:
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[3]/div/div[2]/div/div").click()
                    for ii6 in range(1):
                        time.sleep(1)
                        ActionChains(driver).key_down(Keys.DOWN).perform()
                        time.sleep(1)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                try:
                    Value = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[3]/div/div[2]/div/div/span").text
                    print(Value)
                    TestResult.append(Value+" drop down value in "+Text1 +" is present")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append(Value+" drop down value in "+Text1 +" is not present")
                    TestResultStatus.append("Fail")

            # ------Ben Reported NAV US after F/X Rate---------
            Text1 = "F/X Rate"
            try:
                for ii7 in range(15):
                    Element1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/thead/tr[1]/th["+str(ii7+1)+"]/div").text
                    if Text1 in Element1:
                        print("-------------------"+Element1)
                        Element2 = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/div/div/div[2]/table/thead/tr[1]/th[" + str(
                                ii7 + 2) + "]/div").text
                        print(Element2)
                        if "US$" in Element2:
                            TestResult.append("Ben Reported NAV as of Reporting Date - US$ is immediately after "+Text1 + " column in the Fund listing")
                            TestResultStatus.append("Pass")
                        else:
                            TestResult.append("Ben Reported NAV as of Reporting Date - US$ is not immediately after "+Text1 + " column in the Fund listing")
                            TestResultStatus.append("Fail")
                        break
            except Exception as e1:
                pass

            # ------Sort Field---------
            Text1 = "Export Data to Excel"
            Type = "link text"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div[3]/div[2]/div/p/strong/a").text
                assert Text1 in Element1, Text1 + " " + Type + " is not present"
                TestResult.append(Text1 + " " + Type + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " " + Type + " is not present")
                TestResultStatus.append("Fail")


        except Exception as Mainerror:
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


