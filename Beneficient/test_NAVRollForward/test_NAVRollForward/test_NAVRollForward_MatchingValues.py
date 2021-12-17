import datetime
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global FundNameList
  global FundNameListAfterRemove
  global ct
  global Exe
  global D1
  global D2
  global d1
  global d2
  global DollarDate
  global FundToOpen
  global TotalFundsLengh

  TestName = "test_NAVRollForward_MatchingValues"
  description = "This test scenario is to match Ben NAV - LC [ FUND NAV ROLL ] with GP-Reported FairValue [ Investments ] section of all funds in different quarters"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_NAVRollForward"
  Exe="Yes"
  Directory = 'test_NAVRollForward/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  FundNameList=[]
  FundNameListAfterRemove=[]

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()

      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      today = datetime.date.today()
      D1=today.strftime("%Y-%m-%d")
      #d1 = datetime.datetime.strptime(D1, "%Y-%m-%d")
      d1=D1
      #print(d1)
      DollarDate=datetime.datetime.strptime(d1, '%Y-%m-%d')
      DollarDate="$"+DollarDate.date().__str__()+"$"
      #print("$"+DollarDate.date().__str__()+"$")

  yield
  if Exe == "Yes":
      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        FundToOpen=15
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        try:
            print()
            # ----------------------------------------------------------------------------
            #---------------------------Verify Liquid Trusts page-----------------------------
            PageName="Funds"
            Ptitle1="Investments - BIDS"
            driver.find_element_by_xpath("//*[@title='"+PageName+"']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                try:
                    PageTitle1 = driver.title
                    print(PageTitle1)
                    assert PageTitle1 in Ptitle1, PageName + " not able to open"
                except Exception:
                    Ptitle1="Funds - BIDS"
                    PageTitle1 = driver.title
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName +     " page not able to open")
                TestResultStatus.append("Fail")
            print()
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)
            #---------------------------------------------------------------------------------

            RowsFunds = driver.find_elements_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr")
            for ifund in range(len(RowsFunds)):
                FundName=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr["+str(ifund+1)+"]/td[2]/div/p/a").text
                FundNameList.append(FundName)

            if len(RowsFunds) == 100:
                driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                RowsFunds1 = driver.find_elements_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr")
                RowsFunds=RowsFunds+RowsFunds1
                for ifund in range(len(RowsFunds1)):
                    FundName = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                            ifund + 1) + "]/td[2]/div/p/a").text
                    FundNameList.append(FundName)

            if len(RowsFunds) == 200:
                driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                RowsFunds2 = driver.find_elements_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr")
                RowsFunds = RowsFunds + RowsFunds2
                for ifund in range(len(RowsFunds2)):
                    FundName = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                            ifund + 1) + "]/td[2]/div/p/a").text
                    FundNameList.append(FundName)

            if len(RowsFunds) == 300:
                driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                RowsFunds3 = driver.find_elements_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr")
                RowsFunds = RowsFunds + RowsFunds3
                for ifund in range(len(RowsFunds3)):
                    FundName = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                            ifund + 1) + "]/td[2]/div/p/a").text
                    FundNameList.append(FundName)

            print("length of FundNameList before "+str(len(FundNameList)))

            TotalFundsLengh=len(FundNameList)
            print("TotalFundsLengh "+str(TotalFundsLengh))
            print("=======================================================================================")
            driver.refresh()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            #--------------------------------------------------------------------------------------
            #----------------Appending Funds details in Excel sheet-----------------------------------

            # -----------To add new found Funds in Excel sheet-------------
            ExcelFileName = "FundName"
            loc1 = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
            wb1 = openpyxl.load_workbook(loc1)
            sheet1 = wb1.active
            for i in range(len(FundNameList)):
                try:
                    if sheet1.cell(i + 1, 2).value !=None:
                        if sheet1.cell(i + 1, 2).value in FundNameList:
                            FundNameList.remove(sheet1.cell(i + 1, 2).value)
                            #print("Removed from List: "+sheet1.cell(i + 1, 2).value)
                except Exception as ef:
                    print(ef)
                    pass
            wb1.save(loc1)

            print()
            print("length of FundNameList after " + str(len(FundNameList)))

            noneindex=0
            for iadd in range (TotalFundsLengh):
                if sheet1.cell(iadd + 1, 2).value == None:
                    if noneindex==0:
                        noneindex=iadd+1
                    #print(FundNameList[(iadd+1)-noneindex])
                    sheet1.cell(row=iadd + 1, column=1).value = iadd + 1
                    sheet1.cell(row=iadd + 1, column=2).value = FundNameList[iadd-noneindex]
            wb1.save(loc1)
            #--------------------------------------------------------------------------------------

            # -----------To fetch selected Funds from total list-------------
            ExcelFileName = "FundName"
            loc1 = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
            wb1 = openpyxl.load_workbook(loc1)
            sheet1 = wb1.active

            for i2 in range(TotalFundsLengh):
                if len(FundNameListAfterRemove) <= FundToOpen - 1:
                    if sheet1.cell(i2 + 1, 3).value == None:
                        FundNameListAfterRemove.append(sheet1.cell(i2 + 1, 2).value)
                        sheet1.cell(i2 + 1, 3).value=DollarDate
                    else:
                        if sheet1.cell(i2 + 1, 3).value != None:
                            D2 = sheet1.cell(i2 + 1, 3).value
                            D2 = re.sub('[!@#$]', '', D2)
                            d2 = datetime.datetime.strptime(D2, "%Y-%m-%d")
                            if (d1 - d2).days > 7:
                                print(sheet1.cell(i2 + 1, 2).value)
                                FundNameListAfterRemove.append(sheet1.cell(i2 + 1, 2).value)
                                sheet1.cell(i2 + 1, 3).value = DollarDate
                            else:
                                pass
            print(FundNameListAfterRemove)
            print(str(len(FundNameListAfterRemove)))
            TestResult.append("Below " + str(len(FundNameListAfterRemove)) + " Funds are collected for verification")
            TestResultStatus.append("Pass")
            # ----------------------------------------------------------------------------

            #--------------------------------------------------------------------------------------
            for ifundlist in range(len(FundNameListAfterRemove)):
                print(str(ifundlist))
                #-------------to clear the cache memory of chrome browser-------------
                if (ifundlist/30).is_integer()==True :
                    print("<<<<<<<<<<<Integer value found for 30 digit>>>>>>>>")
                    driver.delete_all_cookies()
                    time.sleep(5)
                    driver.get("https://beneficienttest.appiancloud.com/suite/")
                    driver.find_element_by_id("un").send_keys("neeraj.kumar")
                    driver.find_element_by_id("pw").send_keys("Crochet@786")
                    button = driver.find_element_by_xpath(
                        "//input[@type='submit']")
                    driver.execute_script("arguments[0].click();", button)
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                #--------------------------------------------------------------
                if ifundlist == 100:
                    print()
                    print("Clicked 1 after 100")
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass

                if ifundlist == 200:
                    print()
                    print("Clicked 2 after 200")
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass

                if ifundlist == 300:
                    print()
                    print("Clicked 3 after 300")
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                print(FundNameListAfterRemove[ifundlist])
                main_window = driver.current_window_handle
                print(str(ifundlist)+" "+FundNameListAfterRemove[ifundlist])
                if "'" in FundNameListAfterRemove[ifundlist]:
                    print("******************************************  coma found *********************************************")
                    # print(FundNameList[ifundlist])
                    # fund=FundNameList[ifundlist]
                    # fund=fund.replace("'", "\'")
                    # print("afer changing: " + fund)
                    # driver.find_element_by_xpath("//*[text()='" + fund + "']").click()
                    # break
                    #FundNameList[ifundlist].index
                else:
                    ActionChains(driver).key_down(Keys.CONTROL).perform()
                    try:
                        driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='"+FundNameListAfterRemove[ifundlist]+"']").click()
                    except Exception:
                        print("^^^^^^^^^^^^^^^^^^^^^^")
                        button = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='"+FundNameListAfterRemove[ifundlist]+"']")
                        driver.execute_script("arguments[0].click();", button)

                    ActionChains(driver).key_up(Keys.CONTROL).perform()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    #-------------------Inserting Execution Date in Fund sheet-------------------------
                    TestResult.append(FundNameListAfterRemove[ifundlist])
                    TestResultStatus.append("Pass")

                    #----------------------------------------------------------------------------------
                    driver.switch_to.window(driver.window_handles[1])
                    time.sleep(1)
                    try:
                        Test=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[2]/p").text
                    except Exception:
                        time.sleep(3)
                        #---------------When Page load error occurs---------------------------------
                        try:
                            if driver.find_element_by_xpath("//div[@id='main-frame-error']/div/div[2]/h1/span").is_displayed()==True:
                                driver.refresh()
                                try:
                                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                    WebDriverWait(driver, LONG_TIMEOUT
                                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                except TimeoutException:
                                    pass
                        except Exception:
                            pass
                        #-----------------------------------------------------------------------------
                        print("Waiting in Exception")
                        driver.switch_to.window(driver.window_handles[1])
                        try:
                            Test=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[2]/p").text
                        except Exception:
                            try:
                                time.sleep(2)
                                bool1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                                if bool1 == True:
                                    ErrorFound1 = driver.find_element_by_xpath(
                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                    print(ErrorFound1)
                                    driver.find_element_by_xpath(
                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                                    TestResultStatus.append("Fail")
                                    bool1 = False
                                    Test = ErrorFound1
                            except Exception:
                                try:
                                    time.sleep(2)
                                    bool2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                    if bool2 == True:
                                        ErrorFound2 = driver.find_element_by_xpath(
                                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                        print(ErrorFound2)
                                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                        TestResultStatus.append("Fail")
                                        bool2 = False
                                        Test = ErrorFound2
                                except Exception:
                                    pass
                                pass
                    print("Test text is: "+Test)
                    # --------------------Clicking Fund NAV Roll tab--------------
                    try:
                        PageName = "Fund NAV Roll"
                        Ptitle1 = "Ben Reporting Period"
                        driver.find_element_by_xpath("//button[text()='"+PageName+"']").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        try:
                            time.sleep(2)
                            bool1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                            if bool1 == True:
                                ErrorFound1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                print(ErrorFound1)
                                driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                                TestResultStatus.append("Fail")
                                bool1 = False
                                driver.close()
                        except Exception:
                            try:
                                time.sleep(2)
                                bool2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                if bool2 == True:
                                    ErrorFound2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                    print(ErrorFound2)
                                    TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                    TestResultStatus.append("Fail")
                                    bool2 = False
                                    driver.close()
                            except Exception:
                                pass
                            pass
                        time.sleep(1)
                        try:
                            PageTitle1 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[1]/table/thead/tr/th[1]/div").text
                            print(PageTitle1)
                            assert Ptitle1 in PageTitle1, PageName + " is not able to open "
                            TestResult.append(PageName + " opened successfully")
                            TestResultStatus.append("Pass")
                        except Exception as e1:
                            print(e1)
                            TestResult.append(PageName + " is not able to open ")
                            TestResultStatus.append("Fail")
                        #--------------------------------------------------------------------------

                        #-------Fetching Fund NAV ROll value-------------------------------------
                        BenReportingPeriod=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[1]/table/tbody/tr[2]/td[1]/p").text
                        BenNAVLC=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[1]/table/tbody/tr[2]/td[12]/div/p/span").text
                        print("BenReportingPeriod is "+BenReportingPeriod)
                        print("BenNAVLC is " + BenNAVLC)
                        #------------------------------------------------------------------------
                        # -------Fetching Investments value-------------------------------------
                        PageName = "Investments"
                        Ptitle1 = "Edit Schedule of Investments"
                        driver.find_element_by_xpath("//button[text()='" + PageName + "']").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        try:
                            time.sleep(2)
                            bool1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                            if bool1 == True:
                                ErrorFound1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                print(ErrorFound1)
                                driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                                TestResultStatus.append("Fail")
                                bool1 = False
                                driver.close()
                        except Exception:
                            try:
                                time.sleep(2)
                                bool2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                if bool2 == True:
                                    ErrorFound2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                    print(ErrorFound2)
                                    TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                    TestResultStatus.append("Fail")
                                    bool2 = False
                                    driver.close()
                            except Exception:
                                pass
                            pass
                        time.sleep(1)
                        try:
                            PageTitle1 = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div[1]/div[1]/div/div[2]/div/p/a").text
                            print(PageTitle1)
                            assert Ptitle1 in PageTitle1, PageName + " is not able to open "
                            TestResult.append(PageName + " opened successfully")
                            TestResultStatus.append("Pass")
                        except Exception as e1:
                            print(e1)
                            TestResult.append(PageName + " is not able to open ")
                            TestResultStatus.append("Fail")
                        # --------------------------------------------------------------------------
                        # -------Fetching Ben Remaining NAV value-------------------------------------
                        driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div[2]/div[2]/div/div/div/div[2]/div/div[1]/div/div[2]/div/div").click()
                        time.sleep(1)
                        ActionChains(driver).key_down(Keys.DOWN).perform()
                        time.sleep(1)
                        ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        time.sleep(1)
                        BenRemainingNAV = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div[3]/div/div[2]/div/div[3]/div[2]/div/div/table/tbody/tr[last()]/td[25]/div/p/span/strong").text
                        print("BenRemainingNAV is " + BenRemainingNAV)
                        # ------------------------------------------------------------------------
                        # -------Fetching Ben Remaining NAV value-------------------------------------
                        if BenRemainingNAV not in BenNAVLC:
                            TestResult.append("Ben Remaining NAV [Investments] is not matching with Ben NAV LC for Fund [ "+FundNameListAfterRemove[ifundlist]+" ], quarter [ "+BenReportingPeriod+" ]")
                            TestResultStatus.append("Fail")
                        else:
                            TestResult.append(
                                "Ben Remaining NAV [Investments] "+BenRemainingNAV+" matched with Ben NAV LC "+BenNAVLC+" for Fund [ " +
                                FundNameListAfterRemove[ifundlist] + " ], quarter [ " + BenReportingPeriod + " ]")
                            TestResultStatus.append("Pass")
                        # ------------------------------------------------------------------------

                    except Exception:
                        pass

                    for winclose in range(1,10):
                        time.sleep(1)
                        if len(driver.window_handles)>1:
                            #print("Tab Count is more than 1: "+str(len(driver.window_handles)))
                            driver.switch_to.window(driver.window_handles[1])
                            driver.close()
                        elif len(driver.window_handles)==1:
                            break
                    #driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(1)
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass

        except Exception as Mainerror:
            # stop = time.time()
            # RoundFloatString = round(float(stop - start),2)
            # print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


