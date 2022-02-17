import datetime
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_VerifyAllElements_Investments"
  description = "This test scenario is to verify all the Elements present in Investments Page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_Investments"
  global Exe
  Exe="Yes"
  Directory = 'test_Investments/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@7866")
      driver.find_element_by_xpath("//input[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        try:
            print()
            #---------------------------Verify Liquid Trusts page-----------------------------
            PageName="Investments"
            Ptitle1="Investments - BIDS"
            driver.find_element_by_xpath("//*[@title='"+PageName+"']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    driver.close()
                    bool1 = False
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        driver.close()
                        bool2 = False
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                try:
                    PageTitle1 = driver.title
                    print(PageTitle1)
                    assert PageTitle1 in Ptitle1, PageName + " not able to open"
                except Exception:
                    Ptitle1="Funds - BIDS"
                    PageTitle1 = driver.title
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName +     " page not able to open")
                TestResultStatus.append("Fail")
            print()
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)
            #---------------------------------------------------------------------------------

            #--------------------Checking NAV Concentration--------------
            Text1 = "NAV Concentration"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[1]/div/div[1]/div/div/div").text
                assert Text1 in Element1, Text1 + " section is not present"
                TestResult.append(
                    Text1 + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " section is not present")
                TestResultStatus.append("Fail")

            inside="NAV Concentration"
            # --------------------Checking NAV Concentration content--------------
            Text1 = "NAV Concentration Grouping"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[1]/div/div[2]/div/div/div[1]/div/div/div[1]/span").text
                assert Text1 in Element1, Text1 + " dropdown inside "+inside+" section is not present"
                TestResult.append(
                    Text1 + " dropdown inside "+inside+" section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " dropdown inside "+inside+" section is not present")
                TestResultStatus.append("Fail")

            # --------------------Checking Pie Chart content--------------
            Text1 = "Pie Chart"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[1]/div/div[2]/div/div/div[2]/div/div/div[2]/div/div[1]/div").is_displayed()
                assert Element1==True, Text1 + " dropdown inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------Checking NAV Concentration--------------
            Text1 = "Distributions"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div[1]/div/div/div").text
                assert Text1 in Element1, Text1 + " section is not present"
                TestResult.append(
                    Text1 + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " section is not present")
                TestResultStatus.append("Fail")

            #---------------------------------------------------------------------------------
            inside = "Distributions"
            # --------------------Investment Name search box--------------
            Text1 = "Investment Name"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div[2]/div/div/div[1]/div[2]/div/div[1]/div/div[1]/label").text
                assert Text1 in Element1, Text1 + " search box inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " search box inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " search box inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------Toggle dropdown--------------
            Text1 = "Toggle"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div[2]/div/div/div[1]/div[2]/div/div[2]/div/div[1]/span").text
                assert Text1 in Element1, Text1 + " dropdown inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------Period dropdown--------------
            Text1 = "Period"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div[2]/div/div/div[1]/div[2]/div/div[3]/div/div[1]/span").text
                assert Text1 in Element1, Text1 + " dropdown inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------Export to Excel button--------------
            Text1 = "EXPORT TO EXCEL"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div[2]/div/div/div[1]/div[2]/div/div[4]/div/div/button").text
                assert Text1 in Element1, Text1 + " button inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " button inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " button inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------Investment Name column--------------
            Text1 = "Investment Name"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[1]/table/thead/tr[1]/th[1]/div").text
                assert Text1 in Element1, Text1 + " column inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " column inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " column inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------Checking List of Investments--------------
            Text1 = "List of Investments"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div/div/div[1]/div/div/div").text
                assert Text1 in Element1, Text1 + " section is not present"
                TestResult.append(
                    Text1 + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " section is not present")
                TestResultStatus.append("Fail")

            inside = "List of Investments"
            # --------------------Investment Name search box--------------
            Text1 = "Investment Name"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/div/div[1]/label").text
                assert Text1 in Element1, Text1 + " search box inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " search box inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " search box inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------BEN NAV Shown dropdown--------------
            Text1 = "BEN NAV Shown"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div/div/div/div/div[2]/div/div[2]/div/div[1]/span").text
                assert Text1 in Element1, Text1 + " dropdown inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------Asset Class dropdown--------------
            Text1 = "Asset Class"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div/div/div/div/div[2]/div/div[3]/div/div[1]/span").text
                assert Text1 in Element1, Text1 + " dropdown inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------Sector dropdown--------------
            Text1 = "Sector"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div/div/div/div/div[2]/div/div[4]/div/div[1]/span").text
                assert Text1 in Element1, Text1 + " dropdown inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " dropdown inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------Export to Excel button--------------
            Text1 = "EXPORT TO EXCEL"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div/div/div/div/div[2]/div/div[5]/div/div/button").text
                assert Text1 in Element1, Text1 + " button inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " button inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " button inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # --------------------CREATE NEW INVESTMENT button--------------
            Text1 = "CREATE NEW INVESTMENT"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div/div/div/div/div[2]/div/div[7]/div/div/button").text
                assert Text1 in Element1, Text1 + " button inside " + inside + " section is not present"
                TestResult.append(
                    Text1 + " button inside " + inside + " section is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(
                    Text1 + " button inside " + inside + " section is not present")
                TestResultStatus.append("Fail")

            # ---------------loop for Columns in table for List of Investments----------
            ItemList = ["ID", "Investment Name","Instrument Type","Number of Funds Held In","Ben NAV","Ben Asset Class","Ben Asset Class - Level1","Ben Asset Class - Level2","Sector","Industry Group","Country"
                        ,"Fund Acquired Date","Forecasted Liquidation Date","Investment RVPI for Funds","Operational Currency","Coverage"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div[2]/div/div/div[2]/table/thead/tr[1]/th[" + str(
                            ii + 1) + "]/div").text
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under " + inside + " table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                print("ItemPresent list is not empty")
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under " + inside + " table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                print("ItemNotPresent list is not empty")
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under " + inside + " table\n" + ListD)
                TestResultStatus.append("Fail")


        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start),2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


