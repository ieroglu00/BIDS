from builtins import print
from datetime import datetime, timedelta,date
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
import imaplib
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyodbc
from selenium.webdriver.chrome.options import Options

#----------------SalesForce Username and Password IDs-------------
@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("username").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global FundNameList
  global FundNameListAfterRemove
  global ct
  global Exe
  global D1
  global D2
  global d1
  global d2
  global DollarDate
  global FundToOpen
  global TotalFundsLengh
  global FieldDataFromSF
  global FieldDataSF
  global FoundDataBIDS
  global FoundDataSF

  TestName = "test_DealLog_SFBIDSPhase1"
  description = "This test scenario is to verify integration between Sales Force and BIDS application by creating an opportunity"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_DealLog(SF-BIDS)"
  Exe="Yes"
  Directory = 'test_DealLog(SF-BIDS)/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  FundNameList=[]
  FundNameListAfterRemove=[]
  FieldDataFromSF = {}
  FieldDataSF = {}
  FoundDataBIDS = {}
  FoundDataSF = {}

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      #-----------Disabling access popup from Chrome------------------
      option = Options()
      option.add_argument("--disable-infobars")
      option.add_argument("start-maximized")
      option.add_argument("--disable-extensions")
      option.add_experimental_option("prefs", {"profile.default_content_setting_values.notifications": 2})

      driver=webdriver.Chrome(chrome_options=option, executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()

      ct = datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.now().strftime("%d %B %Y %I %M%p")

      today = date.today()
      D1=today.strftime("%Y-%m-%d")
      d1=D1
      DollarDate=datetime.strptime(d1, '%Y-%m-%d')
      DollarDate="$"+DollarDate.date().__str__()+"$"
      d1 = datetime.strptime(D1, "%Y-%m-%d")

  yield
  if Exe == "Yes":
      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_DealLog_SFBIDSPhase1(test_setup):
    if Exe == "Yes":
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400

        ExcelFileName = "FieldData"
        loc = (path + 'Reference Data/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.get_sheet_by_name("DealLogPhase1")

        VerStatus=sheet.cell(row=1, column=6).value
        print(VerStatus)
        if VerStatus=="Verified":
            print("Verified")
        elif VerStatus=="Unverified":
            print("Unverified")
            OppName=sheet.cell(row=3, column=3).value

        # ----------------------Reading Field data from reference sheet-----------------
        for fielddata in range(1, 50):
            Value1 = sheet.cell(row=fielddata, column=2).value
            try:
                key1 = sheet.cell(row=fielddata, column=1).value
                FieldDataSF[key1] = Value1
            except Exception:
                pass
        print(FieldDataSF)

        try:
            if VerStatus=="Verified":

                # ------------------------Entering Blank values in Ref Excel------------------------
                sheet.cell(row=12, column=3).value = "Blank"
                sheet.cell(row=13, column=3).value = "Blank"
                sheet.cell(row=14, column=3).value = "Blank"
                sheet.cell(row=15, column=3).value = "Blank"
                sheet.cell(row=24, column=3).value = "Blank"
                wb.save(loc)

                # Loader for Sales Force
                LOADING_ELEMENT_XPATH = "//div[@class='slds-spinner_container slds-grid']"

                # ----------------------Now Navigating to salesforce Application----------------------------
                TestResult.append("====================Navigating to salesforce Application==============================")
                TestResultStatus.append("Pass")
                try:
                    driver.get("https://beneficient--int.my.salesforce.com/")
                    enter_username("neeraj.kumar@bitsinglass.com.int")
                    enter_password("Crochet@786")

                    driver.find_element_by_id("Login").click()
                    time.sleep(10)
                    TestResult.append(
                        " Username and Password entered in Sales Force site successfully")
                    TestResultStatus.append("Pass")
                except Exception:
                    PageLoadError = driver.find_element_by_xpath("//span[@jsselect='heading']").text
                    print(PageLoadError)
                    TestResult.append(
                        " Sales Force site is not able to load. Below error found\n" + PageLoadError)
                    TestResultStatus.append("Fail")
                    driver.close()

                try:
                    LoginError = driver.find_element_by_xpath("//div[@role='alert']").text
                    print(LoginError)
                    TestResult.append(
                        " Login attempt denied by Sales Force site. Below error found\n" + LoginError)
                    TestResultStatus.append("Fail")
                    driver.close()
                except Exception:
                    pass

                # ------------------------Clearing old SF data in Ref Excel------------------------
                for fielddata in range(1, 50):
                    sheet.cell(row=fielddata, column=3).value = None
                    sheet.cell(row=fielddata, column=5).value = None

                #------------------------Get verification code from Gmail---------------------------------
                host = 'imap.gmail.com'
                username = 'neeraj.kumar@bitsinglass.com'
                password = 'BitsMoh@2840828'

                # -------------Function to get email content part i.e its body part
                def get_body(msg):
                    if msg.is_multipart():
                        return get_body(msg.get_payload(0))
                    else:
                        return msg.get_payload(None, True)

                # -----------Function to search for a key value pair
                def search(key, value, con):
                    result, data = con.search(None, key, '"{}"'.format(value))
                    return data

                # ---------------Function to get the list of emails under this label
                def get_emails(result_bytes):
                    msgs = []  # all the email data are pushed inside an array
                    for num in result_bytes[0].split():
                        typ, data = con.fetch(num, '(RFC822)')
                        msgs.append(data)
                    return msgs

                con = imaplib.IMAP4_SSL(host)
                con.login(username, password)
                con.select('Inbox')

                # --------------fetching emails from a user
                msgs = get_emails(search('FROM', 'noreply@salesforce.com', con))
                Code=""
                for msg in msgs[::-1]:
                    for sent in msg:
                        if Code!="":
                            break
                        else:
                            if type(sent) is tuple:
                                content = str(sent[1], 'utf-8')
                                data = str(content)
                                try:
                                    indexstart = data.find("ltr")
                                    data2 = data[indexstart + 5: len(data)]
                                    indexend = data2.find("</div>")
                                    indx = data2.find('Verification Code:')
                                    Code = data2[indx + 19] + data2[indx + 20] + data2[indx + 21] + data2[indx + 22] + data2[
                                        indx + 23] + data2[indx + 24] + data2[indx + 25] + data2[indx + 26]
                                    print(Code)
                                    break
                                except UnicodeEncodeError as e:
                                    pass

                #------------Waiting for Verification code email-----------------
                if Code=="":
                    time.sleep(7)
                    print("Waiting for Verification Code ")

                # -----------------To Capture No Verification Code sent error from Sales Force-------------------------
                try:
                    bool1 = driver.find_element_by_id("save").is_displayed()
                except Exception:
                    try:
                        bool1 = driver.find_element_by_xpath(
                            "//div/h2[@class='mb12']").is_displayed()
                        if bool1 == True:
                            ErrorFound = driver.find_element_by_xpath(
                                "//div/h2[@class='mb12']").text
                            print(ErrorFound)
                            ErrorFound2 = driver.find_element_by_xpath(
                                "//div[@id='content']/form/p").text
                            print(ErrorFound2)
                            TestResult.append(" Verification code is not able to send from Sales Force due to below\n" + ErrorFound+ErrorFound2)
                            TestResultStatus.append("Fail")
                            driver.close()
                    except Exception:
                        pass

                # -----------------Login in Sales Force-------------------------
                time.sleep(2)
                driver.find_element_by_id('emc').send_keys(Code)
                TestResult.append("Verification code entered in Sales Force")
                TestResultStatus.append("Pass")
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                # -------------------Clicking on Accounts Tab in Top Menu------------------------
                try:
                    driver.find_element_by_xpath("//a[@title='Accounts']/parent::*").click()
                    TestResult.append("Clicked on Accounts Tab in Sales Force")
                    TestResultStatus.append("Pass")
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                except Exception:
                    PageLoadError = driver.find_element_by_xpath("//span[@jsselect='heading']").text
                    print(PageLoadError)
                    TestResult.append(
                        " Sales Force site is not able to load. Below error found\n" + PageLoadError)
                    TestResultStatus.append("Fail")
                    driver.close()

                # -------------------Checking Accounts present--------------------------
                try:
                    time.sleep(2)
                    AccName=driver.find_element_by_xpath("//span[text()='Account Name']/parent::a/parent::div/parent::th/parent::tr/parent::thead/parent::table/tbody/tr[1]/th/span/a").get_attribute('title')
                    print("AccName found "+AccName)
                    AccPhone = driver.find_element_by_xpath(
                        "//span[text()='Account Name']/parent::a/parent::div/parent::th/parent::tr/parent::thead/parent::table/tbody/tr[1]/td[2]/span/span").text
                    print("AccPhone found " + AccPhone)
                    if AccPhone=="":
                        print("AccPhone found blank")
                        driver.find_element_by_xpath("//span[text()='Account Name']/parent::a/parent::div/parent::th/parent::tr/parent::thead/parent::table/tbody/tr[1]/td[4]/span/div").click()
                        time.sleep(2)
                        ActionChains(driver).key_down(Keys.DOWN).perform()
                        time.sleep(1)
                        ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                        time.sleep(3)
                        for scrolldown in range(1, 10):
                            time.sleep(1)
                            print("scrolldown " + str(scrolldown))
                            try:
                                driver.find_element_by_xpath(
                                    "//h2[text()='Edit Person Account']/parent::article/div[3]/div/div[1]/div/div/div[5]/div[2]/div/div/div/input").send_keys("8877665544")
                                break
                            except Exception:
                                print("Inside Excep")
                                ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                                print("Page Down")
                        driver.find_element_by_xpath("//h2[text()='Edit Person Account']/parent::article/parent::div/parent::div/parent::div/div[2]/div/div/div[2]/button[3]").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                except Exception:
                    driver.find_element_by_xpath("//div[text()='New']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    driver.find_element_by_xpath("//span[text()='Next']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    driver.find_element_by_xpath("//span[text()='First Name']/parent::label/parent::div/input").send_keys("Test")
                    driver.find_element_by_xpath(
                        "//span[text()='Middle Name']/parent::label/parent::div/input").send_keys("account")
                    driver.find_element_by_xpath(
                        "//span[text()='Last Name']/parent::label/parent::div/input").send_keys("name")

                    driver.find_element_by_xpath("//span[text()='Ben Region']/parent::span/parent::div/div").click()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(2)
                    driver.find_element_by_xpath("//span[text()='Industry Detail']/parent::span/parent::div/div").click()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(2)
                    for scrolldown in range(1, 10):
                        time.sleep(1)
                        print("scrolldown " + str(scrolldown))
                        try:
                            driver.find_element_by_xpath(
                                "//h2[text()='Edit Person Account']/parent::article/div[3]/div/div[1]/div/div/div[5]/div[2]/div/div/div/input").send_keys(
                                "8877665544")
                            break
                        except Exception:
                            print("Inside Excep")
                            ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                            print("Page Down")
                    time.sleep(2)
                    driver.find_element_by_xpath("//h2[text()='New Account: Person Account']/parent::article/parent::div/parent::div/parent::div/div[2]/div/div/div[2]/button[3]").click()
                    TestResult.append("Added required account name details in Sales Force")
                    TestResultStatus.append("Pass")

                #-------------------Clicking on Opportunity Tab in Top Menu------------------------
                try:
                    driver.find_element_by_xpath("//a[@title='Opportunities']/parent::*").click()
                    TestResult.append("Clicked on Opportunity Tab in Sales Force")
                    TestResultStatus.append("Pass")
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                except Exception:
                    PageLoadError = driver.find_element_by_xpath("//span[@jsselect='heading']").text
                    print(PageLoadError)
                    TestResult.append(
                        " Sales Force site is not able to load. Below error found\n" + PageLoadError)
                    TestResultStatus.append("Fail")
                    driver.close()

                # -------------------Clikcing on New--------------------------
                driver.find_element_by_xpath("//a[@title='New']").click()
                TestResult.append("Clicked on New Opportunity in Sales Force")
                TestResultStatus.append("Pass")
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                # -------------------Clikcing on Next--------------------------
                driver.find_element_by_xpath("//span[text()='Next']").click()
                TestResult.append("Clicked on Next button to create Opportunity in Sales Force")
                TestResultStatus.append("Pass")
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                # ------------Filling Opportunity details---------------------------

                today = datetime.now()
                d = today.strftime("%b %d, %Y")

                search_key = "Opportunity Created Date"
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = d
                wb.save(loc)

                # --------Opportunity Name-----------
                FieldName = "Opportunity Name"
                abc = datetime.now().strftime('%d%h%I%M')
                OppName = abc
                try:
                    driver.find_element_by_xpath(
                        "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[3]/div[1]/div/div/div/input").send_keys(
                        OppName)
                except Exception:
                    time.sleep(4)
                    driver.find_element_by_xpath(
                        "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[3]/div[1]/div/div/div/input").send_keys(
                        OppName)
                ProjectName = FieldDataSF.get(FieldName)
                TestResult.append("[ " + OppName + " ] Opportunity Name entered in Sales Force")
                TestResultStatus.append("Pass")

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = OppName
                wb.save(loc)

                # --------Opportunity Close Date-----------
                try:
                    FieldName = "Close Date"
                    print(FieldName)
                    print(FieldDataSF.get(FieldName))

                    Duration = int(FieldDataSF.get(FieldName))
                    today = datetime.now()
                    NewDate = today + timedelta(days=Duration)
                    NewDate = NewDate.strftime('%m/%d/%Y')
                    if NewDate[0] == "0":
                        Item = ''.join([NewDate[i] for i in range(len(NewDate)) if i != 0])
                except Exception as ed:
                    print(ed)
                    pass
                print(Item)
                driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[4]/div[2]/div/div/div/div/input").send_keys(
                    Item)
                time.sleep(2)
                TestResult.append("Opportunity Close Date entered in Sales Force")
                TestResultStatus.append("Pass")

                today = datetime.now()
                NewDate = today + timedelta(days=Duration)
                NewDate = NewDate.strftime("%b %d, %Y")

                print(FieldName)
                search_key = FieldName
                print(search_key)
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                print(res)
                sheet.cell(row=res, column=3).value = NewDate
                wb.save(loc)

                # --------Opportunity Type-----------
                FieldName = "Opportunity Type"
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[8]/div[1]/div/div/div/div/div/div/div").click()
                        break
                    except Exception:
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        pass
                time.sleep(2)
                for ii3 in range(1, int(FieldDataSF.get(FieldName)) + 1):
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                value1 = driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[8]/div[1]/div/div/div/div/div/div/div/a").text
                print(value1)
                TestResult.append(FieldName + " entered in Sales Force")
                TestResultStatus.append("Pass")

                ValueToStore = value1
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ValueToStore
                wb.save(loc)

                # -------- Process Type-----------
                FieldName = "Process Type"
                ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[9]/div[1]/div/div/div/div/div/div/div").click()
                        break
                    except Exception:
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        pass
                time.sleep(2)
                for ii3 in range(1, int(FieldDataSF.get(FieldName)) + 1):
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                value1 = driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[9]/div[1]/div/div/div/div/div/div/div/a").text
                print(value1)
                TestResult.append(FieldName + " entered in Sales Force")
                TestResultStatus.append("Pass")
                ValueToStore = value1
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ValueToStore
                wb.save(loc)

                if value1 == "Auction - 2 Stage":
                    # --------IOI Due Date-----------
                    try:
                        FieldName = "IOI Due Date"
                        print(FieldName)
                        print(FieldDataSF.get(FieldName))

                        Duration = int(FieldDataSF.get(FieldName))
                        today = datetime.now()
                        NewDate = today + timedelta(days=Duration)
                        NewDate = NewDate.strftime('%m/%d/%Y')
                        if NewDate[0] == "0":
                            Item = ''.join([NewDate[i] for i in range(len(NewDate)) if i != 0])
                    except Exception as ed:
                        pass
                    print(Item)
                    driver.find_element_by_xpath(
                        "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[6]/div[2]/div/div/div/div/input").send_keys(
                        Item)
                    time.sleep(2)
                    TestResult.append("Opportunity IOI Due Date entered in Sales Force")
                    TestResultStatus.append("Pass")

                    today = datetime.now()
                    NewDate = today + timedelta(days=Duration)
                    NewDate = NewDate.strftime("%b %d, %Y")

                    print(FieldName)
                    search_key = FieldName
                    print(search_key)
                    res = list(FieldDataSF.keys()).index(search_key)
                    res = res + 1
                    print(res)
                    sheet.cell(row=res, column=3).value = NewDate
                    wb.save(loc)

                    # --------Final Bid Due Date-----------
                    try:
                        FieldName = "Final Bid Due Date"
                        print(FieldName)
                        print(FieldDataSF.get(FieldName))

                        Duration = int(FieldDataSF.get(FieldName))
                        today = datetime.now()
                        NewDate = today + timedelta(days=Duration)
                        NewDate = NewDate.strftime('%m/%d/%Y')
                        if NewDate[0] == "0":
                            Item = ''.join([NewDate[i] for i in range(len(NewDate)) if i != 0])
                    except Exception as ed:
                        pass
                    print(Item)
                    driver.find_element_by_xpath(
                        "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[7]/div[2]/div/div/div/div/input").send_keys(
                        Item)
                    time.sleep(2)
                    TestResult.append("Opportunity Final Bid Due Date entered in Sales Force")
                    TestResultStatus.append("Pass")

                    today = datetime.now()
                    NewDate = today + timedelta(days=Duration)
                    NewDate = NewDate.strftime("%b %d, %Y")

                    print(FieldName)
                    search_key = FieldName
                    print(search_key)
                    res = list(FieldDataSF.keys()).index(search_key)
                    res = res + 1
                    print(res)
                    sheet.cell(row=res, column=3).value = NewDate
                    wb.save(loc)

                # -------- NDA Status-----------
                FieldName = "NDA Status"
                print(FieldName)
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[6]/div[1]/div/div/div/div/div/div/div").click()
                        break
                    except Exception:
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        pass
                time.sleep(2)
                for ii3 in range(1, int(FieldDataSF.get(FieldName)) + 1):
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                value1 = driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[6]/div[1]/div/div/div/div/div/div/div/a").text
                print(value1)
                TestResult.append(FieldName + " entered in Sales Force")
                TestResultStatus.append("Pass")
                ValueToStore = value1
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ValueToStore
                wb.save(loc)

                # --------NDA Executed Date-----------
                try:
                    FieldName = "NDA Executed"
                    print(FieldName)
                    print(FieldDataSF.get(FieldName))

                    Duration = int(FieldDataSF.get(FieldName))
                    today = datetime.now()
                    NewDate = today + timedelta(days=Duration)
                    NewDate = NewDate.strftime('%m/%d/%Y')
                    if NewDate[0] == "0":
                        Item = ''.join([NewDate[i] for i in range(len(NewDate)) if i != 0])
                except Exception as ed:
                    pass
                print(Item)
                driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[7]/div[1]/div/div/div/div/input").send_keys(
                    Item)
                time.sleep(2)
                TestResult.append("NDA Executed Date entered in Sales Force")
                TestResultStatus.append("Pass")

                today = datetime.now()
                NewDate = today + timedelta(days=Duration)
                NewDate = NewDate.strftime("%b %d, %Y")

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = NewDate
                wb.save(loc)

                # --------Opportunity Stage-----------
                FieldName = "Stage SF Status"
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[9]/div[2]/div/div/div/div/div/div/div").click()
                        break
                    except Exception:
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        pass
                time.sleep(2)
                x = FieldDataSF.get(FieldName).split(",")
                print(x[0])

                for ii3 in range(1, int(x[0]) + 1):
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                value1 = driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[9]/div[2]/div/div/div/div/div/div/div/a").text
                print(value1)
                StageCheck = value1
                TestResult.append(FieldName + " selected in Sales Force")
                TestResultStatus.append("Pass")

                ValueToStore = value1
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ValueToStore
                wb.save(loc)

                # --------for Sub Stage--------
                ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[10]/div[2]/div/div/div/div/div/div/div").click()
                        break
                    except Exception:
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        pass
                time.sleep(2)
                for ii3 in range(1, int(x[1]) + 1):
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                value1 = driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[10]/div[2]/div/div/div/div/div/div/div/a").text
                print(value1)

                # # --------Stage Lost (if Closed Lost)-----------
                # if StageCheck == "Closed Lost":
                #     FieldName = "Stage Lost (If Closed Lost)"
                #     TestResult.append(FieldName + " selected in Sales Force")
                #     TestResultStatus.append("Pass")
                #
                #     ValueToStore = StageCheck
                #     search_key = FieldName
                #     res = list(FieldDataSF.keys()).index(search_key)
                #     res = res + 1
                #     sheet.cell(row=res, column=3).value = ValueToStore
                #     wb.save(loc)

                # --------Reason Lost (If Closed Lost)-----------
                if StageCheck == "Closed Lost":
                    FieldName = "Reason Lost (If Closed Lost)"
                    TestResult.append(FieldName + " selected in Sales Force")
                    TestResultStatus.append("Pass")

                    ValueToStore = StageCheck
                    search_key = FieldName
                    res = list(FieldDataSF.keys()).index(search_key)
                    res = res + 1
                    sheet.cell(row=res, column=3).value = value1
                    wb.save(loc)

                # --------Liquidity Opportunity-----------
                FieldName = "Liquidity Opportunity"
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[10]/div[1]/div/div/div/div/div/div/div").click()
                        break
                    except Exception:
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        pass
                time.sleep(2)
                for ii3 in range(1, int(FieldDataSF.get(FieldName)) + 1):
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                value1 = driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[10]/div[1]/div/div/div/div/div/div/div/a").text
                print(value1)
                TestResult.append(FieldName + " selected in Sales Force")
                TestResultStatus.append("Pass")

                ValueToStore = value1
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ValueToStore
                wb.save(loc)

                # --------Winning Bid-----------
                FieldName = "Winning Bid"
                print(FieldName)
                print(FieldDataSF.get(FieldName))
                driver.find_element_by_xpath(
                    "//span[text()='Winning Bid']/parent::label/parent::div/input").send_keys(
                    FieldDataSF.get(FieldName))
                time.sleep(2)
                TestResult.append(FieldName + " entered in Sales Force")
                TestResultStatus.append("Pass")

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                a_string = str(FieldDataSF.get(FieldName)).strip("0")
                sheet.cell(row=res, column=3).value = a_string + ".0%"
                wb.save(loc)

                # --------Opportunity Lead and Referral Source-----------
                FieldName = "Lead and Referral Source"
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[12]/div[2]/div/div/div/div/div/div/div").click()
                        break
                    except Exception:
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        pass
                time.sleep(2)
                for ii3 in range(1, int(FieldDataSF.get(FieldName)) + 1):
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                value1 = driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[12]/div[2]/div/div/div/div/div/div/div/a").text
                print(value1)
                TestResult.append(FieldName + " selected in Sales Force")
                TestResultStatus.append("Pass")

                ValueToStore = value1
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ValueToStore
                wb.save(loc)

                # --------Account Name-----------
                FieldName = "Financial Account Name"
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[13]/div[1]/div/div/div/div/div/div[1]/div").click()
                        break
                    except Exception:
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        pass
                time.sleep(2)
                for ii3 in range(1, int(FieldDataSF.get(FieldName)) + 1):
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                try:
                    value1 = driver.find_element_by_xpath(
                        "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[13]/div[1]/div/div/div/div/div/div[2]/div/ul/li[1]/a/span[2]").text
                    print("aa " + value1)
                except Exception:
                    time.sleep(2)
                    FinActBool = driver.find_element_by_xpath("//h2[text()='New Financial Account']").is_displayed()
                    print(FinActBool)
                    driver.find_element_by_xpath("//span[text()='Next']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    driver.find_element_by_xpath(
                        "//span[text()='Financial Account Name']/parent::label/parent::div/input").send_keys(
                        "Test Fin Acc")
                    driver.find_element_by_xpath(
                        "//span[text()='Primary Owner']/parent::label/parent::div/div").click()
                    time.sleep(2)
                    # ActionChains(driver).key_down(Keys.DOWN).perform()
                    # time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()

                    try:
                        value1 = driver.find_element_by_xpath(
                            "//span[text()='Primary Owner']/parent::label/parent::div/div/div/div[2]/div/ul/li[1]/a/span[2]").text
                        print("Primary Contct value found " + value1)
                    except Exception:
                        pass
                        # ------------Need to add code for Add Primary Owner------------------
                        # time.sleep(1)
                        # driver.find_element_by_xpath("//span[text()='Next']").click()

                    driver.find_element_by_xpath(
                        "//h2[text()='New Financial Account: General Account']/parent::div/parent::div/div[3]/div/button[3]").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(2)
                        value1 = driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[13]/div[1]/div/div/div/div/div/div[2]/div/ul/li[1]/a/span[2]").text
                        print("bb " + value1)
                    except Exception:
                        print("cccccc")
                        pass

                # driver.close()

                TestResult.append(FieldName + " selected in Sales Force")
                TestResultStatus.append("Pass")

                ValueToStore = value1
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ValueToStore
                wb.save(loc)

                # --------Financial Account-----------
                FieldName = "Financial Account"
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[12]/div[1]/div/div/div/div/div/div[1]/div").click()
                        break
                    except Exception:
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        pass
                time.sleep(2)
                for ii3 in range(1, int(FieldDataSF.get(FieldName)) + 1):
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                value1 = driver.find_element_by_xpath(
                    "//div[@class='modal-body scrollable slds-modal__content slds-p-around--medium']/div/div/div[1]/div/article/div[3]/div/div[1]/div/div/div[12]/div[1]/div/div/div/div/div/div[2]/div/ul/li/a/span[2]").text
                print(value1)
                TestResult.append(FieldName + " selected in Sales Force")
                TestResultStatus.append("Pass")

                ValueToStore = value1
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ValueToStore
                wb.save(loc)

                # --------Opportunity Description-----------
                FieldName = "Description"
                print(FieldName)
                print(FieldDataSF.get(FieldName))
                driver.find_element_by_xpath(
                    "//textarea[1]").send_keys(FieldDataSF.get(FieldName))
                time.sleep(2)
                TestResult.append(FieldName + " entered in Sales Force")
                TestResultStatus.append("Pass")

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = FieldDataSF.get(FieldName)
                wb.save(loc)

                # -------- Liquidity Request ID-----------
                FieldName = "Liquidity Request ID"
                print(FieldName)
                print(FieldDataSF.get(FieldName))
                driver.find_element_by_xpath(
                    "//span[text()='Liquidity Request ID']/parent::label/parent::div/input").send_keys(
                    FieldDataSF.get(FieldName))
                time.sleep(2)
                TestResult.append(FieldName + " entered in Sales Force")
                TestResultStatus.append("Pass")

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = FieldDataSF.get(FieldName)
                wb.save(loc)

                # --------Liquidity Request Date-----------
                try:
                    FieldName = "Liquidity Request Date"
                    print(FieldName)
                    print(FieldDataSF.get(FieldName))

                    Duration = int(FieldDataSF.get(FieldName))
                    today = datetime.now()
                    NewDate = today + timedelta(days=Duration)
                    NewDate = NewDate.strftime('%m/%d/%Y')
                    if NewDate[0] == "0":
                        Item = ''.join([NewDate[i] for i in range(len(NewDate)) if i != 0])
                except Exception as ed:
                    pass
                print(Item)
                driver.find_element_by_xpath(
                    "//span[text()='Liquidity Request Date']/parent::label/parent::div/div/input").send_keys(Item)
                time.sleep(2)
                TestResult.append("Liquidity Request Date entered in Sales Force")
                TestResultStatus.append("Pass")

                today = datetime.now()
                NewDate = today + timedelta(days=Duration)
                NewDate = NewDate.strftime("%b %d, %Y")

                print(FieldName)
                search_key = FieldName
                print(search_key)
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                print(res)
                sheet.cell(row=res, column=3).value = NewDate
                wb.save(loc)

                # --------------Math done by DWH Elapsed Days since Liquidity Request ----------------------
                search_key = "Math done by DWH Elapsed Days since Liquidity Request"
                print(search_key)
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                print(res)
                sheet.cell(row=res, column=3).value = "-" + str(Duration)
                wb.save(loc)

                # --------Opportunity Owner Id-----------
                FieldName = "Owner Id"

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = "Neeraj Kumar"
                wb.save(loc)

                # ------------Handling Blank fields--------------------------
                # --------Math done by DWH of Funds-----------
                FieldName = "Math done by DWH of Funds"

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = "Blank"
                wb.save(loc)

                # --------Math done by DWH of Investments-----------
                FieldName = "Math done by DWH of Investments"

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = "Blank"
                wb.save(loc)

                # ------------Submitting Opportunity details---------------------------
                driver.find_element_by_xpath(
                    "//div[@class='button-container-inner slds-float_right']/button[3]/span").click()
                time.sleep(10)
                TestResult.append("Opportunity details submitted in Sales Force")
                TestResultStatus.append("Pass")
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                #OppName="26Jan0248"
                driver.find_element_by_xpath("//a[@title='Opportunities']/parent::*").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                driver.find_element_by_xpath("//tbody/tr/th[1]/span/a[text()='"+OppName+"']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                s=driver.current_url
                i = s.index("/")
                s2 = s[i + 1:].strip()
                l = s2.split("/")
                OPPID=l[5]
                print(OPPID)

                # --------Opportunity ID-----------
                FieldName = "Opportunity ID"

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = OPPID
                wb.save(loc)

                # --------Contact Roles-----------
                FieldName = "Contact Roles"
                print(FieldName)
                print(FieldDataSF.get(FieldName))

                for scrolldown in range(1, 10):
                    time.sleep(1)
                    print("scrolldown "+str(scrolldown))
                    try:
                        abc=driver.find_element_by_xpath(
                            "//span[@title='Contact Roles']/parent::a/span[1]").text
                        print(abc)
                        break
                    except Exception:
                        print("Inside Excep")
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        print("Page Down")
                try:
                    ConNum=driver.find_element_by_xpath("//span[@title='Contact Roles']/parent::a/span[2]").get_attribute('title')
                    ConNum=ConNum.replace("(","")
                    ConNum = ConNum.replace(")", "")
                    print(ConNum)
                    if ConNum=="":
                        print("Found ConNum blank")
                        ConNum = 0
                        print(ConNum)
                except Exception:
                    ConNum=0
                    print(ConNum)

                button = driver.find_element_by_xpath("//span[@title='Contact Roles']")
                driver.execute_script("arguments[0].click();", button)
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                driver.find_element_by_xpath("//div[@title='Add Contact Roles']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                time.sleep(1)
                driver.find_element_by_xpath("//tbody/tr/td[2]/span/span/label/span[1]").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                time.sleep(1)
                ContName=driver.find_element_by_xpath("//tbody/tr[@class='selected']/th/span/a").text
                ContPhone = driver.find_element_by_xpath("//tbody/tr[@class='selected']/td[4]/span/span").text
                print(ContName)
                print(ContPhone)
                FieldName = "Contact Name"
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ContName
                wb.save(loc)

                FieldName = "Contact Phone"
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = ContPhone
                wb.save(loc)

                if int(ConNum) == 0:
                    driver.find_element_by_xpath("//button[@title='Next']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass

                    driver.find_element_by_xpath("//span[text()='Back']/parent::button/parent::div/parent::div/parent::div/div[2]/div/div[1]/lightning-grouped-combobox/div/div/lightning-base-combobox/div/div[1]").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(1)
                    driver.find_element_by_xpath("//button[@title='Save']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                else:
                    driver.find_element_by_xpath("//button[@title='Cancel']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass

                TestResult.append(FieldName + " added in Sales Force")
                TestResultStatus.append("Pass")

                #-------Industry-----------
                driver.find_element_by_xpath("//a[@title='Opportunities']/parent::*").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                driver.find_element_by_xpath("//a[text()='"+OppName+"']/preceding::td[last()]//following::td[1]/span/a").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                OPPIndustry=driver.find_element_by_xpath("//div/span[text()='Industry']/parent::div/parent::div/div[2]/span/span").text
                print(OPPIndustry)

                FieldName = "Industry"
                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = OPPIndustry
                wb.save(loc)

                # -------------------Holding Group Name-------------------------
                driver.find_element_by_xpath("//a[@title='Opportunities']/parent::*").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                driver.find_element_by_xpath("//tbody/tr/th[1]/span/a[text()='"+OppName+"']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                driver.find_element_by_xpath("//ul[@role='tablist']/li[1]/a[text()='Related']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                for scrolldown in range(1, 10):
                    time.sleep(2)
                    print("inside Related scrolldown "+str(scrolldown))
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='slds-media slds-media--center slds-has-flexi-truncate']/div[1]/div/div/h2/a/span[text()='Financial Holding Sets']").click()
                        break
                    except Exception:
                        print("Inside Excep")
                        ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                        print("Page Down")
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                FinancialHoldField = "Holding Group Name"
                FinancialHoldFieldValue = "Test Holding"
                Bool=False
                try:
                    Bool=driver.find_element_by_xpath("//tbody/tr/th/span/a[text()='"+FinancialHoldFieldValue+"']").is_displayed()
                    print(Bool)
                except Exception:
                    pass
                if Bool==False:
                    print("Checking boooll")
                    print(Bool)
                    driver.find_element_by_xpath("//div[@class='windowViewMode-normal oneContent active lafPageHost']/div/div/div[1]/div[1]/div[2]/ul/li/a/div").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    driver.find_element_by_xpath("//label[text()='"+FinancialHoldField+"']/parent::lightning-input/div[1]/input").send_keys(FinancialHoldFieldValue)
                    time.sleep(2)
                    driver.find_element_by_xpath("//button[text()='Save']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                time.sleep(2)
                driver.find_element_by_xpath("//tbody/tr/th[1]/span/a[text()='" + FinancialHoldFieldValue + "']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                #-----------------Included Financial Holdings----------------------
                time.sleep(2)
                driver.find_element_by_xpath("//ul[@role='tablist'][count(./li/*) = 3]/li/a[text()='Related']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                time.sleep(2)
                driver.find_element_by_xpath(
                    "//div[@class='slds-media slds-media--center slds-has-flexi-truncate']/div[1]/div/div/h2/a/span[text()='Included Financial Holdings']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                IncludedFinancialHoldingsValue = "Test Fin Holding"
                Bool=False
                try:
                    Bool = driver.find_element_by_xpath(
                        "//tbody/tr[1]/td/span/a[text()='" + IncludedFinancialHoldingsValue + "']").is_displayed()
                    print(Bool)
                except Exception:
                    Bool=False
                    pass
                time.sleep(2)
                if Bool == False:
                    driver.find_element_by_xpath(
                        "//div[@class='windowViewMode-normal oneContent active lafPageHost']/div/div/div[1]/div[1]/div[2]/ul/li/a/div").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    time.sleep(2)
                    driver.find_element_by_xpath(
                        "//input[@placeholder='Search Financial Holdings...']/parent::div").click()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(2)
                    try:
                        driver.find_element_by_xpath("//label/span[text()='Financial Holding Name']/parent::label/parent::div/input").send_keys(IncludedFinancialHoldingsValue)
                        time.sleep(1)
                        driver.find_element_by_xpath("//label/span[text()='Client Indicated Name']/parent::label/parent::div/input").send_keys("Test Cl Indicated")
                        time.sleep(1)
                        driver.find_element_by_xpath("//label/span[text()='Financial Account']/parent::label/parent::div/div/div/div[1]/div").click()
                        try:
                            driver.find_element_by_xpath("//label/span[text()='Financial Account']/parent::label/parent::div/div/div/div[1]/div/div/div[2]/div[2]/span").is_displayed()
                        except Exception:
                            time.sleep(4)
                            pass
                        time.sleep(2)
                        ActionChains(driver).key_down(Keys.DOWN).perform()
                        time.sleep(1)
                        ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                        time.sleep(1)
                        driver.find_element_by_xpath("//h2[text()='New Financial Holding']/parent::div/parent::div/div[3]/div/button[3]/span").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(
                                EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                    except Exception:
                        pass
                    time.sleep(2)
                    driver.find_element_by_xpath("//button[text()='Save']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass

                #----------------Financial Holding details---------------
                driver.find_element_by_xpath(
                    "//tbody/tr[1]/td/span/a[text()='" + IncludedFinancialHoldingsValue + "']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                FieldName = "Original Capital Commitment"
                FieldNameValue=driver.find_element_by_xpath("//lightning-helptext/parent::div/div/span[text()='"+FieldName+"']/parent::div/parent::div/div[2]/span/slot[1]/slot/lightning-formatted-text").text
                print("A "+FieldNameValue)

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = FieldNameValue
                wb.save(loc)

                FieldName = "Price"
                FieldNameValue = driver.find_element_by_xpath(
                    "//span[text()='"+FieldName+"']/parent::div/parent::div/div[2]/span/slot[1]/slot/lightning-formatted-text").text
                print("B "+FieldNameValue)

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = FieldNameValue
                wb.save(loc)

                FieldName = "Remaining Unfunded Commitment"
                FieldNameValue = driver.find_element_by_xpath(
                    "//div/slot/force-record-layout-row[5]/slot/force-record-layout-item/div/div/div/span[text()='"+FieldName+"']/parent::div/parent::div/div[2]/span/slot/slot/lightning-formatted-text").text
                print("C "+FieldNameValue)

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = FieldNameValue
                wb.save(loc)

                FieldName = "Follow on Investment"
                FieldNameValue = driver.find_element_by_xpath(
                    "//div/slot/force-record-layout-row[6]/slot/force-record-layout-item/div/div/div/span[text()='" + FieldName + "']/parent::div/parent::div/div[2]/span/slot/slot/lightning-formatted-text").text
                print("D " + FieldNameValue)

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = FieldNameValue
                wb.save(loc)

                # -------------Asset Class-----------------------
                driver.find_element_by_xpath(
                    "//span[text()='Symbol']/parent::div/parent::div/div[2]/span/slot/slot/force-lookup/div/force-hoverable-link/div/a/slot/slot/span").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                FieldName = "Asset Class"
                FieldNameValue = driver.find_element_by_xpath(
                    "//span[text()='Asset Categories']/parent::button/parent::h3/parent::div/div/div/slot/force-record-layout-row[1]/slot/force-record-layout-item[1]/div/div/div[2]/span/slot[1]/slot/lightning-formatted-text").text
                print("D " + FieldNameValue)

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = "*Test for Neeraj: "+FieldNameValue
                wb.save(loc)

                FieldName = "Asset Category"
                FieldNameValue = driver.find_element_by_xpath(
                    "//span[text()='Asset Categories']/parent::button/parent::h3/parent::div/div/div/slot/force-record-layout-row[2]/slot/force-record-layout-item[1]/div/div/div[2]/span/slot[1]/slot/lightning-formatted-text").text
                print("D " + FieldNameValue)

                search_key = FieldName
                res = list(FieldDataSF.keys()).index(search_key)
                res = res + 1
                sheet.cell(row=res, column=3).value = "*Test for Neeraj: "+FieldNameValue
                wb.save(loc)

                # #-------------Offer-----------------------
                # driver.back()
                # try:
                #     WebDriverWait(driver, SHORT_TIMEOUT
                #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                #
                #     WebDriverWait(driver, LONG_TIMEOUT
                #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                # except TimeoutException:
                #     pass
                #
                # driver.back()
                # try:
                #     WebDriverWait(driver, SHORT_TIMEOUT
                #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                #
                #     WebDriverWait(driver, LONG_TIMEOUT
                #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                # except TimeoutException:
                #     pass
                #
                # driver.back()
                # try:
                #     WebDriverWait(driver, SHORT_TIMEOUT
                #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                #
                #     WebDriverWait(driver, LONG_TIMEOUT
                #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                # except TimeoutException:
                #     pass
                #
                # driver.find_element_by_xpath("//span[@title='Offers']").click()
                # try:
                #     WebDriverWait(driver, SHORT_TIMEOUT
                #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                #
                #     WebDriverWait(driver, LONG_TIMEOUT
                #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                # except TimeoutException:
                #     pass
                #
                # IncludedFinancialHoldingsValue = "Test Offer"
                # Bool = False
                # try:
                #     Bool = driver.find_element_by_xpath(
                #         "//tbody/tr[1]/td/span/a[text()='" + IncludedFinancialHoldingsValue + "']").is_displayed()
                #     print(Bool)
                # except Exception:
                #     pass
                # time.sleep(2)
                # if Bool == False:
                #     driver.find_element_by_xpath(
                #         "//div[@class='windowViewMode-normal oneContent active lafPageHost']/div/div/div[1]/div[1]/div[2]/ul/li/a/div").click()
                #     try:
                #         WebDriverWait(driver, SHORT_TIMEOUT
                #                       ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                #
                #         WebDriverWait(driver, LONG_TIMEOUT
                #                       ).until_not(
                #             EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                #     except TimeoutException:
                #         pass
                #     driver.find_element_by_xpath(
                #         "//input[@placeholder='Search Financial Holdings...']/parent::div").click()
                #     time.sleep(2)
                #     ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                #     driver.find_element_by_xpath("//button[text()='Save']").click()
                #     try:
                #         WebDriverWait(driver, SHORT_TIMEOUT
                #                       ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                #
                #         WebDriverWait(driver, LONG_TIMEOUT
                #                       ).until_not(
                #             EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                #     except TimeoutException:
                #         pass

                #----------------------Now Navigating to BIDS Application----------------------------
                TestResult.append("====================Navigating to BIDS Application==============================")
                TestResultStatus.append("Pass")

                VerStatusCount=0

                #OppName="16Jan1049"
                ProjectName=OppName
                #Loader for BIDS
                LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
                #-------------------For Login in BIDS-------------------
                driver.get("https://beneficienttest.appiancloud.com/suite/")
                driver.find_element_by_id("un").send_keys("neeraj.kumar")
                driver.find_element_by_id("pw").send_keys("Crochet@7866")
                driver.find_element_by_xpath("//input[@type='submit']").click()
                TestResult.append("Navigated to BIDS application")
                TestResultStatus.append("Pass")


                #---------------------------Verify Transactions page-----------------------------
                PageName="Transactions"
                Ptitle1="Transaction Listing "
                driver.find_element_by_xpath("//*[@title='"+PageName+"']").click()
                #start = time.time()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                try:
                    time.sleep(2)
                    bool1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                    if bool1 == True:
                        ErrorFound1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                        print(ErrorFound1)
                        driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                        TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                        TestResultStatus.append("Fail")
                        sheet.cell(row=1, column=6).value = "Unverified"
                        wb.save(loc)
                        driver.close()
                        bool1 = False
                except Exception:
                    try:
                        time.sleep(2)
                        bool2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                        if bool2 == True:
                            ErrorFound2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                            print(ErrorFound2)
                            TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                            TestResultStatus.append("Fail")
                            sheet.cell(row=1, column=6).value = "Unverified"
                            wb.save(loc)
                            driver.close()
                            bool2 = False
                    except Exception:
                        pass
                    pass
                time.sleep(1)
                try:
                    PageTitle1 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[1]/div/div/div").text
                    print(PageTitle1)
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                    TestResult.append(PageName + " page opened successfully")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append(PageName +     " page not able to open")
                    TestResultStatus.append("Fail")
                    driver.close()
                #---------------------------------------------------------------------------------

                try:
                    print()
                    TotalItem=driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/span[2]").text
                    substr = "of"
                    x = TotalItem.split(substr)
                    string_name = x[0]
                    TotalItemAfterOf= x[1]
                    abc = ""
                    countspace = 0
                    for element in range(0, len(string_name)):
                        if string_name[(len(string_name) - 1) - element] == " ":
                            countspace = countspace + 1
                            if countspace == 2:
                                break
                        else:
                            abc = abc + string_name[(len(string_name) - 1) - element]
                    abc = abc[::-1]
                    TotalItemBeforeOf=abc
                    TotalItemAfterOf = re.sub('[^A-Za-z0-9]+', '', TotalItemAfterOf)
                    TotalItemBeforeOf = re.sub('[^A-Za-z0-9]+', '', TotalItemBeforeOf)
                    print("TotalItemAfterOf " + TotalItemAfterOf)
                    print("TotalItemBeforeOf " + TotalItemBeforeOf)

                    IterateNo = int(TotalItemAfterOf) / int(TotalItemBeforeOf)
                    if IterateNo.is_integer() == True:
                        IterateNo = IterateNo - 1
                        pass
                    else:
                        IterateNo = math.ceil(float(IterateNo))
                    #print("IterateNo is " + str(IterateNo))
                    #----------------Searching the Project from Sales Force--------------------
                    LoopExit=0
                    ProejctTOClick = ProjectName
                    PageName = ProejctTOClick
                    start = time.time()
                    WaitingCount=13
                    if IterateNo>=5 and IterateNo<7:
                        WaitingCount=11
                    elif IterateNo >= 7and IterateNo<9:
                        WaitingCount = 9
                    elif IterateNo >= 9:
                        WaitingCount = 5
                    for waiting in range(1,WaitingCount):
                        if LoopExit==0:
                            time.sleep(60)
                            driver.refresh()
                            try:
                                WebDriverWait(driver, SHORT_TIMEOUT
                                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                WebDriverWait(driver, LONG_TIMEOUT
                                              ).until_not(
                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            except TimeoutException:
                                pass

                            loopbreak=0
                            for ii5 in range(1, IterateNo+1):
                                if loopbreak==0:
                                    if ii5 >1:
                                        try:
                                            driver.find_element_by_xpath(
                                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                                            try:
                                                WebDriverWait(driver, SHORT_TIMEOUT
                                                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                                WebDriverWait(driver, LONG_TIMEOUT
                                                              ).until_not(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                            except TimeoutException:
                                                pass
                                        except Exception as q1:
                                            print(q1)
                                            pass

                                    RowsInv = driver.find_elements_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr")
                                    for ii3 in range(1, len(RowsInv)+1):
                                        ProjectNameText = driver.find_element_by_xpath(
                                            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                                                ii3) + "]/td[1]/div/p/a").text
                                        if ProjectNameText==ProejctTOClick:
                                            loopbreak=1
                                            driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr/td[1]/div/p/a[text()='"+ProjectNameText+"']").click()
                                            TestResult.append(PageName + " project clicked in BIDS application")
                                            TestResultStatus.append("Pass")
                                            VerStatusCount=1
                                            try:
                                                WebDriverWait(driver, SHORT_TIMEOUT
                                                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                                WebDriverWait(driver, LONG_TIMEOUT
                                                              ).until_not(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                            except TimeoutException:
                                                pass
                                            try:
                                                time.sleep(2)
                                                bool1 = driver.find_element_by_xpath(
                                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                                                if bool1 == True:
                                                    ErrorFound1 = driver.find_element_by_xpath(
                                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                                    print(ErrorFound1)
                                                    driver.find_element_by_xpath(
                                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                                                    TestResultStatus.append("Fail")
                                                    driver.close()
                                                    bool1 = False
                                            except Exception:
                                                try:
                                                    time.sleep(2)
                                                    bool2 = driver.find_element_by_xpath(
                                                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                                    if bool2 == True:
                                                        ErrorFound2 = driver.find_element_by_xpath(
                                                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                                        print(ErrorFound2)
                                                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                                        TestResultStatus.append("Fail")
                                                        driver.close()
                                                        bool2 = False
                                                except Exception:
                                                    pass
                                                pass
                                            LoopExit=1
                                            break
                                else:
                                    break
                        else:
                            break

                    if LoopExit==0:
                        stop = time.time()
                        RoundFloatString = round(float(stop - start),2)
                        seconds = int(RoundFloatString)
                        min, sec = divmod(seconds, 60)
                        TestResult.append(PageName + " project not able to find in BIDS application\nTotal run time to find the project in BIDS is "+str(min)+" mintues")
                        TestResultStatus.append("Fail")
                        sheet.cell(row=1, column=6).value="Unverified"
                        wb.save(loc)
                        driver.close()

                    # ------------------------clicking Transaction ID--------------------------------
                    PageName = "Transaction ID"
                    driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div/div[2]/div/div/table/tbody/tr/td[2]/div/p/a").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(2)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                        if bool1 == True:
                            ErrorFound1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                            print(ErrorFound1)
                            driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                            TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                            TestResultStatus.append("Fail")
                            sheet.cell(row=1, column=6).value="Unverified"
                            wb.save(loc)
                            driver.close()
                            bool1 = False
                    except Exception:
                        try:
                            time.sleep(2)
                            bool2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                            if bool2 == True:
                                ErrorFound2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                print(ErrorFound2)
                                TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                TestResultStatus.append("Fail")
                                sheet.cell(row=1, column=6).value = "Unverified"
                                wb.save(loc)
                                driver.close()
                                bool2 = False
                        except Exception:
                            pass
                        pass
                    time.sleep(1)
                    TestResult.append(PageName + " clicked in BIDS application")
                    TestResultStatus.append("Pass")

                    #-------------clicking Edit Key Transaction Details--------------------------------
                    driver.find_element_by_xpath("//p/a[text()='Edit Key Transaction Details']").click()
                    PageName="Edit Key Transaction Details"
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(1)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                        if bool1 == True:
                            ErrorFound1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                            print(ErrorFound1)
                            driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                            TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                            TestResultStatus.append("Fail")
                            sheet.cell(row=1, column=6).value = "Unverified"
                            wb.save(loc)
                            driver.close()
                            bool1 = False
                    except Exception:
                        try:
                            time.sleep(1)
                            bool2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                            if bool2 == True:
                                ErrorFound2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                print(ErrorFound2)
                                TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                TestResultStatus.append("Fail")
                                sheet.cell(row=1, column=6).value = "Unverified"
                                wb.save(loc)
                                driver.close()
                                bool2 = False
                        except Exception:
                            pass
                        pass
                    time.sleep(1)
                    TestResult.append(PageName + " clicked in BIDS application")
                    TestResultStatus.append("Pass")

                    #---------------Storing SF all values in a Dic with Reference to BIDS fields-------------------
                    for ff in range(1, 50):
                        try:
                            Value1 = sheet.cell(row=ff, column=3).value
                            key1 = sheet.cell(row=ff, column=4).value
                            if Value1 == "":
                                Value1 = "Blank"
                            FieldDataFromSF[key1] = Value1
                        except Exception:
                            pass

                    # -------------Fetching Key Transaction Details--------------------------------
                    Loop1Number = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div")
                    Loop2Number = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div")
                    Loop3Number = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[3]/div")

                    for loop1 in range(1,len(Loop1Number)+1):
                        foundLabel=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div["+str(loop1)+"]/div[1]/label").text
                        try:
                            foundValue=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div["+str(loop1)+"]/div[2]/div/input").get_attribute('value')
                        except:
                            foundValue=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div["+str(loop1)+"]/div[2]/div/textarea").text
                        if foundValue=="":
                            foundValue="Blank"
                        print("foundLabel: "+foundLabel)
                        print("foundValue: " + foundValue)
                        FoundDataBIDS[foundLabel] = foundValue
                        try:
                            search_key = foundLabel
                            res = list(FieldDataFromSF.keys()).index(search_key)
                            res = res + 1
                            sheet.cell(row=res, column=5).value = foundValue
                            wb.save(loc)
                        except Exception as ss1:
                            print(ss1)
                            pass

                    for loop2 in range(1,len(Loop2Number)+1):
                        foundLabel = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[" + str(loop2) + "]/div[1]/label").text
                        try:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[" + str(loop2) + "]/div[2]/div/input").get_attribute(
                                'value')
                        except:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[" + str(loop2) + "]/div[2]/div/textarea").text
                        if foundValue=="":
                            foundValue="Blank"
                        FoundDataBIDS[foundLabel] = foundValue
                        try:
                            search_key = foundLabel
                            res = list(FieldDataFromSF.keys()).index(search_key)
                            res = res + 1
                            sheet.cell(row=res, column=5).value = foundValue
                            wb.save(loc)
                        except Exception as ss2:
                            print(ss2)
                            pass

                    for loop3 in range(1,len(Loop3Number)+1):
                        foundLabel = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[3]/div[" + str(loop3) + "]/div[1]/label").text
                        try:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[3]/div[" + str(loop3) + "]/div[2]/div/input").get_attribute(
                                'value')
                        except:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[3]/div[" + str(loop3) + "]/div[2]/div/textarea").text
                        if foundValue=="":
                            foundValue="Blank"
                        FoundDataBIDS[foundLabel] = foundValue
                        try:
                            search_key = foundLabel
                            res = list(FieldDataFromSF.keys()).index(search_key)
                            res = res + 1
                            sheet.cell(row=res, column=5).value = foundValue
                            wb.save(loc)
                        except Exception as ss3:
                            print(ss3)
                            pass

                    print("FoundDataBIDS")
                    print(FoundDataBIDS)

                    print("FieldDataFromSF")
                    print(FieldDataFromSF)

                    TestResult.append("===================Now comparing SalesForce Data with BIDS=======================")
                    TestResultStatus.append("Pass")
                    for key in FieldDataFromSF:
                        print()
                        print("---------------------------------")
                        try:
                            print("FieldDataFromSF[key]  "+FieldDataFromSF[key])
                            print("FoundDataBIDS[key]  " + FoundDataBIDS[key])
                        except Exception:
                            pass
                        if (key in FoundDataBIDS and FieldDataFromSF[key] == FoundDataBIDS[key]):
                            print(key + " in Sales Force matched with BIDS data")
                            TestResult.append(key + " in Sales Force matched with BIDS data")
                            TestResultStatus.append("Pass")
                        elif (key in FoundDataBIDS and FieldDataFromSF[key] != FoundDataBIDS[key]):
                            print(key + " in Sales Force does not match with BIDS data")
                            TestResult.append(key + " in Sales Force does not match with BIDS data")
                            TestResultStatus.append("Fail")

                    sheet.cell(row=1, column=6).value = "Verified"
                    for Pname in range(2,100):
                        if sheet.cell(row=Pname, column=7).value == None:
                            sheet.cell(row=Pname, column=7).value=OppName
                            break

                    for Pname1 in range(2,100):
                        if sheet.cell(row=Pname1, column=9).value == None:
                            sheet.cell(row=Pname1, column=9).value=OppName
                            break

                    wb.save(loc)

                except Exception as err:
                    print(err)
                    sheet.cell(row=1, column=6).value = "Unverified"
                    wb.save(loc)
                    pass
            elif VerStatus == "Unverified":
                print("Yesss Unverified")
                print(OppName)

                TestResult.append("Old unverified Opportunity [ " + OppName + " ] picked from Sales Force for re-verification")
                TestResultStatus.append("Pass")
                # ----------------------Now Navigating to BIDS Application----------------------------
                TestResult.append("====================Navigating to BIDS Application==============================")
                TestResultStatus.append("Pass")

                ProjectName = OppName
                # Loader for BIDS
                LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
                # -------------------For Login in BIDS-------------------
                driver.get("https://beneficienttest.appiancloud.com/suite/")
                driver.find_element_by_id("un").send_keys("neeraj.kumar")
                driver.find_element_by_id("pw").send_keys("Crochet@7866")
                driver.find_element_by_xpath("//input[@type='submit']").click()
                TestResult.append("Navigated to BIDS application")
                TestResultStatus.append("Pass")

                # ---------------------------Verify Transactions page-----------------------------
                PageName = "Transactions"
                Ptitle1 = "Transaction Listing "
                driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
                # start = time.time()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                try:
                    time.sleep(2)
                    bool1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                    if bool1 == True:
                        ErrorFound1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                        print(ErrorFound1)
                        driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                        TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                        TestResultStatus.append("Fail")
                        sheet.cell(row=1, column=6).value = "Unverified"
                        wb.save(loc)
                        driver.close()
                        bool1 = False
                except Exception:
                    try:
                        time.sleep(2)
                        bool2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                        if bool2 == True:
                            ErrorFound2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                            print(ErrorFound2)
                            TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                            TestResultStatus.append("Fail")
                            sheet.cell(row=1, column=6).value = "Unverified"
                            wb.save(loc)
                            driver.close()
                            bool2 = False
                    except Exception:
                        pass
                    pass
                time.sleep(1)
                try:
                    PageTitle1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[1]/div/div/div").text
                    print(PageTitle1)
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                    TestResult.append(PageName + " page Opened successfully")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append(PageName + " page not able to open")
                    TestResultStatus.append("Fail")
                    driver.close()
                # ---------------------------------------------------------------------------------

                try:
                    print()
                    TotalItem = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/span[2]").text
                    substr = "of"
                    x = TotalItem.split(substr)
                    string_name = x[0]
                    TotalItemAfterOf = x[1]
                    abc = ""
                    countspace = 0
                    for element in range(0, len(string_name)):
                        if string_name[(len(string_name) - 1) - element] == " ":
                            countspace = countspace + 1
                            if countspace == 2:
                                break
                        else:
                            abc = abc + string_name[(len(string_name) - 1) - element]
                    abc = abc[::-1]
                    TotalItemBeforeOf = abc
                    TotalItemAfterOf = re.sub('[^A-Za-z0-9]+', '', TotalItemAfterOf)
                    TotalItemBeforeOf = re.sub('[^A-Za-z0-9]+', '', TotalItemBeforeOf)
                    print("TotalItemAfterOf " + TotalItemAfterOf)
                    print("TotalItemBeforeOf " + TotalItemBeforeOf)

                    IterateNo = int(TotalItemAfterOf) / int(TotalItemBeforeOf)
                    if IterateNo.is_integer() == True:
                        IterateNo = IterateNo - 1
                        pass
                    else:
                        IterateNo = math.ceil(float(IterateNo))
                    #print("IterateNo is " + str(IterateNo))
                    # ----------------Searching the Project from Sales Force--------------------
                    LoopExit = 0
                    ProejctTOClick = ProjectName
                    PageName = ProejctTOClick
                    start = time.time()
                    WaitingCount = 13
                    if IterateNo >= 5 and IterateNo < 7:
                        WaitingCount = 11
                    elif IterateNo >= 7 and IterateNo < 9:
                        WaitingCount = 9
                    elif IterateNo >= 9:
                        WaitingCount = 5
                    for waiting in range(1, WaitingCount):
                        if LoopExit == 0:
                            time.sleep(60)
                            driver.refresh()
                            try:
                                WebDriverWait(driver, SHORT_TIMEOUT
                                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                WebDriverWait(driver, LONG_TIMEOUT
                                              ).until_not(
                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            except TimeoutException:
                                pass

                            loopbreak = 0
                            for ii5 in range(1, IterateNo + 1):
                                if loopbreak == 0:
                                    if ii5 > 1:
                                        try:
                                            driver.find_element_by_xpath(
                                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                                            try:
                                                WebDriverWait(driver, SHORT_TIMEOUT
                                                              ).until(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                                WebDriverWait(driver, LONG_TIMEOUT
                                                              ).until_not(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                            except TimeoutException:
                                                pass
                                        except Exception as q1:
                                            print(q1)
                                            pass

                                    RowsInv = driver.find_elements_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr")
                                    for ii3 in range(1, len(RowsInv) + 1):
                                        ProjectNameText = driver.find_element_by_xpath(
                                            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                                                ii3) + "]/td[1]/div/p/a").text
                                        if ProjectNameText == ProejctTOClick:
                                            loopbreak = 1
                                            driver.find_element_by_xpath(
                                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr/td[1]/div/p/a[text()='" + ProjectNameText + "']").click()
                                            TestResult.append(PageName + " project clicked in BIDS application")
                                            TestResultStatus.append("Pass")
                                            VerStatusCount = 1
                                            try:
                                                WebDriverWait(driver, SHORT_TIMEOUT
                                                              ).until(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                                WebDriverWait(driver, LONG_TIMEOUT
                                                              ).until_not(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                            except TimeoutException:
                                                pass
                                            try:
                                                time.sleep(2)
                                                bool1 = driver.find_element_by_xpath(
                                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                                                if bool1 == True:
                                                    ErrorFound1 = driver.find_element_by_xpath(
                                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                                    print(ErrorFound1)
                                                    driver.find_element_by_xpath(
                                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                                                    TestResultStatus.append("Fail")
                                                    driver.close()
                                                    bool1 = False
                                            except Exception:
                                                try:
                                                    time.sleep(2)
                                                    bool2 = driver.find_element_by_xpath(
                                                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                                    if bool2 == True:
                                                        ErrorFound2 = driver.find_element_by_xpath(
                                                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                                        print(ErrorFound2)
                                                        TestResult.append(
                                                            PageName + " not able to open\n" + ErrorFound2)
                                                        TestResultStatus.append("Fail")
                                                        driver.close()
                                                        bool2 = False
                                                except Exception:
                                                    pass
                                                pass
                                            LoopExit = 1
                                            break
                                else:
                                    break
                        else:
                            break

                    if LoopExit == 0:
                        stop = time.time()
                        RoundFloatString = round(float(stop - start), 2)
                        seconds = int(RoundFloatString)
                        min, sec = divmod(seconds, 60)
                        TestResult.append(
                            PageName + " project not able to find in BIDS application\nTotal run time to find the project in BIDS is " + str(
                                min) + " mintues")
                        TestResultStatus.append("Fail")
                        sheet.cell(row=1, column=6).value = "Verified"
                        wb.save(loc)
                        driver.close()

                    # ------------------------clicking Transaction ID--------------------------------
                    PageName = "Transaction ID"
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div/div[2]/div/div/table/tbody/tr/td[2]/div/p/a").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(2)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                        if bool1 == True:
                            ErrorFound1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                            print(ErrorFound1)
                            driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                            TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                            TestResultStatus.append("Fail")
                            sheet.cell(row=1, column=6).value = "Unverified"
                            wb.save(loc)
                            driver.close()
                            bool1 = False
                    except Exception:
                        try:
                            time.sleep(2)
                            bool2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                            if bool2 == True:
                                ErrorFound2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                print(ErrorFound2)
                                TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                TestResultStatus.append("Fail")
                                sheet.cell(row=1, column=6).value = "Unverified"
                                wb.save(loc)
                                driver.close()
                                bool2 = False
                        except Exception:
                            pass
                        pass
                    time.sleep(1)
                    TestResult.append(PageName + " clicked in BIDS application")
                    TestResultStatus.append("Pass")

                    # -------------clicking Edit Key Transaction Details--------------------------------
                    driver.find_element_by_xpath("//p/a[text()='Edit Key Transaction Details']").click()
                    PageName = "Edit Key Transaction Details"
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(1)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                        if bool1 == True:
                            ErrorFound1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                            print(ErrorFound1)
                            driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                            TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                            TestResultStatus.append("Fail")
                            sheet.cell(row=1, column=6).value = "Unverified"
                            wb.save(loc)
                            driver.close()
                            bool1 = False
                    except Exception:
                        try:
                            time.sleep(1)
                            bool2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                            if bool2 == True:
                                ErrorFound2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                print(ErrorFound2)
                                TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                TestResultStatus.append("Fail")
                                sheet.cell(row=1, column=6).value = "Unverified"
                                wb.save(loc)
                                driver.close()
                                bool2 = False
                        except Exception:
                            pass
                        pass
                    time.sleep(1)
                    TestResult.append(PageName + " clicked in BIDS application")
                    TestResultStatus.append("Pass")

                    # ---------------Storing SF all values in a Dic with Reference to BIDS fields-------------------
                    for ff in range(1, 50):
                        try:
                            Value1 = sheet.cell(row=ff, column=3).value
                            key1 = sheet.cell(row=ff, column=4).value
                            if Value1 == "":
                                Value1 = "Blank"
                            FieldDataFromSF[key1] = Value1
                        except Exception:
                            pass

                    # -------------Fetching Key Transaction Details--------------------------------
                    Loop1Number = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div")
                    Loop2Number = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div")
                    Loop3Number = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[3]/div")

                    for loop1 in range(1, len(Loop1Number) + 1):
                        foundLabel = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div[" + str(
                                loop1) + "]/div[1]/label").text
                        try:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div[" + str(
                                    loop1) + "]/div[2]/div/input").get_attribute('value')
                        except:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div[" + str(
                                    loop1) + "]/div[2]/div/textarea").text
                        if foundValue == "":
                            foundValue = "Blank"
                        print("foundLabel: " + foundLabel)
                        print("foundValue: " + foundValue)
                        FoundDataBIDS[foundLabel] = foundValue
                        try:
                            search_key = foundLabel
                            res = list(FieldDataFromSF.keys()).index(search_key)
                            res = res + 1
                            sheet.cell(row=res, column=5).value = foundValue
                            wb.save(loc)
                        except Exception as ss1:
                            print(ss1)
                            pass

                    for loop2 in range(1, len(Loop2Number) + 1):
                        foundLabel = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[" + str(
                                loop2) + "]/div[1]/label").text
                        try:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[" + str(
                                    loop2) + "]/div[2]/div/input").get_attribute(
                                'value')
                        except:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[2]/div[" + str(
                                    loop2) + "]/div[2]/div/textarea").text
                        if foundValue == "":
                            foundValue = "Blank"
                        FoundDataBIDS[foundLabel] = foundValue
                        try:
                            search_key = foundLabel
                            res = list(FieldDataFromSF.keys()).index(search_key)
                            res = res + 1
                            sheet.cell(row=res, column=5).value = foundValue
                            wb.save(loc)
                        except Exception as ss2:
                            print(ss2)
                            pass

                    for loop3 in range(1, len(Loop3Number) + 1):
                        foundLabel = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[3]/div[" + str(
                                loop3) + "]/div[1]/label").text
                        try:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[3]/div[" + str(
                                    loop3) + "]/div[2]/div/input").get_attribute(
                                'value')
                        except:
                            foundValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[1]/div[2]/div/div/div[3]/div[" + str(
                                    loop3) + "]/div[2]/div/textarea").text
                        if foundValue == "":
                            foundValue = "Blank"
                        FoundDataBIDS[foundLabel] = foundValue
                        try:
                            search_key = foundLabel
                            res = list(FieldDataFromSF.keys()).index(search_key)
                            res = res + 1
                            sheet.cell(row=res, column=5).value = foundValue
                            wb.save(loc)
                        except Exception as ss3:
                            print(ss3)
                            pass

                    print("FoundDataBIDS")
                    print(FoundDataBIDS)

                    print("FieldDataFromSF")
                    print(FieldDataFromSF)

                    TestResult.append(
                        "===================Now comparing SalesForce Data with BIDS=======================")
                    TestResultStatus.append("Pass")
                    for key in FieldDataFromSF:
                        print()
                        print("---------------------------------")
                        try:
                            print("FieldDataFromSF[key]  " + FieldDataFromSF[key])
                            print("FoundDataBIDS[key]  " + FoundDataBIDS[key])
                        except Exception:
                            pass
                        if (key in FoundDataBIDS and FieldDataFromSF[key] == FoundDataBIDS[key]):
                            print(key + " in Sales Force matched with BIDS data")
                            TestResult.append(key + " in Sales Force matched with BIDS data")
                            TestResultStatus.append("Pass")
                        elif (key in FoundDataBIDS and FieldDataFromSF[key] != FoundDataBIDS[key]):
                            print(key + " in Sales Force does not match with BIDS data")
                            TestResult.append(key + " in Sales Force does not match with BIDS data")
                            TestResultStatus.append("Fail")

                    sheet.cell(row=1, column=6).value = "Verified"
                    for Pname in range(2,100):
                        if sheet.cell(row=Pname, column=7).value == None:
                            sheet.cell(row=Pname, column=7).value=OppName
                            break

                    for Pname1 in range(2,100):
                        if sheet.cell(row=Pname1, column=9).value == None:
                            sheet.cell(row=Pname1, column=9).value=OppName
                            break
                    wb.save(loc)
                except Exception as err:
                    print(err)
                    sheet.cell(row=1, column=6).value = "Unverified"
                    wb.save(loc)
                    pass

        except Exception as Mainerror:
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


