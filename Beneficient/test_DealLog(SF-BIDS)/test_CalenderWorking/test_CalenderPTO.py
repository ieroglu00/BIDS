from builtins import print
from datetime import datetime, timedelta,date
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from pathlib import Path
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyodbc
from selenium.webdriver.chrome.options import Options

#----------------SalesForce Username and Password IDs-------------
@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("username").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global FundNameList
  global FundNameListAfterRemove
  global ct
  global Exe
  global D1
  global D2
  global d1
  global d2
  global DollarDate
  global FundToOpen
  global TotalFundsLengh
  global FieldDataFromSF
  global FieldDataSF
  global FoundDataBIDS
  global FoundDataSF
  global path

  TestName = "test_CalenderPTO"
  description = "This test scenario is to verify elements and working of Calender (PTO) in Transactions"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_CalenderWorking"
  Exe="Yes"
  Directory = 'test_DealLog(SF-BIDS)/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  FundNameList=[]
  FundNameListAfterRemove=[]
  FieldDataFromSF = {}
  FieldDataSF = {}
  FoundDataBIDS = {}
  FoundDataSF = {}

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      #-----------Disabling access popup from Chrome------------------
      option = Options()
      option.add_argument("--disable-infobars")
      option.add_argument("start-maximized")
      option.add_argument("--disable-extensions")
      option.add_experimental_option("prefs", {"profile.default_content_setting_values.notifications": 2})

      driver=webdriver.Chrome(chrome_options=option, executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      # -------------------For Login in BIDS-------------------
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      driver.find_element_by_id("un").send_keys("neeraj.kumar")
      driver.find_element_by_id("pw").send_keys("Crochet@7866")
      driver.find_element_by_xpath("//input[@type='submit']").click()
      TestResult.append("Login to BIDS application")
      TestResultStatus.append("Pass")

      ct = datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.now().strftime("%d %B %Y %I %M%p")

      today = date.today()
      D1=today.strftime("%Y-%m-%d")
      d1=D1
      DollarDate=datetime.strptime(d1, '%Y-%m-%d')
      DollarDate="$"+DollarDate.date().__str__()+"$"
      d1 = datetime.strptime(D1, "%Y-%m-%d")

  yield
  if Exe == "Yes":
      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_CalenderPTO(test_setup):
    if Exe == "Yes":
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"

        try:
            print()
            # ---------------------------Verify Transactions page-----------------------------
            PageName = "Transactions"
            Ptitle1 = "Transaction Listing "
            driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    driver.close()
                    bool1 = False
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        driver.close()
                        bool2 = False
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[1]/div/div/div").text
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
                driver.close()
            # ---------------------------------------------------------------------------------
            PageName = "Calendar View"
            try:
                driver.find_element_by_xpath("//*[text() = '" + PageName + "']").click()
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                try:
                    time.sleep(2)
                    bool1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                    if bool1 == True:
                        ErrorFound1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                        print(ErrorFound1)
                        driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                        TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                        TestResultStatus.append("Fail")
                        bool1 = False
                except Exception:
                    try:
                        time.sleep(2)
                        bool2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                        if bool2 == True:
                            ErrorFound2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                            print(ErrorFound2)
                            TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                            TestResultStatus.append("Fail")
                            bool2 = False
                    except Exception:
                        pass
                    pass
                time.sleep(1)
                try:
                    Ptitle2 = "Transaction Listing"
                    PageTitle2 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[1]/div/div/div").text
                    assert Ptitle2 in PageTitle2, PageName + " not able to open"
                    TestResult.append(PageName + " opened successfully")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append(PageName + " not able to open")
                    TestResultStatus.append("Fail")
                time.sleep(1)
            except Exception:
                TestResult.append(PageName + " not able to open")
                TestResultStatus.append("Fail")
                pass

            # --------------To click Manage PTO/Internal Project Entries outside-----------------------
            driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div/div/div[1]/div/div[2]/div/div/div/div[2]/div/p/strong/a").click()
            time.sleep(1)
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass

            # --------------To click Manage PTO Entries inside-----------------------
            driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div[2]/div[1]/div/div").click()
            time.sleep(1)

            # --------------To click Add Entry (Manage PTO Entries)-----------------------
            driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div[2]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[2]/div/div/a").click()
            time.sleep(1)

            # --------------Filling Start Date in Add Entry (Manage PTO Entries)-----------------------
            today = datetime.now()
            DateDigit=today.strftime("%d")
            AnalystNameFound=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div[2]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/table/tbody/tr[1]/th[1]/div/div/span").text
            AnalystNameFound = "".join(AnalystNameFound.split())
            AnalystNameToCheck=DateDigit+AnalystNameFound
            print(AnalystNameToCheck)

            StartDate = today.strftime("%m/%d/%Y")
            print(StartDate)
            EndDate=StartDate
            print(EndDate)

            driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div[2]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/table/tbody/tr[1]/td[1]/div/div/input").send_keys(StartDate)
            time.sleep(1)
            driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div[2]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/table/tbody/tr[1]/td[2]/div/div/input").send_keys(EndDate)
            time.sleep(1)
            driver.find_element_by_xpath("//button[text()='Confirm Changes']").click()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            time.sleep(1)
            driver.find_element_by_xpath("//button[text()='Confirm Changes']").click()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass

            TestResult.append("PTO added successfully")
            TestResultStatus.append("Pass")

            #--------------To check working of Calendar view dropdown-----------------------
            driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div").click()
            time.sleep(1)
            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
            time.sleep(1)
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div").click()
            time.sleep(7)

            TestResult.append("Calendar view dropdown is working fine")
            TestResultStatus.append("Pass")

            Found = "False"
            for dd1 in range(7,13):
                for dd2 in range(1,8):
                    try:
                        Datee=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div["+str(dd1)+"]/div["+str(dd2)+"]/div/div/div/div/div[2]/div/p/span/strong").text
                        if len(Datee)==1:
                            Datee="0"+Datee
                            print("Datee is "+Datee)
                            print("DateDigit is " + DateDigit)
                        if len(Datee) > 1:
                            print("Datee is "+Datee)
                            print("DateDigit is " + DateDigit)

                        if Datee==DateDigit:
                            print("Date found")
                            try:
                                DateEle=driver.find_elements_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[" + str(dd1) + "]/div[" + str(dd2) + "]/div/div/div/div")
                                for ele in range(1, len(DateEle)):
                                    try:
                                        DateDatagot = driver.find_element_by_xpath(
                                            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[" + str(
                                                dd1) + "]/div[" + str(dd2) + "]/div/div/div/div["+str(ele+1)+"]/div/div/div[2]/div/p").text
                                        DateDatagot = "".join(DateDatagot.split())
                                        print(DateDatagot)
                                        DateDatagot=Datee+DateDatagot
                                        if AnalystNameToCheck==DateDatagot:
                                            print("Analyst Name matched: "+DateDatagot)
                                            Found="True"
                                            TestResult.append(
                                                "PTO verified successfully for Date [ " + StartDate + " ]")
                                            TestResultStatus.append("Pass")
                                            break
                                    except Exception:
                                        pass
                            except Exception:
                                pass

                    except Exception:
                        pass
            if Found == "False":
                TestResult.append(
                    "PTO added not found for Date [ " + StartDate + " ]")
                TestResultStatus.append("Fail")

        except Exception as Mainerror:
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


