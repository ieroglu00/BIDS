from builtins import print
from datetime import datetime, timedelta,date
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from pathlib import Path
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyodbc
from selenium.webdriver.chrome.options import Options

#----------------SalesForce Username and Password IDs-------------
@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("username").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global FundNameList
  global FundNameListAfterRemove
  global ct
  global Exe
  global D1
  global D2
  global d1
  global d2
  global DollarDate
  global FundToOpen
  global TotalFundsLengh
  global FieldDataFromSF
  global FieldDataSF
  global FoundDataBIDS
  global FoundDataSF
  global path

  TestName = "test_TransactionDealTemplate"
  description = "This test scenario is to verify elements inside Transactions Deal Template section and mapping Fund with Transaction"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_TransationDealTemplate"
  Exe="Yes"
  Directory = 'test_DealLog(SF-BIDS)/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  FundNameList=[]
  FundNameListAfterRemove=[]
  FieldDataFromSF = {}
  FieldDataSF = {}
  FoundDataBIDS = {}
  FoundDataSF = {}

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      #-----------Disabling access popup from Chrome------------------
      option = Options()
      option.add_argument("--disable-infobars")
      option.add_argument("start-maximized")
      option.add_argument("--disable-extensions")
      option.add_experimental_option("prefs", {"profile.default_content_setting_values.notifications": 2})

      driver=webdriver.Chrome(chrome_options=option, executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      # -------------------For Login in BIDS-------------------
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      driver.find_element_by_id("un").send_keys("neeraj.kumar")
      driver.find_element_by_id("pw").send_keys("Crochet@7866")
      driver.find_element_by_xpath("//input[@type='submit']").click()
      TestResult.append("Login to BIDS application")
      TestResultStatus.append("Pass")

      ct = datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.now().strftime("%d %B %Y %I %M%p")

      today = date.today()
      D1=today.strftime("%Y-%m-%d")
      d1=D1
      DollarDate=datetime.strptime(d1, '%Y-%m-%d')
      DollarDate="$"+DollarDate.date().__str__()+"$"
      d1 = datetime.strptime(D1, "%Y-%m-%d")

  yield
  if Exe == "Yes":
      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_DealLog_SFBIDSPhase1(test_setup):
    if Exe == "Yes":
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"

        try:
            print()
            # ---------------------------Verify Transactions page-----------------------------
            PageName = "Transactions"
            Ptitle1 = "Transaction Listing "
            driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
            # start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    driver.close()
                    bool1 = False
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        driver.close()
                        bool2 = False
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[1]/div/div/div").text
                print(PageTitle1)
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
                driver.close()
            # ---------------------------------------------------------------------------------
            print()
            TotalItem = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/span[2]").text
            substr = "of"
            x = TotalItem.split(substr)
            string_name = x[0]
            TotalItemAfterOf = x[1]
            abc = ""
            countspace = 0
            for element in range(0, len(string_name)):
                if string_name[(len(string_name) - 1) - element] == " ":
                    countspace = countspace + 1
                    if countspace == 2:
                        break
                else:
                    abc = abc + string_name[(len(string_name) - 1) - element]
            abc = abc[::-1]
            TotalItemBeforeOf = abc
            TotalItemAfterOf = re.sub('[^A-Za-z0-9]+', '', TotalItemAfterOf)
            TotalItemBeforeOf = re.sub('[^A-Za-z0-9]+', '', TotalItemBeforeOf)
            print("TotalItemAfterOf " + TotalItemAfterOf)
            print("TotalItemBeforeOf " + TotalItemBeforeOf)

            IterateNo = int(TotalItemAfterOf) / int(TotalItemBeforeOf)
            if IterateNo.is_integer() == True:
                IterateNo = IterateNo - 1
                pass
            else:
                IterateNo = math.ceil(float(IterateNo))
            print("IterateNo is " + str(IterateNo))

            #----------Reading Project Name to open--------
            ExcelFileName = "FieldData"
            loc = (path + 'Reference Data/' + ExcelFileName + '.xlsx')
            wb = openpyxl.load_workbook(loc)
            sheet = wb.get_sheet_by_name("DealLogPhase1")

            for Pname in range(2, 100):
                if sheet.cell(row=Pname, column=8).value == None:
                    ProjectName = sheet.cell(row=Pname, column=7).value
                    break
            print(ProjectName)
            #ProjectName="Funds for (LR67511)"
            FundToMap="21 Centrale Partners III"

            # ----------------Searching the Project from Sales Force--------------------
            LoopExit = 0
            ProejctTOClick = ProjectName
            PageName = ProejctTOClick
            start = time.time()
            WaitingCount = 13
            if IterateNo >= 5 and IterateNo < 7:
                WaitingCount = 11
            elif IterateNo >= 7 and IterateNo < 9:
                WaitingCount = 9
            elif IterateNo >= 9:
                WaitingCount = 5
            print(WaitingCount)
            for waiting in range(1, WaitingCount):
                if LoopExit == 0:
                    print("waiting is "+str(waiting))
                    if waiting>0:
                        driver.refresh()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(
                                EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass

                    loopbreak = 0
                    for ii5 in range(1, IterateNo + 1):
                        if loopbreak == 0:
                            if ii5 > 1:
                                try:
                                    driver.find_element_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                                    try:
                                        WebDriverWait(driver, SHORT_TIMEOUT
                                                      ).until(
                                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                        WebDriverWait(driver, LONG_TIMEOUT
                                                      ).until_not(
                                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                    except TimeoutException:
                                        pass
                                except Exception as q1:
                                    print(q1)
                                    pass

                            RowsInv = driver.find_elements_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr")
                            for ii3 in range(1, len(RowsInv) + 1):
                                ProjectNameText = driver.find_element_by_xpath(
                                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                                        ii3) + "]/td[1]/div/p/a").text
                                if ProjectNameText == ProejctTOClick:
                                    loopbreak = 1
                                    driver.find_element_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr/td[1]/div/p/a[text()='"+ProjectName+"']/parent::p/parent::div//parent::td/parent::tr/td[2]/div/p/a").click()
                                    TestResult.append("Transaction ID clicked for Fund [ "+PageName + " ]")
                                    TestResultStatus.append("Pass")
                                    try:
                                        WebDriverWait(driver, SHORT_TIMEOUT
                                                      ).until(
                                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                        WebDriverWait(driver, LONG_TIMEOUT
                                                      ).until_not(
                                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                    except TimeoutException:
                                        pass
                                    try:
                                        time.sleep(2)
                                        bool1 = driver.find_element_by_xpath(
                                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                                        if bool1 == True:
                                            ErrorFound1 = driver.find_element_by_xpath(
                                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                            print(ErrorFound1)
                                            driver.find_element_by_xpath(
                                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                            TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                                            TestResultStatus.append("Fail")
                                            driver.close()
                                    except Exception:
                                        try:
                                            time.sleep(2)
                                            bool2 = driver.find_element_by_xpath(
                                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                            if bool2 == True:
                                                ErrorFound2 = driver.find_element_by_xpath(
                                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                                print(ErrorFound2)
                                                TestResult.append(
                                                    PageName + " not able to open\n" + ErrorFound2)
                                                TestResultStatus.append("Fail")
                                                driver.close()
                                        except Exception:
                                            pass
                                        pass
                                    LoopExit = 1
                                    break
                        else:
                            break
                else:
                    break

            if LoopExit == 0:
                stop = time.time()
                RoundFloatString = round(float(stop - start), 2)
                seconds = int(RoundFloatString)
                min, sec = divmod(seconds, 60)
                TestResult.append(
                    PageName + " transactions ID not able to find in BIDS application\nTotal run time to find the project in BIDS is " + str(
                        min) + " mintues")
                TestResultStatus.append("Fail")
                driver.close()

            # ------------------------clicking Deal Template--------------------------------
            PageName = "Deal Template"
            driver.find_element_by_xpath(
                "//button[text()='Deal Template']").click()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(
                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    driver.close()
                    bool1 = False
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        driver.close()
                        bool2 = False
                except Exception:
                    pass
                pass
            time.sleep(1)
            TestResult.append(PageName + " button clicked in Transactions")
            TestResultStatus.append("Pass")

            # --------------------Deal Template Status Summary table--------------
            inside="Deal Template Status Summary"
            # ---------------loop for Columns in table---------
            ItemList = ["Fund⠀Name", "PT Assoc?","Status","Date⠀Approved","Coverage","Include in template?"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[1]/div/div[2]/div/div/div[2]/div/div/div[2]/table/thead/tr[1]/th[" + str(
                            ii + 1) + "]/div").text
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under  [ " + inside + " ]  table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under  [ " + inside + " ]  table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under  [ " + inside + " ]  table\n" + ListD)
                TestResultStatus.append("Fail")

            # --------------------Deal Templates Sent to Risk table--------------
            inside = "Deal Templates Sent to Risk"
            # ---------------loop for Columns in table----------
            ItemList = ["Template", "Document", "Filenames"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div[2]/table/thead/tr[1]/th[1]/div").text
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under  [ " + inside + " ]  table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under  [ " + inside + " ]  table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under  [ " + inside + " ]  table\n" + ListD)
                TestResultStatus.append("Fail")

            # --------------------Deal Template Support table--------------
            inside = "Deal Template Support"
            # ---------------loop for Columns in table---------
            ItemList = ["Source Documents", "Fund", "Uploaded On"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[3]/div/div[2]/div/div/div[2]/div/div/div[2]/table/thead/tr[1]/th[" + str(
                            ii + 1) + "]/div").text
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under  [ " + inside + " ]  table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under  [ " + inside + " ]  table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under  [ " + inside + " ]  table\n" + ListD)
                TestResultStatus.append("Fail")

            # --------------------Deal Template Review table--------------
            inside = "Deal Template Review"
            # ---------------loop for Columns in table---------
            ItemList = ['Template Status', 'UW Analyst', 'CIQ⠀ID', 'Fund', 'Asset', 'Date⠀(NAV)', 'Fund⠀NAV⠀ ⠀(local)', 'LP⠀NAV⠀⠀ ⠀(local)', 'LP Commitment (local)','LP Unfunded Commitment (local)', 'LP NAV + Unfunded Comm (local)', 'Fund⠀NAV (USD)', 'LP⠀NAV (USD)','LP Commitment (USD)', 'LP Unfunded Commitment (USD)', 'LP⠀NAV + Unfunded Comm (USD)', 'Country Code','Currency', 'GICS⠀Code', 'Sector⠀Key', 'Industry⠀Group⠀Key', 'Industry⠀Key','Sub⠀Industry⠀Key', 'Vintage Year', 'BCS⠀L1', 'BCS⠀L2', 'BCS⠀L3', 'BCS⠀L4', 'BCS⠀L5']

            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[3]/div[2]/div/div[2]/div[2]/div/div/div[2]/table/thead/tr[1]/th[" + str(
                            ii + 1) + "]/div").text
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under  [ " + inside + " ]  table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under  [ " + inside + " ]  table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under  [ " + inside + " ]  table\n" + ListD)
                TestResultStatus.append("Fail")

            #-------------Checking if Fund available in Deal Template Status Summary table-------------
            FundPresent=False
            try:
                FundPresent=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[1]/div/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td[2]").is_displayed()
            except Exception:
                pass
            if FundPresent==False:
                #------Adding fund------------
                TestResult.append("No Fund was found in Deal Template Status Summary table")
                TestResultStatus.append("Pass")
                driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[2]/div[1]/div[1]/div[2]/div/div").click()
                import pyperclip
                TEXT = FundToMap
                pyperclip.copy(TEXT)
                time.sleep(1)
                ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(1)
                ActionChains(driver).key_down(Keys.DOWN).key_up(Keys.DOWN).perform()
                time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                try:
                    time.sleep(2)
                    bool1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                    if bool1 == True:
                        ErrorFound1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                        print(ErrorFound1)
                        driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                        TestResult.append("Not able to add Fund in Deal Template Status Summary table. Below error found\n" + ErrorFound1)
                        TestResultStatus.append("Fail")
                        driver.close()
                        bool1 = False
                except Exception:
                    try:
                        time.sleep(2)
                        bool2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                        if bool2 == True:
                            ErrorFound2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                            print(ErrorFound2)
                            TestResult.append("Not able to add Fund in Deal Template Status Summary table. Below error found\n" + ErrorFound2)
                            TestResultStatus.append("Fail")
                            driver.close()
                            bool2 = False
                    except Exception:
                        pass
                    pass
                time.sleep(1)
                driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[2]/div[2]/div/div/button").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                try:
                    time.sleep(2)
                    bool1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                    if bool1 == True:
                        ErrorFound1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                        print(ErrorFound1)
                        driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                        TestResult.append("----- not able to click. Below error found\n" + ErrorFound1)
                        TestResultStatus.append("Fail")
                        driver.close()
                        bool1 = False
                except Exception:
                    try:
                        time.sleep(2)
                        bool2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                        if bool2 == True:
                            ErrorFound2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                            print(ErrorFound2)
                            TestResult.append("----- not able to click. Below error found\n" + ErrorFound2)
                            TestResultStatus.append("Fail")
                            driver.close()
                            bool2 = False
                    except Exception:
                        pass
                    pass
                time.sleep(1)
                TestResult.append("Fund [ "+FundToMap+" ] added in Deal Template Status Summary table")
                TestResultStatus.append("Pass")
                driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[1]/div/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td[1]/div/p/a[text()='"+TEXT+"']").click()
                TestResult.append("Fund [ " + FundToMap + " ] clicked to add Potential Trust")
                TestResultStatus.append("Pass")
                driver.switch_to.window(driver.window_handles[1])
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div[1]/div/div/div/div[1]/div[3]/div/div[2]/div/div").click()
                TEXT = ProjectName
                pyperclip.copy(TEXT)
                ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(1)
                ActionChains(driver).key_down(Keys.DOWN).key_up(Keys.DOWN).perform()
                time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(
                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                #Get FundDocNum and other Funds details from Funds page ------------------------

                RowFund=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[3]/div/div[1]/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[3]/div/input").get_attribute("value")
                if RowFund=="21 Centrale Partners III":
                    FUNDNAV = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div[1]/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[6]/div/input").get_attribute(
                        "value")
                    if FUNDNAV == "_":
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[3]/div/div[1]/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[6]/div/input").send_keys(
                            "120000")
                    time.sleep(1)
                    LPNAV= driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div[1]/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[7]/div/input").get_attribute(
                        "value")
                    if LPNAV=="_":
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[3]/div/div[1]/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[7]/div/input").send_keys("10000")
                    time.sleep(1)
                    driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[3]/div/div[1]/div/div/div/div[1]/div[4]/div/div[2]/div/p/a").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div[3]/div/div/div/div/div[2]/div/div/button").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    for winclose in range(1, 10):
                        time.sleep(1)
                        if len(driver.window_handles) > 1:
                            driver.switch_to.window(driver.window_handles[1])
                            driver.close()
                        elif len(driver.window_handles) == 1:
                            break
                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(1)
                    driver.refresh()
                    time.sleep(1)
                    try:
                        driver.switch_to_alert().accept()
                    except Exception:
                        pass
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    # ------------------------clicking Deal Template--------------------------------
                    time.sleep(2)
                    PageName = "Deal Template"
                    driver.find_element_by_xpath(
                        "//button[text()='Deal Template']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(
                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(2)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                        if bool1 == True:
                            ErrorFound1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                            print(ErrorFound1)
                            driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                            TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                            TestResultStatus.append("Fail")
                            driver.close()
                            bool1 = False
                    except Exception:
                        try:
                            time.sleep(2)
                            bool2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                            if bool2 == True:
                                ErrorFound2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                print(ErrorFound2)
                                TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                TestResultStatus.append("Fail")
                                driver.close()
                                bool2 = False
                        except Exception:
                            pass
                        pass
                    time.sleep(1)
                    TestResult.append(PageName + " clicked in Transactions for Fund [ "+FundToMap+" ]")
                    TestResultStatus.append("Pass")
                    #----Verify Fund is approved----
                    try:
                        FundVerifyStatus = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[1]/div/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td[3]/div/p/span").text

                        if FundVerifyStatus=="Approved⠀":
                            print("FundVerifyStatus 2 is "+FundVerifyStatus)
                            TestResult.append(
                                "Fund mapping with Transaction in Deal Template page is working correctly")
                            TestResultStatus.append("Pass")

                        elif FundVerifyStatus=="Unapproved⠀":
                            print("FundVerifyStatus 3 is "+FundVerifyStatus)
                            TestResult.append("Fund mapping with Transaction in Deal Template page is not working correctly\nFund found Unapproved after approval process")
                            TestResultStatus.append("Fail")

                    except Exception:
                        FundVerifyStatus = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[1]/div/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td").text
                        print("FundVerifyStatus 4 is " + FundVerifyStatus)

                        if FundVerifyStatus=="No items available":
                            print("FundVerifyStatus 5 is "+FundVerifyStatus)
                            TestResult.append("Fund mapping with Transaction in Deal Template page is not working correctly\nNo Fund found in Deal Template Status Summary table after approval process")
                            TestResultStatus.append("Fail")
                        pass

                    if FundVerifyStatus == "Approved⠀":
                        print("Perform further tasks")

                        TestResult.append("Verification started for Deal Templates Sent to Risk table data")
                        TestResultStatus.append("Pass")

                        driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[2]/div[6]/div/div/button").click()
                        TestResult.append("Export to Excel button clicked")
                        TestResultStatus.append("Pass")
                        DownloadDoc=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr[1]/td[1]/div/p/a").text
                        TestResult.append("Deal Template document in Deal Templates Sent to Risk table found after Export to Excel. Below Doc found\n"+DownloadDoc)
                        TestResultStatus.append("Pass")

                        TestResult.append("Verification started for source documents in [ Deal Template Support ] table")
                        TestResultStatus.append("Pass")

                        SourceDocNum = driver.find_elements_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[3]/div/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr")
                        print("Total SourceDocNum found "+str(len(SourceDocNum)))

                        FundDocNum=37
                        if FundDocNum==len(SourceDocNum):
                            TestResult.append(
                                "Total source documents in [ Deal Template Support ] table matched with docs present in Fund")
                            TestResultStatus.append("Pass")
                        else:
                            TestResult.append(
                                "Total source documents in [ Deal Template Support ] table does not match with docs present in Fund")
                            TestResultStatus.append("Fail")
                else:
                    driver.close()
                    TestResult.append("Not able to map Potential Trust with Fund")
                    TestResultStatus.append("Fail")

            elif FundPresent==True:
                print("Fund is already Present _____________________")
                FundVerifyStatus = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[1]/div/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td").text
                print("FundVerifyStatus 4 is " + FundVerifyStatus)

                if FundVerifyStatus == "Approved⠀":
                    TestResult.append("Verification started for Deal Templates Sent to Risk table data")
                    TestResultStatus.append("Pass")

                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[2]/div[6]/div/div/button").click()
                    TestResult.append("Export to Excel button clicked")
                    TestResultStatus.append("Pass")
                    DownloadDoc = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr[1]/td[1]/div/p/a").text
                    TestResult.append(
                        "Deal Template document in Deal Templates Sent to Risk table found after Export to Excel. Below Doc found\n" + DownloadDoc)
                    TestResultStatus.append("Pass")

                    TestResult.append("Verification started for source documents in [ Deal Template Support ] table")
                    TestResultStatus.append("Pass")

                    SourceDocNum = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[3]/div/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr")
                    print("Total SourceDocNum found " + str(len(SourceDocNum)))

                    FundDocNum = 37
                    if FundDocNum == len(SourceDocNum):
                        TestResult.append(
                            "Total source documents in [ Deal Template Support ] table matched with docs present in Fund")
                        TestResultStatus.append("Pass")
                    else:
                        TestResult.append(
                            "Total source documents in [ Deal Template Support ] table does not match with docs present in Fund")
                        TestResultStatus.append("Fail")

            sheet.cell(row=Pname, column=8).value = "Verified"
            wb.save(loc)

        except Exception as Mainerror:
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


