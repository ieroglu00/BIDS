import datetime
import time
from telnetlib import EC

import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_VerifyClickables"
  description = "This test scenario is to verify all the clickable elements such as Buttons, Hyperlinks and clickable tabs"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_COVE_to_BIDS"
  global Exe
  Exe="Yes"
  Directory = 'test_COVE_to_BIDS/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_AllModulesVerifyCOVE(test_setup):
    if Exe == "Yes":
        print()
        PageName="Transactions"
        Ptitle1="Transactions - BIDS"
        driver.find_element_by_xpath("//*[@title='"+PageName+"']").click()
        for iat1 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        try:
            try:
                PageTitle1 = driver.title
                print(PageTitle1)
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
            except Exception:
                Ptitle1="Funds - BIDS"
                PageTitle1 = driver.title
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " page not able to open")
            TestResultStatus.append("Fail")

        # PageName = "Task Management"
        # driver.find_element_by_xpath("//*[text() = '"+PageName+"']").click()
        # for iat2 in range(1000):
        #     try:
        #         bool = driver.find_element_by_xpath(
        #             "//div[@id='appian-working-indicator-hidden']").is_enabled()
        #     except Exception:
        #         time.sleep(1)
        #         break
        # time.sleep(1)
        # Ptitle2 = "Display Tasks For Analyst:"
        # PageTitle2 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div/div/div/div[2]/p").text
        # try:
        #     print(PageTitle2)
        #     assert Ptitle2 in PageTitle2, PageName + " not able to open"
        #     TestResult.append(PageName + " page Opened successfully")
        #     TestResultStatus.append("Pass")
        # except Exception:
        #     TestResult.append(PageName + " page not able to open")
        #     TestResultStatus.append("Fail")
        # driver.find_element_by_xpath("//*[@title='Transactions']").click()
        # for iat3 in range(1000):
        #     try:
        #         bool = driver.find_element_by_xpath(
        #             "//div[@id='appian-working-indicator-hidden']").is_enabled()
        #     except Exception:
        #         time.sleep(1)
        #         break
        # time.sleep(1)

        PageName = "Transaction ID"
        try:
            driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[2]/div/p/a").click()
        except Exception:
            time.sleep(7)
            try:
                driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[2]/div/p/a").click()
            except Exception:
                TestResult.append(PageName + " not able to open on click")
                TestResultStatus.append("Fail")
        for iat4 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        Ptitle3 = "Transaction NAV Concentration"
        print(driver.title)
        PageTitle3 = driver.find_element_by_xpath(
            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[1]/div[1]/div[1]/div/div/div").text
        try:
            assert Ptitle3 in PageTitle3, PageName + " not able to open"
            TestResult.append(PageName + " clicked and opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " not able to open on click")
            TestResultStatus.append("Fail")

        PageName = "Transaction Mgmt"
        driver.find_element_by_xpath(
            "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/button").click()
        for iat4 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        Ptitle4 = "Transaction Workflow"
        PageTitle4 = driver.find_element_by_xpath(
            "//div[@class='ContentLayout---content_layout']/div[4]/div[1]/div/div[1]/div[1]").text
        try:
            print(PageTitle3)
            assert Ptitle4 in PageTitle4, PageName + " not able to open"
            TestResult.append(PageName + " button clicked successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " button not able to open on click")
            TestResultStatus.append("Fail")

        HyperlinksCount=19
        for ii in range (1,HyperlinksCount+1):
            try:
                time.sleep(1)
                try:
                    try:
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div["+str(ii)+"]/div[2]/div/div[2]/div/p/span/a").click()
                    except Exception:
                        time.sleep(1)
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div[" + str(
                                ii) + "]/div[2]/div/div[2]/div/p/span/strong/a").click()
                    #print("A1")
                    for iat5 in range(1000):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        except Exception:
                            time.sleep(1)
                            break
                    Element = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div[" + str(
                            ii) + "]/div[2]/div/div[2]/div/p/span/strong/a").text
                    print(Element)
                    Element = Element.split(' ', 1)
                    Element = Element[1]
                    TestResult.append(Element + " opened successfully")
                    TestResultStatus.append("Pass")

                except Exception:
                    try:
                        time.sleep(2)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div/h4").is_displayed()
                        ErrorText = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div/p").text
                        print(ErrorText)
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[2]/div/button").click()
                        time.sleep(5)
                        TestResult.append("Transaction Workflow Hyperlink at "+str(ii)+") position is not able to open on click\n"+ErrorText)
                        TestResultStatus.append("Fail")
                    except Exception as alertExp:
                        TestResult.append("Transaction Workflow Hyperlink at "+str(ii)+") position is not able to open on click")
                        TestResultStatus.append("Fail")
                        pass

            except Exception:
                TestResult.append("Transaction Workflow under Transaction MGMT Tanb is not able to open")
                TestResultStatus.append("Fail")
                pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


