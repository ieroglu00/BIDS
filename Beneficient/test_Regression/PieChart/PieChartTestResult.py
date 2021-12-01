import numpy as np
import openpyxl
import pytest
from matplotlib import pyplot as plt

@pytest.mark.smoke
def test_PieChartTestResult():
    TestStatus = []
    Directory = 'test_Regression/'
    path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

    #-------------------To read content to send in e-Mail--------------------
    ExcelFileName = "FileName"
    loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
    wb=openpyxl.load_workbook(loc)
    sheet = wb.active
    for i in range(1, 100):
        if sheet.cell(i, 1).value == None:
            break
        else:
            TestStatus.append(sheet.cell(i, 5).value)
    #--------------To create Pie Chart and attach in email------------------

    T_Tests=len(TestStatus)
    PassCount = TestStatus.count("Pass")
    FailCount = TestStatus.count("Fail")
    SkippedCount = TestStatus.count("Skipped")
    print("PassCount"+ str(PassCount))
    print("FailCount"+ str(FailCount))
    print("SkippedCount"+ str(SkippedCount))

    PassCountPer = round((PassCount / T_Tests) * 100, 2)
    FailCountPer=round((FailCount/T_Tests)*100 , 2)
    SkippedCountPer=round((SkippedCount/T_Tests)*100 , 2)
    print("PassCountPer"+ str(PassCountPer))
    print("FailCountPer"+ str(FailCountPer))
    print("SkippedCountPer"+ str(SkippedCountPer))

    y = np.array([PassCountPer, FailCountPer, SkippedCountPer])
    mylabels = ["Pass "+str(PassCount), "Fail "+str(FailCount), "Skipped "+str(SkippedCount)]
    mycolors = ["Green", "Red", "Grey"]
    plt.pie(y, labels=mylabels, startangle=90, colors=mycolors)
    plt.legend(title="Testing Suite "
                     "Status: "+str((PassCount+FailCount+SkippedCount)))
    plt.savefig(path+'/TestPieResult.png', format='png', dpi=300)

    #-----------------------------------------------------------------------