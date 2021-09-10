import time

import openpyxl
import xlrd
from openpyxl import Workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.expected_conditions import staleness_of
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import pytest
from selenium import webdriver
import allure
import pandas as pd

# @allure.step("Entering username ")
# def enter_username(username):
#       driver.find_element_by_id("un").send_keys(username)
#
# @allure.step("Entering password ")
# def enter_password(password):
#       driver.find_element_by_id("pw").send_keys(password)

# @pytest.fixture()
# def test_setup():
#       global driver
#       driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
#       # driver.implicitly_wait(10)
#       # driver.maximize_window()
#       # driver.get("https://beneficienttest.appiancloud.com/suite/")
#       # enter_username("neeraj.kumar")
#       # enter_password("Crochet@786")
#       # driver.find_element_by_xpath("//input[@type='submit']").click()
#
#       yield
#       driver.quit()
#
# @pytest.mark.regression
# @allure.description("Test case to verfiy all links at Funds Inside Page")
# @allure.severity(severity_level="High")
def test_VerfyAllLinksFundsInsidePage():
    loc = ('C:/BIDS/beneficienttest/Beneficient/Testt/TestExcel.xlsx')
    wb = openpyxl.load_workbook(loc)
    sheet = wb.active
    print()

    check="ab1c1"
    PdfName="bbb"
    checkcount=0
    InputPDFName = []

    for i in range(1, 100):
        if sheet.cell(i, 1).value=="jwj":
            print("Filename found now")
            print(str(i))

    # This function is for getting "PDF name" from the PDFName excel file
    # for i in range(1, 100):
    #    if sheet.cell(i, 1).value==None:
    #        print("None found now")
    #        for ab in range(1,len(InputPDFName)+1):
    #             sheet.cell(row=ab, column=1).value =InputPDFName[ab-1]
    #        if checkcount == 0:
    #          sheet.cell(row=ab + 1, column=1).value = check
    #          sheet.cell(row=ab + 1, column=2).value = PdfName
    #          checkcount = 1
    #        wb.save(loc)
    #        break
    #    else:
    #        print(sheet.cell(i, 1).value)
    #        if sheet.cell(i, 1).value==check:
    #            if checkcount==0:
    #              sheet.cell(row=i, column=2).value = PdfName
    #              checkcount = 1
    #        InputPDFName.append(sheet.cell(i, 1).value)

