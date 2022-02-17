import datetime
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global FundNameList
  global FundNameListAfterRemove
  global ct
  global Exe
  global D1
  global D2
  global d1
  global d2
  global DollarDate
  global FundToOpen
  global TotalFundsLengh
  global ItemList

  TestName = "test_FundSOIMatchingValues"
  description = "This test scenario is to compare different fund level SOI values"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_FundSOI"
  Exe="Yes"
  Directory = 'test_FundSOI/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  FundNameList=[]
  FundNameListAfterRemove=[]
  ItemList = []

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()

      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      today = datetime.date.today()
      D1=today.strftime("%Y-%m-%d")
      d1=D1
      DollarDate=datetime.datetime.strptime(d1, '%Y-%m-%d')
      DollarDate="$"+DollarDate.date().__str__()+"$"
      d1 = datetime.datetime.strptime(D1, "%Y-%m-%d")

  yield
  if Exe == "Yes":
      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        FundToOpen=15
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        try:
            print()
            # ----------------------------------------------------------------------------
            #---------------------------Verify Funds page-----------------------------
            PageName="Funds"
            Ptitle1="Funds - BIDS"
            driver.find_element_by_xpath("//*[@title='"+PageName+"']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                PageTitle1 = driver.title
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName +     " page not able to open")
                TestResultStatus.append("Fail")
            #---------------------------------------------------------------------------------
            try:
                TotalItem = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[2]").text
                substr = "of"
                x = TotalItem.split(substr)
                string_name = x[0]
                TotalItemAfterOf = x[1]
                abc = ""
                countspace = 0
                for element in range(0, len(string_name)):
                    if string_name[(len(string_name) - 1) - element] == " ":
                        countspace = countspace + 1
                        if countspace == 2:
                            break
                    else:
                        abc = abc + string_name[(len(string_name) - 1) - element]
                abc = abc[::-1]
                TotalItemBeforeOf = abc
                print("TotalItemAfterOf " + TotalItemAfterOf)
                print("TotalItemBeforeOf " + TotalItemBeforeOf)

                IterateNo = int(TotalItemAfterOf) / int(TotalItemBeforeOf)
                IterateNo = math.ceil(float(IterateNo))
                print(IterateNo)
            except Exception:
                pass

            for ii5 in range(1, IterateNo + 1):
                RowsInv = driver.find_elements_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr")

                for ii3 in range(1, len(RowsInv)+1):
                    FundNameInvText = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                            ii3) + "]/td[2]/div/p/a").text
                    FundNameList.append(FundNameInvText)
                print(str(len(FundNameList)))

                if ii5 > 1 and ii5 < IterateNo:
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div /div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                elif ii5 == IterateNo:
                    pass
                else:
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                try:
                    time.sleep(2)
                    bool1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                    if bool1 == True:
                        ErrorFound1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                        print(ErrorFound1)
                        driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                        TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                        TestResultStatus.append("Fail")
                        bool1 = False
                        driver.close()
                except Exception:
                    try:
                        time.sleep(2)
                        bool2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                        if bool2 == True:
                            ErrorFound2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                            print(ErrorFound2)
                            TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                            TestResultStatus.append("Fail")
                            bool2 = False
                            driver.close()
                    except Exception:
                        pass
                    pass
                print()
                print(*FundNameList, sep="\n")

            print("length of FundNameList before "+str(len(FundNameList)))
            TestResult.append("Total "+str(len(FundNameList))+" Funds are present in Funds section")
            TestResultStatus.append("Pass")

            TotalFundsLengh=len(FundNameList)
            print("=======================================================================================")
            driver.refresh()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            #--------------------------------------------------------------------------------------
            #----------------Appending Funds details in Excel sheet-----------------------------------

            # -----------To add new found Funds in Excel sheet-------------
            ExcelFileName = "FundName"
            loc1 = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
            wb1 = openpyxl.load_workbook(loc1)
            sheet1 = wb1.active
            for i in range(len(FundNameList)):
                try:
                    if sheet1.cell(i + 1, 2).value !=None:
                        if sheet1.cell(i + 1, 2).value in FundNameList:
                            FundNameList.remove(sheet1.cell(i + 1, 2).value)
                except Exception as ef:
                    print(ef)
                    pass
            wb1.save(loc1)

            noneindex=0
            for iadd in range (TotalFundsLengh):
                if sheet1.cell(iadd + 1, 2).value == None:
                    if noneindex==0:
                        noneindex=iadd+1
                    sheet1.cell(row=iadd + 1, column=1).value = iadd + 1
                    sheet1.cell(row=iadd + 1, column=2).value = FundNameList[iadd-noneindex]
            wb1.save(loc1)
            #--------------------------------------------------------------------------------------

            # -----------To fetch selected Funds from total list-------------
            ExcelFileName = "FundName"
            loc1 = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
            wb1 = openpyxl.load_workbook(loc1)
            sheet1 = wb1.active

            for i2 in range(TotalFundsLengh):
                if len(FundNameListAfterRemove) <= FundToOpen - 1:
                    if sheet1.cell(i2 + 1, 3).value == None:
                        FundNameListAfterRemove.append(sheet1.cell(i2 + 1, 2).value)
                        sheet1.cell(i2 + 1, 3).value=DollarDate
                    else:
                        if sheet1.cell(i2 + 1, 3).value != None:
                            D2 = sheet1.cell(i2 + 1, 3).value
                            D2 = re.sub('[!@#$]', '', D2)
                            d2 = datetime.datetime.strptime(D2, "%Y-%m-%d")
                            if (d1 - d2).days > 7:
                                FundNameListAfterRemove.append(sheet1.cell(i2 + 1, 2).value)
                                sheet1.cell(i2 + 1, 3).value = DollarDate
                            else:
                                pass
            wb1.save(loc1)
            print(FundNameListAfterRemove)
            print(str(len(FundNameListAfterRemove)))
            TestResult.append("Below " + str(len(FundNameListAfterRemove)) + " funds are collected for verification")
            TestResultStatus.append("Pass")

            #--------------------------------------------------------------------------------------
            for ifundlist in range(len(FundNameListAfterRemove)):
                #print()
                #-------------to clear the cache memory of chrome browser-------------
                if (ifundlist/30).is_integer()==True and (ifundlist/30)!=0.0:
                    print("<<<<<<<<<<<Integer value found for 30 digit>>>>>>>>")
                    print(str((ifundlist / 30)))
                    driver.delete_all_cookies()
                    time.sleep(5)
                    driver.get("https://beneficienttest.appiancloud.com/suite/")
                    driver.find_element_by_id("un").send_keys("neeraj.kumar")
                    driver.find_element_by_id("pw").send_keys("Crochet@7866")
                    button = driver.find_element_by_xpath(
                        "//input[@type='submit']")
                    driver.execute_script("arguments[0].click();", button)
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                #--------------------------------------------------------------
                #print(FundNameListAfterRemove[ifundlist])
                main_window = driver.current_window_handle
                print(str(ifundlist)+" "+FundNameListAfterRemove[ifundlist])
                if "'" in FundNameListAfterRemove[ifundlist]:
                    print("******************************************  coma found *********************************************")
                    # print(FundNameList[ifundlist])
                    # fund=FundNameList[ifundlist]
                    # fund=fund.replace("'", "\'")
                    # print("afer changing: " + fund)
                    # driver.find_element_by_xpath("//*[text()='" + fund + "']").click()
                    # break
                    #FundNameList[ifundlist].index
                else:
                    ActionChains(driver).key_down(Keys.CONTROL).perform()
                    try:
                        driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='"+FundNameListAfterRemove[ifundlist]+"']").click()
                    except Exception:
                        try:
                            button = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='"+FundNameListAfterRemove[ifundlist]+"']")
                            driver.execute_script("arguments[0].click();", button)
                        except Exception:
                            print("Clicking on First pagination icon")
                            driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                            try:
                                WebDriverWait(driver, SHORT_TIMEOUT
                                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                WebDriverWait(driver, LONG_TIMEOUT
                                              ).until_not(
                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            except TimeoutException:
                                pass
                            #--------------------------------On Second Page--------------------------------
                            try:
                                driver.find_element_by_xpath(
                                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='" +
                                    FundNameListAfterRemove[ifundlist] + "']").click()
                            except Exception:
                                try:
                                    button = driver.find_element_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='" +
                                        FundNameListAfterRemove[ifundlist] + "']")
                                    driver.execute_script("arguments[0].click();", button)
                                except Exception:
                                    print("Clicking on Second pagination icon")
                                    driver.find_element_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                                    try:
                                        WebDriverWait(driver, SHORT_TIMEOUT
                                                      ).until(
                                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                        WebDriverWait(driver, LONG_TIMEOUT
                                                      ).until_not(
                                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                    except TimeoutException:
                                        pass
                                    # ----------------------------------On Third Page--------------------------------
                                    try:
                                        driver.find_element_by_xpath(
                                            "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='" +
                                            FundNameListAfterRemove[ifundlist] + "']").click()
                                    except Exception:
                                        try:
                                            button = driver.find_element_by_xpath(
                                                "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='" +
                                                FundNameListAfterRemove[ifundlist] + "']")
                                            driver.execute_script("arguments[0].click();", button)
                                        except Exception:
                                            print("Clicking on Third pagination icon")
                                            driver.find_element_by_xpath(
                                                "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a[1]").click()
                                            try:
                                                WebDriverWait(driver, SHORT_TIMEOUT
                                                              ).until(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                                WebDriverWait(driver, LONG_TIMEOUT
                                                              ).until_not(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                            except TimeoutException:
                                                pass
                                            # ---------------------------------On Fourth Page------------------------------
                                            try:
                                                driver.find_element_by_xpath(
                                                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='" +
                                                    FundNameListAfterRemove[ifundlist] + "']").click()
                                            except Exception:
                                                try:
                                                    button = driver.find_element_by_xpath(
                                                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td/div/p/a[text()='" +
                                                        FundNameListAfterRemove[ifundlist] + "']")
                                                    driver.execute_script("arguments[0].click();", button)
                                                except Exception:
                                                    pass

                    ActionChains(driver).key_up(Keys.CONTROL).perform()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass

                    #----------------------------------------------------------------------------------
                    driver.switch_to.window(driver.window_handles[1])
                    time.sleep(1)
                    try:
                        try:
                            Test=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[2]/p").text
                        except Exception:
                            time.sleep(5)
                            Test = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[2]/p").text
                    except Exception:
                        time.sleep(3)
                        #---------------When Page load error occurs---------------------------------
                        try:
                            if driver.find_element_by_xpath("//div[@id='main-frame-error']/div/div[2]/h1/span").is_displayed()==True:
                                driver.refresh()
                                try:
                                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                    WebDriverWait(driver, LONG_TIMEOUT
                                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                except TimeoutException:
                                    pass
                        except Exception:
                            pass
                        #-----------------------------------------------------------------------------
                        driver.switch_to.window(driver.window_handles[1])
                        try:
                            try:
                                Test=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[2]/p").text
                            except Exception:
                                time.sleep(5)
                                Test = driver.find_element_by_xpath(
                                    "//div[@class='ContentLayout---content_layout']/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[2]/p").text
                        except Exception:
                            try:
                                time.sleep(2)
                                bool1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                                if bool1 == True:
                                    ErrorFound1 = driver.find_element_by_xpath(
                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                    print(ErrorFound1)
                                    driver.find_element_by_xpath(
                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                                    TestResultStatus.append("Fail")
                                    bool1 = False
                                    Test = ErrorFound1
                            except Exception:
                                try:
                                    time.sleep(2)
                                    bool2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                    if bool2 == True:
                                        ErrorFound2 = driver.find_element_by_xpath(
                                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                        print(ErrorFound2)
                                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                                        TestResultStatus.append("Fail")
                                        bool2 = False
                                        Test = ErrorFound2
                                except Exception:
                                    pass
                                pass
                    # --------------------Clicking Fund-BS tab--------------
                    PageName = "Fund-BS"
                    Ptitle1 = "New/Edit Balance Sheet"
                    driver.find_element_by_xpath("//button[text()='"+PageName+"']").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(2)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                        if bool1 == True:
                            ErrorFound1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                            print(ErrorFound1)
                            driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                            TestResult.append("=> Fund [ "+FundNameListAfterRemove[ifundlist]+" ]\n"+PageName + " not able to open\n" + ErrorFound1)
                            TestResultStatus.append("Fail")
                            bool1 = False
                            driver.close()
                    except Exception:
                        try:
                            time.sleep(2)
                            bool2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                            if bool2 == True:
                                ErrorFound2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                print(ErrorFound2)
                                TestResult.append("=> Fund [ "+FundNameListAfterRemove[ifundlist]+" ]\n"+PageName + " not able to open\n" + ErrorFound2)
                                TestResultStatus.append("Fail")
                                bool2 = False
                                driver.close()
                        except Exception:
                            pass
                        pass
                    time.sleep(1)
                    try:
                        PageTitle1 = driver.find_element_by_xpath("//a[text()='New/Edit Balance Sheet']").text
                        assert Ptitle1 in PageTitle1, PageName + " is not able to open "
                        TestResult.append(PageName + " tab opened successfully for Fund [ "+ FundNameListAfterRemove[ifundlist]+" ]")
                        TestResultStatus.append("Pass")
                    except Exception as e1:
                        print(e1)
                        TestResult.append("=> Fund [ "+FundNameListAfterRemove[ifundlist]+" ]\n"+PageName + " is not able to open ")
                        TestResultStatus.append("Fail")
                    #--------------------------------------------------------------------------

                    inside = "Main Table"
                    # ---------------loop for Columns in table for Funds View----------
                    for ii1 in range(10):
                        try:
                            Element1 = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div[1]//table/thead/tr/th[" + str(ii1+2) + "]/div").text
                            ItemList.append(Element1)
                        except Exception as qw:
                            pass
                    #print(ItemList)

                    for ii2 in range(1,25):
                        try:
                            LabelName = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div[1]//table/tbody/tr[" + str(ii2) + "]/td[1]/div/p/span").text
                            if LabelName=="Total Assets":
                                TotalAssetsIndex=ii2
                            if LabelName=="Net Assets":
                                NetAssetsIndex=ii2
                            if LabelName=="Fund NAV":
                                FundNAVIndex=ii2
                            if LabelName=="Ben NAV":
                                BenNAVIndex=ii2
                            if LabelName=="Ben Ownership%":
                                BenOwnershipIndex=ii2
                        except Exception:
                            pass

                    ItemDic = {}
                    for ii3 in range(len(ItemList)):
                        #print()
                        #print()
                        Text1 = ItemList[ii3]
                        try:
                            TotalAssetsValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div[1]//table/tbody/tr["+str(TotalAssetsIndex)+"]/td[" + str(
                                    ii3 + 2) + "]/div/p/span").text
                            #print("TotalAssetsValue "+TotalAssetsValue)
                            NetAssetsValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div[1]//table/tbody/tr[" + str(NetAssetsIndex) + "]/td[" + str(
                                    ii3 + 2) + "]/div/p/span").text
                            #print("NetAssetsValue " + NetAssetsValue)
                            FundNAVValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div[1]//table/tbody/tr[" + str(FundNAVIndex) + "]/td[" + str(
                                    ii3 + 2) + "]/div/p/span").text
                            #print("FundNAVValue " + FundNAVValue)
                            BenNAVValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div[1]//table/tbody/tr[" + str(BenNAVIndex) + "]/td[" + str(
                                    ii3 + 2) + "]/div/p/span").text
                            #print("BenNAVValue " + BenNAVValue)
                            BenNAVValue = BenNAVValue.replace(" ", "")
                            BenNAVValue = re.sub(r'[?|$|!|,|-]', r'', BenNAVValue)
                            BenNAVValue = round(float(BenNAVValue))

                            BenOwnershipValue = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div[2]/div/div[1]//table/tbody/tr[" + str(BenOwnershipIndex) + "]/td[" + str(
                                    ii3 + 2) + "]/div/p/span").text
                            #print("BenOwnershipValue " + BenOwnershipValue)

                            ItemDic[ItemList[ii3]]=[TotalAssetsValue,NetAssetsValue,FundNAVValue,BenNAVValue,BenOwnershipValue]
                        except Exception:
                            pass
                    #print(ItemDic)

                    for ii4 in range(len(ItemList)):
                        QuarterText=ItemList[ii4]
                        substr = " "
                        x = QuarterText.split(substr)
                        string_name = x[1]
                        string_name1= string_name.split("Q")
                        QuarterYear=string_name1[1]
                        QuarterYear="/20"+QuarterYear
                        QuarterNumber = string_name1[0]
                        if QuarterNumber=="1":
                            QuarterNumber="03/31"
                        elif QuarterNumber=="2":
                            QuarterNumber="06/30"
                        elif QuarterNumber=="3":
                            QuarterNumber="09/30"
                        elif QuarterNumber=="4":
                            QuarterNumber="12/31"
                        QuarterText=QuarterNumber+QuarterYear
                        ItemDic[QuarterText] = ItemDic.pop(ItemList[ii4])
                    #print(ItemDic)

                    for ii5 in range(len(ItemList)):
                        print()
                        for ii6 in range(10):
                            DrpDwnYear = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div[5]/div/div[1]/div[1]/div/div[2]/div/div/span").text
                            if DrpDwnYear in ItemDic.keys():
                                #print(ItemDic[DrpDwnYear])
                                #print(ItemDic[DrpDwnYear][3])
                                TestResult.append(
                                    "For Quarter " + DrpDwnYear + " \n Total Assets Value: " + ItemDic[DrpDwnYear][0] + "\n Net Assets Value: " + ItemDic[DrpDwnYear][1] + "\n Fund NAV Value: " + ItemDic[DrpDwnYear][2] + "\n Ben NAV Value: " + str(ItemDic[DrpDwnYear][3]) + "\n Ben Ownership Value: " + ItemDic[DrpDwnYear][4])
                                TestResultStatus.append("Pass")

                                BenNavCout=driver.find_elements_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div[5]/div/div[2]/div[2]/div/div/div/div[2]/div/div/table/tbody/tr")
                                BenNAVFloat=0.0
                                for ii7 in range(1,len(BenNavCout)+1):
                                    print("ii7  "+str(ii7))
                                    BenNAV = driver.find_element_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div[5]/div/div[2]/div[2]/div/div/div/div[2]/div/div/table/tbody/tr["+str(ii7)+"]/td[6]/p").text
                                    BenNAV = BenNAV.replace(" ", "")
                                    BenNAV = re.sub(r'[?|$|!|,|-]', r'', BenNAV)
                                    if "_" in BenNAV:
                                        BenNAV="0"
                                    BenNAVFloat =BenNAVFloat+ round(float(BenNAV))

                                if ItemDic[DrpDwnYear][3]==BenNAVFloat:
                                    print("BenNAV matched ")
                                    TestResult.append("Ben NAV matched. Value found in main table [ "+str(ItemDic[DrpDwnYear][3])+" ] and  in GP Reported Ben Details section [ "+str(BenNAVFloat)+" ]" )
                                    TestResultStatus.append("Pass")
                                else:
                                    print("BenNAV doesn't match ")
                                    TestResult.append("Ben NAV doesn't match. Value found in main table [ "+str(ItemDic[DrpDwnYear][3])+" ] and  in GP Reported Ben Details section [ "+str(BenNAVFloat)+" ]")
                                    TestResultStatus.append("Pass")

                                driver.find_element_by_xpath(
                                    "//div[@class='ContentLayout---content_layout']/div[5]/div/div[1]/div[1]/div/div[2]/div/div").click()
                                time.sleep(1)
                                ActionChains(driver).key_down(Keys.DOWN).perform()
                                time.sleep(1)
                                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                                try:
                                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(
                                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                    WebDriverWait(driver, LONG_TIMEOUT
                                                  ).until_not(
                                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                except TimeoutException:
                                    pass
                                break
                            else:
                                driver.find_element_by_xpath(
                                    "//div[@class='ContentLayout---content_layout']/div[5]/div/div[1]/div[1]/div/div[2]/div/div").click()
                                time.sleep(1)
                                ActionChains(driver).key_down(Keys.DOWN).perform()
                                time.sleep(1)
                                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                                try:
                                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(
                                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                    WebDriverWait(driver, LONG_TIMEOUT
                                                  ).until_not(
                                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                except TimeoutException:
                                    pass

                    for winclose in range(1,10):
                        time.sleep(1)
                        if len(driver.window_handles)>1:
                            #print("Tab Count is more than 1: "+str(len(driver.window_handles)))
                            driver.switch_to.window(driver.window_handles[1])
                            driver.close()
                        elif len(driver.window_handles)==1:
                            break
                    #driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(1)
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass

        except Exception as Mainerror:
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


