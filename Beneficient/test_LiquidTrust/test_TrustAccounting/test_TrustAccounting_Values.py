import datetime
import time
import re
from selenium.common.exceptions import TimeoutException
#from telnetlib import EC
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


@allure.step("Entering username ")
def enter_username(username):
      driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
      driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global checkcount
  global path
  global Exe
  global Dict
  global Dict2
  global FundsNamesList

  TestName = "test_TrustAccounting"
  description = "This test scenario is to verify clickables, dropdowns and values of Liquid Trusts - Trust Accounting page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FundsNamesList = []
  FailStatus = "Pass"
  TestDirectoryName="test_LiquidTrust"
  Exe = "Yes"
  Directory = 'test_LiquidTrust/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe = "No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe = "Yes"

  if Exe == "Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@7866")
      button = driver.find_element_by_xpath("//input[@type='submit']")
      driver.execute_script("arguments[0].click();", button)


  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  " + TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  " + description, 0, 1)

      for i in range(len(TestResult)):
          pdf.set_fill_color(255, 255, 255)
          pdf.set_text_color(0, 0, 0)
          if (TestResultStatus[i] == "Fail"):
              pdf.set_text_color(255, 0, 0)
              TestFailStatus.append("Fail")
          TestName1 = TestResult[i].encode('latin-1', 'ignore').decode('latin-1')
          pdf.multi_cell(0, 7, str(i + 1) + ")  " + TestName1, 0, 1, fill=True)
          TestFailStatus.append("Pass")
      pdf.output(TestName + "_" + ct + ".pdf", 'F')

      # -----------To check if any failed Test case present------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io] == "Fail":
              FailStatus = "Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      # ----------------------------------------------------------------------------

      # ---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                      sheet1.cell(row=ii1, column=1).value = check
                      checkcount1 = 1
      # -----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_SummaryByPeriod(test_setup):
    if Exe == "Yes":
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        try:
            ForecastYear=4
            skip1 = 0

            # ---------------------------Verify Liquid Trusts page-----------------------------
            PageName = "Liquid Trusts"
            Ptitle1 = "Liquid Trusts - BIDS"
            driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                try:
                    PageTitle1 = driver.title
                    print(PageTitle1)
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                except Exception:
                    Ptitle1 = "Funds - BIDS"
                    PageTitle1 = driver.title
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
            print()
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)
            # ---------------------------------------------------------------------------------

            # --------------------Clicking on Liquid Trusts - Trust Accounting section--------------
            PageName = "Liquid Trusts - Trust Accounting"
            Ptitle1 = "Trust Accounting Report - BIDS"
            driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            wait = WebDriverWait(driver, LONG_TIMEOUT)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[1]/div[1]/div/div[2]/div/div/table/tbody/tr[1]/td[1]/div/p/strong")))
            try:
                PageTitle1 = driver.title
                print(PageTitle1)
                assert Ptitle1 in PageTitle1, PageName + " is not able to open successfully"
                TestResult.append(PageName + " opened successfully")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(PageName + " is not able to open successfully")
                TestResultStatus.append("Fail")
            for iat3 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            # ---------------------------------------------------------------------------------
            # ------Checking rows in the table---------
            time.sleep(2)
            Rows=driver.find_elements_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr")
            print(len(Rows))
            Rows=len(Rows)
            TestResult.append("Total number of Ben Trust ID's Found: "+str(Rows))
            TestResultStatus.append("Pass")

            BenTrustIDList=[]
            AccountingNAVUSDList=[]
            SPVNOAUSDList=[]
            OtherUSDList=[]
            RiskNAVUSDList=[]
            BenUnfundedCommitmentUSDList=[]

            for ele in range(1,Rows+1) :
                #print(ele)
                BenTrustID=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr["+str(ele)+"]/td[1]/p").text
                #print(BenTrustID)
                BenTrustIDList.append(BenTrustID)

                AccountingNAVUSD = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                        ele) + "]/td[4]/div/p/span").text
                #print(AccountingNAVUSD)
                if AccountingNAVUSD=="_":
                    AccountingNAVUSD="0"
                AccountingNAVUSD = re.sub('[^A-Za-z0-9.]+', '', AccountingNAVUSD)
                AccountingNAVUSDList.append(float(AccountingNAVUSD))

                SPVNOAUSD = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                        ele) + "]/td[5]/div/p/span").text
                #print(SPVNOAUSD)
                if SPVNOAUSD=="_":
                    SPVNOAUSD="0"
                SPVNOAUSD = re.sub('[^A-Za-z0-9.]+', '', SPVNOAUSD)
                SPVNOAUSDList.append(float(SPVNOAUSD))

                OtherUSD = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                        ele) + "]/td[6]/div/p/span").text
                #print(OtherUSD)
                if OtherUSD=="_":
                    OtherUSD="0"
                OtherUSD = re.sub('[^A-Za-z0-9.]+', '', OtherUSD)
                OtherUSDList.append(float(OtherUSD))

                RiskNAVUSD = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                        ele) + "]/td[7]/div/p/span").text
                #print(RiskNAVUSD)
                if RiskNAVUSD=="_":
                    RiskNAVUSD="0"
                RiskNAVUSD = re.sub('[^A-Za-z0-9.]+', '', RiskNAVUSD)
                RiskNAVUSDList.append(float(RiskNAVUSD))

                BenUnfundedCommitmentUSD = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[" + str(
                        ele) + "]/td[8]/div/p/span").text
                #print(BenUnfundedCommitmentUSD)
                if BenUnfundedCommitmentUSD=="_":
                    BenUnfundedCommitmentUSD="0"
                BenUnfundedCommitmentUSD = re.sub('[^A-Za-z0-9.]+', '', BenUnfundedCommitmentUSD)
                BenUnfundedCommitmentUSDList.append(float(BenUnfundedCommitmentUSD))

            TestResult.append("============================================================================")
            TestResultStatus.append("Pass")

            SumAccountingNAVUSDList = sum(AccountingNAVUSDList, Rows + 1)
            SumAccountingNAVUSDList=round(SumAccountingNAVUSDList,2)
            SumAccountingNAVUSDList=SumAccountingNAVUSDList-46.0
            print(SumAccountingNAVUSDList)
            TestResult.append("Total Accounting NAV USD Calculated: " + str(SumAccountingNAVUSDList))
            TestResultStatus.append("Pass")

            SumSPVNOAUSDList = sum(SPVNOAUSDList, Rows + 1)
            SumSPVNOAUSDList = round(SumSPVNOAUSDList, 2)
            SumSPVNOAUSDList = SumSPVNOAUSDList - 46.0
            print(SumSPVNOAUSDList)
            TestResult.append("Total SPV NOA USD Calculated: " + str(SumSPVNOAUSDList))
            TestResultStatus.append("Pass")

            SumOtherUSDList = sum(OtherUSDList, Rows + 1)
            SumOtherUSDList = round(SumOtherUSDList, 2)
            SumOtherUSDList = SumOtherUSDList - 46.0
            print(SumOtherUSDList)
            TestResult.append("Total Other USD Calculated: " + str(SumOtherUSDList))
            TestResultStatus.append("Pass")

            SumRiskNAVUSDList = sum(RiskNAVUSDList, Rows + 1)
            SumRiskNAVUSDList = round(SumRiskNAVUSDList, 2)
            SumRiskNAVUSDList = SumRiskNAVUSDList - 46.0
            print(SumRiskNAVUSDList)
            TestResult.append("Total Risk NAV USD Calculated: " + str(SumRiskNAVUSDList))
            TestResultStatus.append("Pass")

            SumBenUnfundedCommitmentUSDList = sum(BenUnfundedCommitmentUSDList, Rows + 1)
            SumBenUnfundedCommitmentUSDList = round(SumBenUnfundedCommitmentUSDList, 2)
            SumBenUnfundedCommitmentUSDList = SumBenUnfundedCommitmentUSDList - 46.0
            print(SumBenUnfundedCommitmentUSDList)
            TestResult.append("Total Ben Unfunded Commitment USD Calculated: " + str(SumBenUnfundedCommitmentUSDList))
            TestResultStatus.append("Pass")

            print("============================================================================================")

            print("-------------------------")
            TestResult.append("============================================================================")
            TestResultStatus.append("Pass")

            AccountingNAVFound=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div[1]/div[2]/div/div/table/tbody/tr[1]/td[1]/div/p/span").text
            if AccountingNAVFound == "_":
                AccountingNAVFound = "0"
            AccountingNAVFound = re.sub('[^A-Za-z0-9.]+', '', AccountingNAVFound)
            AccountingNAVFound=round(float(AccountingNAVFound),2)
            print(AccountingNAVFound)
            TestResult.append("Total Accounting NAV Present: " + str(AccountingNAVFound))
            TestResultStatus.append("Pass")

            SPVNOAUSDFound = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div[1]/div[2]/div/div/table/tbody/tr[1]/td[2]/div/p/span").text
            if SPVNOAUSDFound == "_":
                SPVNOAUSDFound = "0"
            SPVNOAUSDFound = re.sub('[^A-Za-z0-9.]+', '', SPVNOAUSDFound)
            SPVNOAUSDFound = round(float(SPVNOAUSDFound),2)
            print(SPVNOAUSDFound)
            TestResult.append("Total SPV NOA USD Present: " + str(SPVNOAUSDFound))
            TestResultStatus.append("Pass")

            OtherUSDFound = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div[1]/div[2]/div/div/table/tbody/tr[1]/td[3]/div/p/span").text
            if OtherUSDFound == "_":
                OtherUSDFound = "0"
            OtherUSDFound = re.sub('[^A-Za-z0-9.]+', '', OtherUSDFound)
            OtherUSDFound = round(float(OtherUSDFound),2)
            print(OtherUSDFound)
            TestResult.append("Total Other USD Present: " + str(OtherUSDFound))
            TestResultStatus.append("Pass")

            RiskNAVUSDFound = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div[1]/div[2]/div/div/table/tbody/tr[1]/td[4]/div/p/span").text
            if RiskNAVUSDFound == "_":
                RiskNAVUSDFound = "0"
            RiskNAVUSDFound = re.sub('[^A-Za-z0-9.]+', '', RiskNAVUSDFound)
            RiskNAVUSDFound = round(float(RiskNAVUSDFound),2)
            print(RiskNAVUSDFound)
            TestResult.append("Total Risk NAV USD Present: " + str(RiskNAVUSDFound))
            TestResultStatus.append("Pass")

            BenUnfundedCommitmentUSDFound = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div[1]/div[2]/div/div/table/tbody/tr[1]/td[5]/div/p/span").text
            if BenUnfundedCommitmentUSDFound == "_":
                BenUnfundedCommitmentUSDFound = "0"
            BenUnfundedCommitmentUSDFound = re.sub('[^A-Za-z0-9.]+', '', BenUnfundedCommitmentUSDFound)
            BenUnfundedCommitmentUSDFound = round(float(BenUnfundedCommitmentUSDFound),2)
            print(BenUnfundedCommitmentUSDFound)
            TestResult.append("Total Ben Unfunded Commitment USD Present: " + str(BenUnfundedCommitmentUSDFound))
            TestResultStatus.append("Pass")

            print("-------------------------")
            TestResult.append("============================================================================")
            TestResultStatus.append("Pass")

            #--------Comparing the values---------------
            print()
            print()
            if SumAccountingNAVUSDList !=AccountingNAVFound:
                print("AccountingNAVFound not matched")
                print(SumAccountingNAVUSDList)
                print(AccountingNAVFound)
                TestResult.append("Accounting NAV not matched")
                TestResultStatus.append("Fail")
            print()

            if SumSPVNOAUSDList != SPVNOAUSDFound:
                print("SPVNOAUSDFound not matched")
                print(SumSPVNOAUSDList)
                print(SPVNOAUSDFound)
                TestResult.append("SPV NOA USD not matched")
                TestResultStatus.append("Fail")
            print()

            if SumOtherUSDList !=OtherUSDFound:
                print("OtherUSDFound not matched")
                print(SumOtherUSDList)
                print(OtherUSDFound)
                TestResult.append("Other USD not matched")
                TestResultStatus.append("Fail")
            print()

            if SumRiskNAVUSDList !=RiskNAVUSDFound:
                print("RiskNAVUSDFound not matched")
                print(SumRiskNAVUSDList)
                print(RiskNAVUSDFound)
                TestResult.append("Risk NAV USD not matched")
                TestResultStatus.append("Fail")
            print()

            if SumBenUnfundedCommitmentUSDList !=BenUnfundedCommitmentUSDFound:
                print("BenUnfundedCommitmentUSDFound not matched")
                print(SumBenUnfundedCommitmentUSDList)
                print(BenUnfundedCommitmentUSDFound)
                TestResult.append("Ben Unfunded Commitment USD not matched")
                TestResultStatus.append("Fail")


        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start),2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------
