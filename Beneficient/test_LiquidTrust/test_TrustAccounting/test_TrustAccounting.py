import datetime
import time
from telnetlib import EC
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


@allure.step("Entering username ")
def enter_username(username):
      driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
      driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global checkcount
  global path
  global Exe
  global Dict
  global Dict2
  global FundsNamesList

  TestName = "test_TrustAccounting"
  description = "This test scenario is to verify clickables, dropdowns and values of Liquid Trusts - Trust Accounting page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FundsNamesList = []
  FailStatus = "Pass"
  TestDirectoryName="test_LiquidTrust"
  Exe = "Yes"
  Directory = 'test_LiquidTrust/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe = "No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe = "Yes"

  if Exe == "Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      button = driver.find_element_by_xpath("//input[@type='submit']")
      driver.execute_script("arguments[0].click();", button)


  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  " + TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  " + description, 0, 1)

      for i in range(len(TestResult)):
          pdf.set_fill_color(255, 255, 255)
          pdf.set_text_color(0, 0, 0)
          if (TestResultStatus[i] == "Fail"):
              pdf.set_text_color(255, 0, 0)
              TestFailStatus.append("Fail")
          TestName1 = TestResult[i].encode('latin-1', 'ignore').decode('latin-1')
          pdf.multi_cell(0, 7, str(i + 1) + ")  " + TestName1, 0, 1, fill=True)
          TestFailStatus.append("Pass")
      pdf.output(TestName + "_" + ct + ".pdf", 'F')

      # -----------To check if any failed Test case present------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io] == "Fail":
              FailStatus = "Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      # ----------------------------------------------------------------------------

      # ---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                      sheet1.cell(row=ii1, column=1).value = check
                      checkcount1 = 1
      # -----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_SummaryByPeriod(test_setup):
    if Exe == "Yes":
        try:
            ForecastYear=4
            skip1 = 0

            # ---------------------------Verify Liquid Trusts page-----------------------------
            PageName = "Liquid Trusts"
            Ptitle1 = "Liquid Trusts - BIDS"
            driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
            start = time.time()
            for iat1 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(1)
            try:
                try:
                    PageTitle1 = driver.title
                    print(PageTitle1)
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                except Exception:
                    Ptitle1 = "Funds - BIDS"
                    PageTitle1 = driver.title
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
            print()
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)
            # ---------------------------------------------------------------------------------

            # --------------------Clicking on Liquid Trusts - Trust Accounting section--------------
            PageName = "Liquid Trusts - Trust Accounting"
            Ptitle1 = "Trust Accounting Report - BIDS"
            driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
            start = time.time()
            for iat2 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            wait = WebDriverWait(driver, 200)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[1]/div[1]/div/div[2]/div/div/table/tbody/tr[1]/td[1]/div/p/strong")))
            try:
                PageTitle1 = driver.title
                print(PageTitle1)
                assert Ptitle1 in PageTitle1, PageName + " is not able to open successfully"
                TestResult.append(PageName + " opened successfully")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(PageName + " is not able to open successfully")
                TestResultStatus.append("Fail")
            for iat3 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            print()
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)
            # ---------------------------------------------------------------------------------
            # ------Checking Ben Reporting Period dropdown---------
            time.sleep(2)
            Text1 = "Ben Reporting Period"
            Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[1]/div[1]/div/div[2]/div/div/table/tbody/tr[1]/td[1]/div/p/strong").text
            try:
                assert Text1 in Element1, "Ben Reporting Period dropdown at Liquid Trusts - Trust Accounting is not present"
                TestResult.append("Ben Reporting Period dropdown at Liquid Trusts - Trust Accounting is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append("Ben Reporting Period dropdown at Liquid Trusts - Trust Accounting is not present")
                TestResultStatus.append("Fail")

            # ------Checking Total Accounting NAV (USD)---------
            Text1 = "Total Accounting NAV (USD)"
            Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div/div[2]/div/div/table/thead/tr/th[1]/div").text
            try:
                assert Text1 in Element1, "Total Accounting NAV (USD) at Liquid Trusts - Trust Accounting is not present"
                TestResult.append("Total Accounting NAV (USD) at Liquid Trusts - Trust Accounting is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append("Total Accounting NAV (USD) at Liquid Trusts - Trust Accounting is not present")
                TestResultStatus.append("Fail")

            # ------Checking Total SPV NOA (USD)---------
            Text1 = "Total SPV NOA (USD)"
            Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div/div[2]/div/div/table/thead/tr/th[2]/div").text
            try:
                assert Text1 in Element1, "Total SPV NOA (USD) at Liquid Trusts - Trust Accounting is not present"
                TestResult.append("Total SPV NOA (USD) at Liquid Trusts - Trust Accounting is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append("Total SPV NOA (USD) at Liquid Trusts - Trust Accounting is not present")
                TestResultStatus.append("Fail")

            # ------Checking Total Other (USD)---------
            Text1 = "Total Other (USD)"
            Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div/div[2]/div/div/table/thead/tr/th[3]/div").text
            try:
                assert Text1 in Element1, "Total Other (USD) at Liquid Trusts - Trust Accounting is not present"
                TestResult.append("Total Other (USD) at Liquid Trusts - Trust Accounting is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append("Total Other (USD) at Liquid Trusts - Trust Accounting is not present")
                TestResultStatus.append("Fail")

            # ------Checking Total Risk NAV (USD)---------
            Text1 = "Total Risk NAV (USD)"
            Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div/div[2]/div/div/table/thead/tr/th[4]/div").text
            try:
                assert Text1 in Element1, "Total Risk NAV (USD) at Liquid Trusts - Trust Accounting is not present"
                TestResult.append("Total Risk NAV (USD) at Liquid Trusts - Trust Accounting is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append("Total Risk NAV (USD) at Liquid Trusts - Trust Accounting is not present")
                TestResultStatus.append("Fail")

            # ------Checking Total Ben Unfunded Commitment (USD)---------
            Text1 = "Total Ben Unfunded Commitment (USD)"
            Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div/div[2]/div/div/table/thead/tr/th[5]/div").text
            try:
                assert Text1 in Element1, "Total Ben Unfunded Commitment (USD) at Liquid Trusts - Trust Accounting is not present"
                TestResult.append("Total Ben Unfunded Commitment (USD) at Liquid Trusts - Trust Accounting is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append("Total Ben Unfunded Commitment (USD) at Liquid Trusts - Trust Accounting is not present")
                TestResultStatus.append("Fail")

            # ---------------loop for Columns in table for Trust Accounting----------
            ItemList1 = ["Ben Trust ID", "Trust Name", "Collective Trust", "Accounting NAV (USD)",
                         "SPV NOA (USD)", "Other (USD)", "Risk NAV (USD)", "Ben Unfunded Commitment (USD)"]
            for ii1 in range(len(ItemList1)):
                Text1 = ItemList1[ii1]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[3]/div[2]/div/div/div[2]/table/thead/tr[1]/th[" + str(
                            ii1 + 1) + "]/div").text
                except Exception as e1:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column is not present in table"
                    TestResult.append(
                        Text1 + " column is present in table")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    TestResult.append(
                        Text1 + " column is not present in table")
                    TestResultStatus.append("Fail")

        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start),2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            TestResult.append(stringMainerror)
            TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------
