import datetime
import math
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


@allure.step("Entering username ")
def enter_username(username):
      driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
      driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global checkcount
  global path
  global Exe
  global Dict
  global Dict2
  global FundsNamesList

  TestName = "test_SummaryByPeriod"
  description = "This test scenario is to verify clickables, dropdowns and values of Capital Activity Summary By Period page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FundsNamesList = []
  Dict = {}
  Dict2 = {}

  FailStatus = "Pass"
  TestDirectoryName="test_LiquidTrust"
  Exe = "Yes"
  Directory = 'test_LiquidTrust/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe = "No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe = "Yes"

  if Exe == "Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@7866")
      button = driver.find_element_by_xpath("//input[@type='submit']")
      driver.execute_script("arguments[0].click();", button)


  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  " + TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  " + description, 0, 1)

      for i in range(len(TestResult)):
          pdf.set_fill_color(255, 255, 255)
          pdf.set_text_color(0, 0, 0)
          if (TestResultStatus[i] == "Fail"):
              pdf.set_text_color(255, 0, 0)
              TestFailStatus.append("Fail")
          TestName1 = TestResult[i].encode('latin-1', 'ignore').decode('latin-1')
          pdf.multi_cell(0, 7, str(i + 1) + ")  " + TestName1, 0, 1, fill=True)
          TestFailStatus.append("Pass")
      pdf.output(TestName + "_" + ct + ".pdf", 'F')

      # -----------To check if any failed Test case present------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io] == "Fail":
              FailStatus = "Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      # ----------------------------------------------------------------------------

      # ---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                      sheet1.cell(row=ii1, column=1).value = check
                      checkcount1 = 1
      # -----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_SummaryByPeriod(test_setup):
    if Exe == "Yes":
        try:
            ForecastYear=4
            skip1 = 0
            SHORT_TIMEOUT = 5
            LONG_TIMEOUT = 400
            LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"

            # ---------------------------Verify Liquid Trusts page-----------------------------
            PageName = "Liquid Trusts"
            Ptitle1 = "Liquid Trusts - BIDS"
            driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                try:
                    PageTitle1 = driver.title
                    print(PageTitle1)
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                except Exception:
                    Ptitle1 = "Funds - BIDS"
                    PageTitle1 = driver.title
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)
            print()
            # ---------------------------------------------------------------------------------

            # --------------------Clicking on Capital Activity Summary By Period section--------------
            PageName = "Capital Activity Summary By Period"
            Ptitle1 = "COR_CapitalActivityPeriodTemplate - BIDS"
            driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                PageTitle1 = driver.title
                print(PageTitle1)
                assert Ptitle1 in PageTitle1, PageName + " is not able to open successfully"
                TestResult.append(PageName + " opened successfully")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(PageName + " is not able to open successfully")
                TestResultStatus.append("Fail")
            for iat3 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(1)
            # ---------------------------------------------------------------------------------

            # ---------------------------------------------------View by Investments--------------------------------------------------
            TestResult.append(
                "---------------Now Verifying content on the page [ View by Funds]--------------")
            TestResultStatus.append("Pass")
            # -------------------------------------------------------------------------------------------------------------------------
            #------------------------------Verifying elemenets inside Summary By Period section---------------------------
            try:
                time.sleep(2)
                driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[1]/div[2]/div/div[2]").click()
                TestResult.append("View by Funds at Capital Call & Distribution Activity button clicked successfully")
                TestResultStatus.append("Pass")

                #driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[1]/div[2]/div/div[2]").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                time.sleep(1)

            #     # ------Checking View by Funds page---------
                Text1="Filter By"
                Element1=driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div[1]/div/div[2]/div/div[1]/div[2]/div/div[1]/div/div[2]/div/p/strong").text
                try:
                    assert Text1 in Element1, "Capital Call & Distribution Activity is not able to open"
                    TestResult.append("Capital Call & Distribution Activity page opened successfully after clicking [ View by Funds ]")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append("Capital Call & Distribution Activity page not able to open after clicking [ View by Funds ]")
                    TestResultStatus.append("Fail")

            #     # ------Checking Filter By drop down---------
                Text1 = "Both"
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div[1]/div/div[2]/div/div[1]/div[2]/div/div[2]/div/div[2]/div/div/span").text
                try:
                    assert Text1 in Element1, "Filter By drop down at Capital Call & Distribution Activity is not present"
                    TestResult.append("Filter By drop down at Capital Call & Distribution Activity is present")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append("Filter By drop down at Capital Call & Distribution Activity is not present")
                    TestResultStatus.append("Fail")

            #     # ------Checking Grand Total---------
                Text1 = "Grand Total"
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div/div/table/tbody/tr[3]/td[1]/div/p").text
                try:
                    assert Text1 in Element1, "Grand Total at Capital Call & Distribution Activity under Settlement Totals section is not present"
                    TestResult.append("Grand Total at Capital Call & Distribution Activity under Settlement Totals section is present")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append("Grand Total at Capital Call & Distribution Activity under Settlement Totals section is not present")
                    TestResultStatus.append("Fail")

            #     # ------Checking Funds present---------
                Text1 = "Fund Name"
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[1]/div[2]/table/thead/tr/th[5]/div").text
                try:
                    assert Text1 in Element1, "Funds Grid at Capital Call & Distribution Activity is not present"
                    TestResult.append(
                        "Funds Grid at Capital Call & Distribution Activity is present")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append(
                        "Funds Grid at Capital Call & Distribution Activity is not present")
                    TestResultStatus.append("Fail")

            #     # ---------------loop for Columns in table for Funds View----------
                ItemList1 = ["Transaction Type", "Issue Date", "Due Date of Cap Call or Dist", "BEN Fund ID","Fund Name", "LP Name", "Local Ccy", "Distribution","Capital Call", "Total", "Impact to Unfunded", "Functional Ccy", "Settlement Date","Settlement FX","Dist.","CC","Other","Settlement Amount (Func. Ccy)","Pending Transactions (Local Ccy)","Reviewed By","Date Reviewed","Reviewer Notes"]
                for ii1 in range(len(ItemList1)):
                    Text1 = ItemList1[ii1]
                    try:
                        Element1 = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div[2]/div/div/div[2]/table/thead/tr[1]/th[" + str(
                                ii1 + 1) + "]/div").text
                    except Exception:
                        pass
                    try:
                        assert Text1 in Element1, Text1 + " column is not present in table"
                        TestResult.append(
                            Text1 + " column is present in table")
                        TestResultStatus.append("Pass")
                    except Exception as e1:
                        print(e1)
                        TestResult.append(
                            Text1 + " column is not present in table")
                        TestResultStatus.append("Fail")


            #     # ------Compare Grand Total at Settlement Totals with Settlement Amount (Func. Ccy) value---------
                Text1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div/div/table/tbody/tr[3]/td[2]/div/p/strong").text
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[1]/div[2]/table/tbody/tr[position()=last()]/td[18]/div/p/strong").text
                try:
                    assert Text1 in Element1, "Grand Total at Settlement Totals is not matching with Settlement Amount (Func. Ccy) value"
                    TestResult.append(
                        "Grand Total at Settlement Totals matched with Settlement Amount (Func. Ccy) value")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append(
                        "Grand Total at Settlement Totals is not matching with Settlement Amount (Func. Ccy) value")
                    TestResultStatus.append("Fail")
            #     #-----------------------------------------------------------------------------------
            except Exception as e1:
                print(e1)
                TestResult.append("View by Funds button at Capital Call & Distribution Activity is not able to click")
                TestResultStatus.append("Fail")

            # ---------------------------------------------------View by Investments--------------------------------------------------
            TestResult.append(
                "---------------Now Verifying content on the page [ View by Investment]--------------")
            TestResultStatus.append("Pass")
            #-------------------------------------------------------------------------------------------------------------------------
            try:
                time.sleep(2)
                driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[1]/div[2]/div/div[1]").click()
                TestResult.append("View by Investments at Capital Call & Distribution Activity button clicked successfully")
                TestResultStatus.append("Pass")
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                time.sleep(1)

                # ---------------Ben Reporting Period label present in page----------
                Text1 = "Ben Reporting Period"
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/p/strong").text
                try:
                    assert Text1 in Element1, "Ben Reporting Period label is not present"
                    TestResult.append(
                        "Ben Reporting Period label is present")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append(
                        "Ben Reporting Period label is not present")
                    TestResultStatus.append("Fail")

                # ---------------Checking Ben Reporting Period dropdown----------
                Text1 = "Ben Reporting Period"
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout ContentLayout---padding_less']/div/div[2]/div/div/div/div[2]/div/p/strong").text
                try:
                    assert Text1 in Element1, "Ben Reporting Period dropdown (View by Investements) at Capital Call & Distribution Activity is not present"
                    TestResult.append("Ben Reporting Period dropdown (View by Investements) at Capital Call & Distribution Activity is present")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append("Ben Reporting Period dropdown (View by Investements) at Capital Call & Distribution Activity is not present")
                    TestResultStatus.append("Fail")

                # ---------------Filter present in page----------
                Text1 = "Filter"
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[2]/div/div/div/div/div/div/div[1]/div").text
                try:
                    assert Text1 in Element1, "Filter label is not present"
                    TestResult.append(
                        "Filter label is present")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append(
                        "Filter label is not present")
                    TestResultStatus.append("Fail")

                # ---------------Checking Funds listing table----------
                time.sleep(3)
                Text1 = "Fund"
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/thead/tr[1]/th[2]/div").text
                print(Element1)
                try:
                    assert Text1 in Element1, "Funds listing table (View by Investements) at Capital Call & Distribution Activity is not present"
                    TestResult.append(
                        "Funds listing table (View by Investements) at Capital Call & Distribution Activity is present")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append(
                        "Funds listing table (View by Investements) at Capital Call & Distribution Activity is not present")
                    TestResultStatus.append("Fail")

                # ---------------loop for Columns in table for Investment View----------
                ItemList2=["Notice Date","Fund\u2007Name⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀","Fund ID","Investment\u2007Name⠀⠀⠀⠀⠀⠀⠀","Investment ID", "Currency","Settled/Unsettled","Distribution","Holdback or Capital Call","Total","FX Rate","USD Total","Liquid Trust"]
                for ii2 in range(len(ItemList2)):
                    Text2 = ItemList2[ii2]
                    try:
                        Element2 = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/thead/tr[1]/th["+str(ii2+1)+"]/div").text
                    except Exception:
                        pass

                    try:
                        assert Text2 in Element2, Text2+" column is not present in table"
                        TestResult.append(
                            Text2+" column is present in table")
                        TestResultStatus.append("Pass")
                    except Exception as e1:
                        print(e1)
                        TestResult.append(
                            Text2+" column is not present in table")
                        TestResultStatus.append("Fail")

            except Exception as e1:
                print(e1)
                TestResult.append("View by Investments button at Capital Call & Distribution Activity is not able to click")
                TestResultStatus.append("Fail")

        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start),2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------
