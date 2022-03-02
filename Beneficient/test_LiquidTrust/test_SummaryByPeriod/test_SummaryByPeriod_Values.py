import datetime
import math
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

@allure.step("Entering username ")
def enter_username(username):
      driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
      driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global checkcount
  global path
  global Exe
  global Dict
  global Dict2
  global FundsNamesList

  TestName = "test_SummaryByPeriod_Values"
  description = "This test scenario is to match SOS [Funds/Investment] values with Distribution values in Investments and Funds table at Capital Activity Summary By Period page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FundsNamesList = []
  Dict = {}
  Dict2 = {}

  FailStatus = "Pass"
  TestDirectoryName="test_LiquidTrust"
  Exe = "Yes"
  Directory = 'test_LiquidTrust/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe = "No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe = "Yes"

  if Exe == "Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@7866")
      button = driver.find_element_by_xpath("//input[@type='submit']")
      driver.execute_script("arguments[0].click();", button)


  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  " + TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  " + description, 0, 1)

      for i in range(len(TestResult)):
          pdf.set_fill_color(255, 255, 255)
          pdf.set_text_color(0, 0, 0)
          if (TestResultStatus[i] == "Fail"):
              pdf.set_text_color(255, 0, 0)
              TestFailStatus.append("Fail")
          TestName1 = TestResult[i].encode('latin-1', 'ignore').decode('latin-1')
          pdf.multi_cell(0, 7, str(i + 1) + ")  " + TestName1, 0, 1, fill=True)
          TestFailStatus.append("Pass")
      pdf.output(TestName + "_" + ct + ".pdf", 'F')

      # -----------To check if any failed Test case present------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io] == "Fail":
              FailStatus = "Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      # ----------------------------------------------------------------------------

      # ---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                      sheet1.cell(row=ii1, column=1).value = check
                      checkcount1 = 1
      # -----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_SummaryByPeriod(test_setup):
    if Exe == "Yes":
        try:
            ForecastYear=4
            skip1 = 0
            SHORT_TIMEOUT = 5
            LONG_TIMEOUT = 400
            LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"

            # ---------------------------Verify Liquid Trusts page-----------------------------
            PageName = "Liquid Trusts"
            Ptitle1 = "Liquid Trusts - BIDS"
            driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                try:
                    PageTitle1 = driver.title
                    print(PageTitle1)
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                except Exception:
                    Ptitle1 = "Funds - BIDS"
                    PageTitle1 = driver.title
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)
            print()
            # ---------------------------------------------------------------------------------

            # --------------------Clicking on Capital Activity Summary By Period section--------------
            PageName = "Capital Activity Summary By Period"
            Ptitle1 = "COR_CapitalActivityPeriodTemplate - BIDS"
            driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                PageTitle1 = driver.title
                print(PageTitle1)
                assert Ptitle1 in PageTitle1, PageName + " is not able to open successfully"
                TestResult.append(PageName + " opened successfully")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(PageName + " is not able to open successfully")
                TestResultStatus.append("Fail")
            for iat3 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(1)
            # ---------------------------------------------------------------------------------

            # # -----------------------------------------View by Funds---------------------------
            # TestResult.append(
            #     "---------------Now Verifying content on the page [ View by Funds]--------------")
            # TestResultStatus.append("Pass")
            # # -------------------------------------------------------------------------------------------------------------------------
            # #------------------------------Verifying elemenets inside Summary By Period section---------------------------
            # try:
            #     time.sleep(2)
            #     driver.find_element_by_xpath(
            #         "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[1]/div[2]/div/div[2]").click()
            #     TestResult.append("View by Funds at Capital Call & Distribution Activity button clicked successfully")
            #     TestResultStatus.append("Pass")
            #     try:
            #         WebDriverWait(driver, SHORT_TIMEOUT
            #                       ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #         WebDriverWait(driver, LONG_TIMEOUT
            #                       ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #     except TimeoutException:
            #         pass
            #     time.sleep(1)
            #
            # except Exception as e1:
            #     print(e1)
            #     TestResult.append("[ View by Funds ] button at Capital Call & Distribution Activity is not able to click")
            #     TestResultStatus.append("Fail")

            # ---------------------------------------------------View by Investments--------------------------------------------------
            TestResult.append(
                "---------------Now Verifying content on the page [ View by Investment]--------------")
            TestResultStatus.append("Pass")
            #-------------------------------------------------------------------------------------------------------------------------
            try:
                time.sleep(2)
                driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[1]/div[2]/div/div[1]").click()
                TestResult.append("View by Investments at Capital Call & Distribution Activity button clicked successfully")
                TestResultStatus.append("Pass")
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                time.sleep(1)


                #---------------------------------Fetching all Funds from Table values---------------------------
                FundNameInvList=[]
                DistributionInvList=[]
                LiquidTrustInvList=[]

                try:
                    TotalItem=driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[2]/div[2]/div/p").text
                    substr = "of"
                    count1 = TotalItem.index(substr)

                    TotalItemAfterOf=TotalItem[count1+3]+TotalItem[count1 +4]
                    TotalItemBeforeOf = TotalItem[count1 - 3] + TotalItem[count1 - 2]
                    #print("TotalItemAfterOf "+TotalItemAfterOf)
                    #print("TotalItemBeforeOf " + TotalItemBeforeOf)

                    IterateNo=int(TotalItemAfterOf)/int(TotalItemBeforeOf)
                    #print(str(float(IterateNo)))
                    IterateNo=math.ceil(float(IterateNo))
                    #print(IterateNo)
                    print()
                except Exception:
                    pass

                for ii5 in range(1, IterateNo+1):
                    print("ii5 "+str(ii5))
                    RowsInv=driver.find_elements_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/tbody/tr")
                    print("RowsInv "+str(ii5)+" "+str(len(RowsInv)))

                    for ii3 in range(1, len(RowsInv)):
                        FundNameInvText = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/tbody/tr["+str(ii3)+"]/td[2]/div/p/a/span").text
                        print(FundNameInvText)
                        FundNameInvList.append(FundNameInvText)

                        try:
                            DistributionInvText= driver.find_element_by_xpath(
                                    "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/tbody/tr["+str(ii3)+"]/td[8]/div/p/span/a/span").text
                        except Exception:
                            DistributionInvText = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/tbody/tr[" + str(
                                    ii3) + "]/td[8]/div/p/span/span/span").text
                        if DistributionInvText in "_":
                            DistributionInvText="0"
                        DistributionInvList.append(DistributionInvText)
                        print(DistributionInvText)

                        LiquidTrustInvText=driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/tbody/tr["+str(ii3)+"]/td[13]/div/p/a").text
                        LiquidTrustInvList.append(LiquidTrustInvText)
                        print(LiquidTrustInvText)
                    #print(str(len(FundNameInvList)))

                    if ii5>1 and ii5<IterateNo:
                        driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[2]/div[2]/div/p/a[2]").click()
                    elif ii5==IterateNo:
                        pass
                    else:
                        driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[2]/div[2]/div/p/a").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(2)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                        if bool1 == True:
                            ErrorFound1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                            print(ErrorFound1)
                            driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                            TestResult.append("Forward icon in pagination is not able to open on click\n" + ErrorFound1)
                            TestResultStatus.append("Fail")
                            bool1 = False
                    except Exception:
                        try:
                            time.sleep(2)
                            bool2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                            if bool2 == True:
                                ErrorFound2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                print(ErrorFound2)
                                TestResult.append("Forward icon in pagination is not able to open on click\n" + ErrorFound2)
                                TestResultStatus.append("Fail")
                                bool2 = False
                        except Exception:
                            pass
                        pass
                for p in range(len(FundNameInvList)):
                    Dict[FundNameInvList[p]+LiquidTrustInvList[p]] = DistributionInvList[p]
                print(Dict)

                #------------Returning back to Capital Activity Summary By Period page-------------------------
                PageName = "Liquid Trusts"
                driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                try:
                    time.sleep(2)
                    bool1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                    if bool1 == True:
                        ErrorFound1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                        print(ErrorFound1)
                        driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                        TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                        TestResultStatus.append("Fail")
                        bool1 = False
                        driver.close()
                except Exception:
                    try:
                        time.sleep(2)
                        bool2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                        if bool2 == True:
                            ErrorFound2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                            print(ErrorFound2)
                            TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                            TestResultStatus.append("Fail")
                            bool2 = False
                            driver.close()
                    except Exception:
                        pass
                    pass

                PageName = "Capital Activity Summary By Period"
                driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                try:
                    time.sleep(2)
                    bool1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                    if bool1 == True:
                        ErrorFound1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                        print(ErrorFound1)
                        driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                        TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                        TestResultStatus.append("Fail")
                        bool1 = False
                        driver.close()
                except Exception:
                    try:
                        time.sleep(2)
                        bool2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                        if bool2 == True:
                            ErrorFound2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                            print(ErrorFound2)
                            TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                            TestResultStatus.append("Fail")
                            bool2 = False
                            driver.close()
                    except Exception:
                        pass
                    pass

                #-------------------Checking and comparing All Funds Cont and Dist values--------------------------------
                CheckedFund=[]
                FundIterate=1
                for ii4 in range(len(FundNameInvList)):
                    print()
                    #print("ii4 is " + str(ii4))
                    print(FundNameInvList[ii4])
                    #print("FundIterate "+str(FundIterate))
                    main_window = driver.current_window_handle

                    #print("ii4/int(TotalItemBeforeOf) is "+str(ii4/int(TotalItemBeforeOf)))
                    if ii4/int(TotalItemBeforeOf)>=1 and ii4/int(TotalItemBeforeOf)<2 :
                        #print("111")
                        if ii4/int(TotalItemBeforeOf)==1:
                            FundIterate=1
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[2]/div[2]/div/p/a").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(
                                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            WebDriverWait(driver, LONG_TIMEOUT
                                                  ).until_not(
                                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        try:
                            time.sleep(2)
                            bool1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                            if bool1 == True:
                                ErrorFound1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                print(ErrorFound1)
                                driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                TestResult.append("Forward icon in pagination is not able to open on click\n" + ErrorFound1)
                                TestResultStatus.append("Fail")
                                driver.close()
                        except Exception:
                            try:
                                time.sleep(2)
                                bool2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                if bool2 == True:
                                    ErrorFound2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                    print(ErrorFound2)
                                    TestResult.append("Forward icon in pagination is not able to open on click\n" + ErrorFound2)
                                    TestResultStatus.append("Fail")
                                    driver.close()
                            except Exception:
                                pass
                            pass
                    elif ii4/int(TotalItemBeforeOf)>=2 and (ii4/int(TotalItemBeforeOf)).is_integer():
                        #print("222")
                        if ii4/int(TotalItemBeforeOf)==2:
                            FundIterate = 1
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[2]/div[2]/div/p/a").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(
                                EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(
                                EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        try:
                            time.sleep(2)
                            bool1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                            if bool1 == True:
                                ErrorFound1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                print(ErrorFound1)
                                driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                TestResult.append("Forward icon in pagination is not able to open on click\n" + ErrorFound1)
                                TestResultStatus.append("Fail")
                                driver.close()
                        except Exception:
                            try:
                                time.sleep(2)
                                bool2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                if bool2 == True:
                                    ErrorFound2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                    print(ErrorFound2)
                                    TestResult.append("Forward icon in pagination is not able to open on click\n" + ErrorFound2)
                                    TestResultStatus.append("Fail")
                                    driver.close()
                            except Exception:
                                pass
                            pass
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[2]/div[2]/div/p/a[2]").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(
                                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            WebDriverWait(driver, LONG_TIMEOUT
                                                  ).until_not(
                                        EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        try:
                            time.sleep(2)
                            bool1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                            if bool1 == True:
                                ErrorFound1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                print(ErrorFound1)
                                driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                TestResult.append("Forward icon in pagination is not able to open on click\n" + ErrorFound1)
                                TestResultStatus.append("Fail")
                                driver.close()
                        except Exception:
                            try:
                                time.sleep(2)
                                bool2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                if bool2 == True:
                                    ErrorFound2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                    print(ErrorFound2)
                                    TestResult.append("Forward icon in pagination is not able to open on click\n" + ErrorFound2)
                                    TestResultStatus.append("Fail")
                                    driver.close()
                            except Exception:
                                pass
                            pass

                    FundToCheck = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/tbody/tr[" + str(FundIterate) + "]/td[2]/div/p/a/span").text
                    print("FundToCheck is "+FundToCheck)
                    if FundToCheck in CheckedFund:
                        print("Fund found "+FundNameInvList[ii4])
                        pass
                    else:
                        print("Fund is unchecked Fund")
                        #try:
                        #print(str(FundIterate))
                        try:
                            driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/tbody/tr["+str(FundIterate)+"]/td[2]/div/p/a").click()
                        except Exception:
                            #print("Exception")
                            button = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[4]/div/div[1]/div[2]/div/div/div[2]/table/tbody/tr["+str(FundIterate)+"]/td[2]/div/p/a")
                            driver.execute_script("arguments[0].click();", button)
                        CheckedFund.append(FundNameInvList[ii4])
                        time.sleep(3)
                        driver.switch_to.window(driver.window_handles[1])
                        time.sleep(3)
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        try:
                            time.sleep(2)
                            bool1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                            if bool1 == True:
                                ErrorFound1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                print(ErrorFound1)
                                driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                TestResult.append(FundNameInvList[ii4] + " not able to open\n" + ErrorFound1)
                                TestResultStatus.append("Fail")
                                driver.close()
                        except Exception:
                            try:
                                time.sleep(2)
                                bool2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                if bool2 == True:
                                    ErrorFound2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                    print(ErrorFound2)
                                    TestResult.append(FundNameInvList[ii4] + " not able to open\n" + ErrorFound2)
                                    TestResultStatus.append("Fail")
                                    driver.close()
                            except Exception:
                                pass
                            pass
                        #---Cont & Dist button-------------
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div[2]/button").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        try:
                            time.sleep(2)
                            bool1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                            if bool1 == True:
                                ErrorFound1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                print(ErrorFound1)
                                driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                TestResult.append("Cont & Dist button not able to open\n" + ErrorFound1)
                                TestResultStatus.append("Fail")
                                driver.close()
                        except Exception:
                            try:
                                time.sleep(2)
                                bool2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                if bool2 == True:
                                    ErrorFound2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                    print(ErrorFound2)
                                    TestResult.append("Cont & Dist button not able to open\n" + ErrorFound2)
                                    TestResultStatus.append("Fail")
                                    driver.close()
                            except Exception:
                                pass
                            pass
                        #--------Drop icon clikcked for Period----------
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div/div/table/tbody/tr/td/div/p/span/a/span").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        try:
                            time.sleep(2)
                            bool1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                            if bool1 == True:
                                ErrorFound1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                print(ErrorFound1)
                                driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                TestResult.append("Drop icon not able to open\n" + ErrorFound1)
                                TestResultStatus.append("Fail")
                                driver.close()
                        except Exception:
                            try:
                                time.sleep(2)
                                bool2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                if bool2 == True:
                                    ErrorFound2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                    print(ErrorFound2)
                                    TestResult.append("Drop icon not able to open\n" + ErrorFound2)
                                    TestResultStatus.append("Fail")
                                    driver.close()
                            except Exception:
                                pass
                            pass
                        #-------------Edit icon clicked----------------
                        driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div/div/table/tbody/tr[2]/td[9]/div/p/a").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        try:
                            time.sleep(2)
                            bool1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                            if bool1 == True:
                                ErrorFound1 = driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                print(ErrorFound1)
                                driver.find_element_by_xpath(
                                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                TestResult.append("Edit icon not able to open on click\n" + ErrorFound1)
                                TestResultStatus.append("Fail")
                                driver.close()
                        except Exception:
                            try:
                                time.sleep(2)
                                bool2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                if bool2 == True:
                                    ErrorFound2 = driver.find_element_by_xpath(
                                        "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                    print(ErrorFound2)
                                    TestResult.append("Edit icon not able to open on click\n" + ErrorFound2)
                                    TestResultStatus.append("Fail")
                                    driver.close()
                            except Exception:
                                pass
                            pass

                        for periodloop in range(1,50):
                            print()
                            if periodloop==1:
                                P = driver.find_element_by_xpath(
                                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div[5]/div[1]/div/div[2]/div/div/span").text
                                if P in "--- Select a Value ---":
                                    print("Select found")
                                    try:
                                        time.sleep(2)
                                        bool1 = driver.find_element_by_xpath(
                                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                                        if bool1 == True:
                                            ErrorFound1 = driver.find_element_by_xpath(
                                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                            print(ErrorFound1)
                                            driver.find_element_by_xpath(
                                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                            TestResult.append(P + " not able to open\n" + ErrorFound1)
                                            TestResultStatus.append("Fail")
                                            break
                                            driver.close()
                                    except Exception:
                                        try:
                                            time.sleep(2)
                                            bool2 = driver.find_element_by_xpath(
                                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                            if bool2 == True:
                                                ErrorFound2 = driver.find_element_by_xpath(
                                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                                print(ErrorFound2)
                                                TestResult.append(P + " not able to open\n" + ErrorFound2)
                                                TestResultStatus.append("Fail")
                                                break
                                                driver.close()
                                        except Exception:
                                            pass
                                        pass

                                    break
                            else:
                                driver.find_element_by_xpath(
                                             "//div[@class='ContentLayout---content_layout']/div[3]/div/div[5]/div[1]/div/div[2]/div/div").click()
                                time.sleep(1)
                                ActionChains(driver).key_down(Keys.DOWN).perform()
                                time.sleep(1)
                                P = driver.find_element_by_xpath(
                                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div[5]/div[1]/div/div[2]/div/div/span").text
                                time.sleep(1)
                                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                                try:
                                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                    WebDriverWait(driver, LONG_TIMEOUT
                                                  ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                except TimeoutException:
                                    pass
                                try:
                                    time.sleep(2)
                                    bool1 = driver.find_element_by_xpath(
                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                                    if bool1 == True:
                                        ErrorFound1 = driver.find_element_by_xpath(
                                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                        print(ErrorFound1)
                                        driver.find_element_by_xpath(
                                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                        TestResult.append(P + " not able to open\n" + ErrorFound1)
                                        TestResultStatus.append("Fail")
                                        break
                                        driver.close()
                                except Exception:
                                    try:
                                        time.sleep(2)
                                        bool2 = driver.find_element_by_xpath(
                                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                        if bool2 == True:
                                            ErrorFound2 = driver.find_element_by_xpath(
                                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                            print(ErrorFound2)
                                            TestResult.append(P + " not able to open\n" + ErrorFound2)
                                            TestResultStatus.append("Fail")
                                            break
                                            driver.close()
                                    except Exception:
                                        pass
                                    pass
                                P = driver.find_element_by_xpath(
                                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div[5]/div[1]/div/div[2]/div/div/span").text

                            print("P is "+P)
                            if P in "--- Select a Value ---":
                                print("Select found")
                                try:
                                    time.sleep(2)
                                    bool1 = driver.find_element_by_xpath(
                                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                                    if bool1 == True:
                                        ErrorFound1 = driver.find_element_by_xpath(
                                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                                        print(ErrorFound1)
                                        driver.find_element_by_xpath(
                                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                                        TestResult.append(P + " not able to open\n" + ErrorFound1)
                                        TestResultStatus.append("Fail")
                                        break
                                        driver.close()
                                except Exception:
                                    try:
                                        time.sleep(2)
                                        bool2 = driver.find_element_by_xpath(
                                            "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                                        if bool2 == True:
                                            ErrorFound2 = driver.find_element_by_xpath(
                                                "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                            print(ErrorFound2)
                                            TestResult.append(P + " not able to open\n" + ErrorFound2)
                                            TestResultStatus.append("Fail")
                                            break
                                            driver.close()
                                    except Exception:
                                        pass
                                    pass

                                break
                            else:
                                if Dict.get(FundNameInvList[ii4] + P) != None:
                                    print("Dict has " + Dict.get(FundNameInvList[ii4] + P))
                                    ValueFound = driver.find_element_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div[6]/div/div/div[2]/div/div/table/tbody/tr[last()]/td[last()]/div/p/strong").text
                                    if ValueFound in "_":
                                        ValueFound="0"
                                    print("We Found " + ValueFound)
                                    if Dict.get(FundNameInvList[ii4] + P) in ValueFound:
                                        TestResult.append(
                                            "Data matched for Fund [ " + FundNameInvList[ii4] + " ] Liquid Trust is "+P)
                                        TestResultStatus.append("Pass")
                                        pass
                                    else:
                                        TestResult.append("Data not matching for Fund [ "+FundNameInvList[ii4]+" ] \nValue at Fund listing (Investement view) is: " + Dict.get(FundNameInvList[ii4] + P)+" , value found in ContDist section is: "+ValueFound)
                                        TestResultStatus.append("Fail")

                                else:
                                    print("Dict has None")
                                    ValueFound = driver.find_element_by_xpath(
                                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div[6]/div/div/div[2]/div/div/table/tbody/tr[1]/td[15]/div/input").get_attribute(
                                        "value")
                                    if ValueFound in "_":
                                        ValueFound="0"
                                    print("We Found " + ValueFound)
                        #-------------------------------------------------------------***************************************

                        # indexLT=LiquidTrustInvList.index(P)
                        # print("indexLT "+str(indexLT))
                        # print(FundNameInvList[indexLT])
                        # print(DistributionInvList[indexLT])

                        driver.close()
                        driver.switch_to.window(main_window)
                        #driver.refresh()
                        PageName = "Liquid Trusts"
                        driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
                        try:
                            driver.switch_to_alert().accept()
                        except Exception:
                            pass
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass

                        PageName = "Capital Activity Summary By Period"
                        driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                    FundIterate=FundIterate+1
            except Exception as e1:
                print(e1)
                TestResult.append("View by Investments button at Capital Call & Distribution Activity is not able to click")
                TestResultStatus.append("Fail")
        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start),2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------
