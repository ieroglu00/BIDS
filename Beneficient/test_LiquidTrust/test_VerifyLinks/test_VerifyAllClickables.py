import datetime
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_VerifyAllClickables_LiquidTrust"
  description = "This test scenario is to verify all the clickable sections at Liquid Trust Page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_LiquidTrust"
  global Exe
  Exe="Yes"
  Directory = 'test_LiquidTrust/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        print()
        #---------------------------Verify Liquid Trusts page-----------------------------
        PageName="Liquid Trusts"
        Ptitle1="Liquid Trusts - BIDS"
        driver.find_element_by_xpath("//*[@title='"+PageName+"']").click()
        for iat1 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        try:
            try:
                PageTitle1 = driver.title
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
            except Exception:
                Ptitle1="Funds - BIDS"
                PageTitle1 = driver.title
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName +     " page not able to open")
            TestResultStatus.append("Fail")
        print()
        #---------------------------------------------------------------------------------

        #--------------------Clicking on Capital Activity Summary By Period section--------------
        PageName = "Capital Activity Summary By Period"
        Ptitle1 = "COR_CapitalActivityPeriodTemplate - BIDS"
        driver.find_element_by_xpath("//strong[contains(text(),'"+PageName+"')]").click()
        for iat2 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        try:
            #PageTitle1 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div[2]/button").text
            PageTitle1= driver.title
            print(PageTitle1)
            assert PageTitle1 in Ptitle1, PageName + " is not able to open successfully"
            TestResult.append(PageName + " opened successfully")
            TestResultStatus.append("Pass")
        except Exception as e1:
            print(e1)
            TestResult.append(PageName + " is not able to open successfully")
            TestResultStatus.append("Fail")
        driver.find_element_by_xpath("//*[@title='Liquid Trusts']").click()
        for iat3 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        print()
        #---------------------------------------------------------------------------------

        # --------------------Clicking on Liquid Trusts - Trust Accounting section--------------
        PageName = "Liquid Trusts - Trust Accounting"
        Ptitle1 = "Trust Accounting Report - BIDS"
        driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
        for iat4 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        wait = WebDriverWait(driver, 60)
        wait.until(EC.presence_of_element_located((By.XPATH,
                                                   "//div[@class='ContentLayout---content_layout']/div/div[2]/div/div[1]/div[1]/div/div[2]/div/div/table/tbody/tr[1]/td[1]/div/p/strong")))
        try:
            # PageTitle1 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div[2]/button").text
            PageTitle1 = driver.title
            print(PageTitle1)
            assert PageTitle1 in Ptitle1, PageName + " is not able to open successfully"
            TestResult.append(PageName + " opened successfully")
            TestResultStatus.append("Pass")
        except Exception as e1:
            print(e1)
            TestResult.append(PageName + " is not able to open successfully")
            TestResultStatus.append("Fail")
        driver.find_element_by_xpath("//*[@title='Liquid Trusts']").click()
        for iat5 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        print()
        # ---------------------------------------------------------------------------------
        # --------------------Clicking on Unfunded Commitments section--------------
        PageName = "Unfunded Commitments"
        Ptitle1 = "Unfunded Rollforward/Commitment Page - BIDS"
        driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
        for iat6 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        wait = WebDriverWait(driver, 60)
        wait.until(EC.presence_of_element_located((By.XPATH,
                                                   "//div[@class='ContentLayout---content_layout']/div[2]/div[1]/div/div[1]/span")))
        try:
            # PageTitle1 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div[2]/button").text
            PageTitle1 = driver.title
            print(PageTitle1)
            assert PageTitle1 in Ptitle1, PageName + " is not able to open successfully"
            TestResult.append(PageName + " opened successfully")
            TestResultStatus.append("Pass")
        except Exception as e1:
            print(e1)
            TestResult.append(PageName + " is not able to open successfully")
            TestResultStatus.append("Fail")
        driver.find_element_by_xpath("//*[@title='Liquid Trusts']").click()
        for iat7 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        print()
        # ---------------------------------------------------------------------------------
        # --------------------Clicking on Liquid Trusts - Underwriting section--------------
        PageName = "Liquid Trusts - Underwriting"
        Ptitle1 = "COR_ListOfLiquidTrusts - BIDS"
        driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
        for iat8 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        try:
            # PageTitle1 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div[2]/button").text
            PageTitle1 = driver.title
            print(PageTitle1)
            assert PageTitle1 in Ptitle1, PageName + " is not able to open successfully"
            TestResult.append(PageName + " opened successfully")
            TestResultStatus.append("Pass")
        except Exception as e1:
            print(e1)
            TestResult.append(PageName + " is not able to open successfully")
            TestResultStatus.append("Fail")
        driver.find_element_by_xpath("//*[@title='Liquid Trusts']").click()
        for iat9 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        print()
        # ---------------------------------------------------------------------------------
        # --------------------Clicking on K-1 Report section--------------
        PageName = "K-1 Report"
        Ptitle1 = "K1 Report - BIDS"
        driver.find_element_by_xpath("//strong[contains(text(),'" + PageName + "')]").click()
        for iat10 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        try:
            # PageTitle1 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div[2]/button").text
            PageTitle1 = driver.title
            print(PageTitle1)
            assert PageTitle1 in Ptitle1, PageName + " is not able to open successfully"
            TestResult.append(PageName + " opened successfully")
            TestResultStatus.append("Pass")
        except Exception as e1:
            print(e1)
            TestResult.append(PageName + " is not able to open successfully")
            TestResultStatus.append("Fail")
        driver.find_element_by_xpath("//*[@title='Liquid Trusts']").click()
        for iat11 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        # ---------------------------------------------------------------------------------

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------

