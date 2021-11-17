import datetime
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  TestName = "test_ComparisonSOSwithMissionControl"
  description = "This is smoke test case to verify compared values between SOS Funds / Investments AND Mission Control Funds / Investments values"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_MissionControl"
  global Exe
  Exe="Yes"

  ExcelFileName = "Execution"
  locx = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()
      for iat1 in range(1000):
          try:
              bool = driver.find_element_by_xpath(
                  "//div[@id='appian-working-indicator-hidden']").is_enabled()
          except Exception:
              time.sleep(1)
              break
      time.sleep(1)

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_ComparisonSOSwithMissionControl(test_setup):
    YearCounterNumber=8
    FirstQuarter="12/31/2021"
    YearList = []
    # Lists for Mission Control Funds/Investments
    AmountListFundLevel = []
    AmountListInvestmentLevel = []

    # Lists for SOS Funds/Investments
    AmountListSOSFunds = []
    AmountListSOSSPV = []
    AmountListSOSInvestmenst = []

    if Exe == "Yes":
        try:
            print()
            PageName = "Quarterly NAV Close"
            Ptitle1="Quarterly NAV Close - BIDS"
            driver.find_element_by_xpath("//*[@title='Quarterly NAV Close']").click()
            start = time.time()
            for iat2 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(1)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)

            ###############################################
            PageName = "Mission Control"
            Ptitle1 = "COR_ReportMissionControl - BIDS"
            driver.find_element_by_xpath("//*[text() = '"+PageName+"']").click()
            start = time.time()
            for iat3 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(5)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")

            wait = WebDriverWait(driver, 300)
            wait.until(EC.presence_of_element_located((By.XPATH,
                                                       "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div/span")))
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)

            # Setting first quarter in the quarter dropdown list
            for iat4 in range(10):
                P = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div/span").text
                #print("P found is "+P)
                if P in FirstQuarter:
                    #print()
                    #print()
                    #print("Found the first quarter for "+PageName)
                    break
                else:
                    print("Trying again as P found is "+P)
                    driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div").click()
                    ActionChains(driver).key_down(Keys.UP).perform()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    for iat16 in range(1000):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        except Exception:
                            time.sleep(1)
                            break
                    time.sleep(5)

            wait = WebDriverWait(driver, 300)
            wait.until(EC.presence_of_element_located((By.XPATH,
                 "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div/span")))
            for ia in range(YearCounterNumber):
                # print("Count is " + str(ia))
                if ia == 0:
                    # print("First Data")
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div/span").text
                elif ia > 0:
                    time.sleep(5)
                    # print("Other Data")
                    driver.find_element_by_xpath(
                        "//input[@class='PickerWidget---picker_input PickerWidget---placeholder']").send_keys()
                    elements = driver.find_elements_by_xpath(
                        "//div[@class='DropdownWidget---dropdown_value DropdownWidget---inSideBySideItem']")
                    for elem in elements:
                        elem.click()
                        break
                    time.sleep(5)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    for iat5 in range(1000):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                            # print("Loader present")
                        except Exception:
                            # print("Loader finished")
                            time.sleep(5)
                            break
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div/span").text
                for iat6 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath("//div[@id='appian-working-indicator-hidden']").is_enabled()
                    except Exception:
                        # print("Loader finished")
                        time.sleep(3)
                        break
                time.sleep(5)
                try:
                    AmtFundLevel = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[5]/div[2]/div/div/div[2]/div[10]/div[2]/div/p/span/strong").text
                    AmtInvestmentLevel = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[5]/div[2]/div/div/div[3]/div[10]/div[2]/div/p/span/strong").text
                    YearList.append(P)
                    AmountListFundLevel.append(AmtFundLevel)
                    AmountListInvestmentLevel.append(AmtInvestmentLevel)
                except Exception:
                    YearList.append(P)
                    AmtFundLevel="0"
                    AmtInvestmentLevel="0"
                    AmountListFundLevel.append(AmtFundLevel)
                    AmountListInvestmentLevel.append(AmtInvestmentLevel)

            # Navigating back to Quarterly NAV Close Page
            driver.find_element_by_xpath("//*[@title='Quarterly NAV Close']").click()
            for iat7 in range(1000):
                try:
                    bool = driver.find_element_by_xpath("//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    # print("Loader finished")
                    time.sleep(3)
                    break


            ####################################################
            PageName = "Sign-Off Summary: SPVs"
            Ptitle1 = "User Input Task - BIDS"
            driver.find_element_by_xpath("//*[text() = '" + PageName + "']").click()
            for iat13 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(5)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")

            # Setting first quarter in the quarter dropdown list
            wait = WebDriverWait(driver, 300)
            wait.until(EC.presence_of_element_located((By.XPATH,
                                                       "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div/span")))
            for iat15 in range(10):
                P = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div/span").text
                #print("P found is " + P)
                if P in FirstQuarter:
                    # print()
                    # print()
                    # print("Found the first quarter for "+PageName)
                    break
                else:
                    print("Trying again as P found is "+P)
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div").click()
                    ActionChains(driver).key_down(Keys.UP).perform()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    for iat18 in range(1000):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        except Exception:
                            time.sleep(1)
                            break
                    time.sleep(5)

            for iaa in range(YearCounterNumber):
                # print("Count is " + str(ia))
                if iaa == 0:
                    # print("First Data")
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div").text
                elif iaa > 0:
                    time.sleep(5)
                    # print("Other Data")
                    driver.find_element_by_xpath(
                        "//input[@class='PickerWidget---picker_input PickerWidget---placeholder']").send_keys()
                    elements = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div")
                    for elem in elements:
                        elem.click()
                        break
                    time.sleep(5)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    for iat8 in range(1000):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                            # print("Loader present")
                        except Exception:
                            # print("Loader finished")
                            time.sleep(5)
                            break
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div").text
                for iat9 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath("//div[@id='appian-working-indicator-hidden']").is_enabled()
                    except Exception:
                        # print("Loader finished")
                        time.sleep(3)
                        break
                time.sleep(5)
                try:
                    element = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[3]/div/div[2]/div/input")
                    SPVSOSTotalPartnerNAVEndingUSD = element.get_attribute("value")
                    AmountListSOSSPV.append(SPVSOSTotalPartnerNAVEndingUSD)
                except Exception:
                    SPVSOSTotalPartnerNAVEndingUSD="0"
                    AmountListSOSSPV.append(SPVSOSTotalPartnerNAVEndingUSD)

            # Navigating back to Quarterly NAV Close Page
            driver.find_element_by_xpath("//*[@title='Quarterly NAV Close']").click()
            for iat10 in range(1000):
                try:
                    bool = driver.find_element_by_xpath("//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    # print("Loader finished")
                    time.sleep(3)
                    break

            ####################################################
            PageName = "Sign-Off Summary: Funds"
            Ptitle1 = "User Input Task - BIDS"
            driver.find_element_by_xpath("//*[text() = '" + PageName + "']").click()
            for iat13 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(5)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")

            # Setting first quarter in the quarter dropdown list
            wait = WebDriverWait(driver, 300)
            wait.until(EC.presence_of_element_located((By.XPATH,
                                                       "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div[1]/div[1]/div/div[2]/div/div/span")))
            for iat15 in range(10):
                P = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div[1]/div[1]/div/div[2]/div/div/span").text
                #print("P found is " + P)
                if P in FirstQuarter:
                    # print()
                    # print()
                    # print("Found the first quarter for " + PageName)
                    break
                else:
                    print("Trying again as P found is " + P)
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div[1]/div[1]/div/div[2]/div/div").click()
                    ActionChains(driver).key_down(Keys.UP).perform()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    for iat18 in range(1000):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        except Exception:
                            time.sleep(1)
                            break
                    time.sleep(5)

            for iaa in range(YearCounterNumber):
                # print("Count is " + str(ia))
                if iaa == 0:
                    # print("First Data")
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div/span").text
                elif iaa > 0:
                    time.sleep(5)
                    # print("Other Data")
                    driver.find_element_by_xpath(
                        "//input[@class='PickerWidget---picker_input PickerWidget---placeholder']").send_keys()
                    elements = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div")
                    for elem in elements:
                        elem.click()
                        break
                    time.sleep(5)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    for iat8 in range(1000):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                            time.sleep(1)
                        except Exception:
                            # print("Loader finished")
                            time.sleep(1)
                            break
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[1]/div/div[3]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div/span").text
                for iat9 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath(
                            "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        time.sleep(1)
                    except Exception:
                        # print("Loader finished")
                        time.sleep(1)
                        break
                time.sleep(10)
                try:
                    element = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[8]/div/div[2]/div/input")
                    FundsSOSTotalPartnerNAVEndingUSD = element.get_attribute("value")
                    AmountListSOSFunds.append(FundsSOSTotalPartnerNAVEndingUSD)
                except Exception:
                    FundsSOSTotalPartnerNAVEndingUSD="0"
                    AmountListSOSFunds.append(FundsSOSTotalPartnerNAVEndingUSD)

            # Navigating back to Quarterly NAV Close Page
            driver.find_element_by_xpath("//*[@title='Quarterly NAV Close']").click()
            for iat10 in range(1000):
                try:
                    bool = driver.find_element_by_xpath("//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    # print("Loader finished")
                    time.sleep(3)
                    break



            ####################################################
            PageName = "Sign-Off Summary: Investments"
            Ptitle1 = "User Input Task - BIDS"
            driver.find_element_by_xpath("//*[text() = '" + PageName + "']").click()
            for iat14 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(5)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")

            # Setting first quarter in the quarter dropdown list
            wait = WebDriverWait(driver, 300)
            wait.until(EC.presence_of_element_located((By.XPATH,
                                                       "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div/span")))
            for iat17 in range(10):
                P = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div/span").text
                #print("P found is " + P)
                if P in FirstQuarter:
                    # print()
                    # print()
                    # print("Found the first quarter for "+PageName)
                    break
                else:
                    print("Trying again as P found is "+P)
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div").click()
                    ActionChains(driver).key_down(Keys.UP).perform()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    for iat19 in range(1000):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        except Exception:
                            time.sleep(1)
                            break
                    time.sleep(5)

            for iaaa in range(YearCounterNumber):
                # print("Count is " + str(ia))
                if iaaa == 0:
                    # print("First Data")
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div/span").text
                elif iaaa > 0:
                    time.sleep(5)
                    # print("Other Data")
                    driver.find_element_by_xpath(
                        "//input[@class='PickerWidget---picker_input PickerWidget---placeholder']").send_keys()
                    elements = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div")
                    for elem in elements:
                        elem.click()
                        break
                    time.sleep(5)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    for iat11 in range(1000):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                            time.sleep(1)
                        except Exception:
                            time.sleep(1)
                            break
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/div/span").text
                for iat12 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath("//div[@id='appian-working-indicator-hidden']").is_enabled()
                        time.sleep(1)
                    except Exception:
                        # print("Loader finished")
                        time.sleep(1)
                        break
                time.sleep(10)
                try:
                    element = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[5]/div/div/div[2]/div/div[8]/div/div[2]/div/input")
                    InvestmentSOSBenLevelStruckNAVatEndofPeriod = element.get_attribute("value")
                    AmountListSOSInvestmenst.append(InvestmentSOSBenLevelStruckNAVatEndofPeriod)
                except Exception:
                    InvestmentSOSBenLevelStruckNAVatEndofPeriod="0"
                    AmountListSOSInvestmenst.append(InvestmentSOSBenLevelStruckNAVatEndofPeriod)

            # print("Now Comparing the numbers")
            for x1 in range(len(AmountListFundLevel)):
                print()
                print("Fund Level For year: "+YearList[x1])
                print("AmountListSOSFunds is"+AmountListSOSFunds[x1])
                print("AmountListSOSSPV is" + AmountListSOSSPV[x1])
                print("AmountListFundLevel is" + AmountListFundLevel[x1])

                Value1=AmountListSOSFunds[x1]
                if "_" in AmountListSOSFunds[x1]:
                    Value1 ="0"
                Value1 = Value1.replace(" ", "")
                Value1 = re.sub(r'[?|$|€|£|!|,]', r'', Value1)
                Value1SOSFundsFloat=float(Value1)

                Value2 = AmountListSOSSPV[x1]
                if "_" in AmountListSOSSPV[x1]:
                    Value2 ="0"
                Value2 = Value2.replace(" ", "")
                Value2 = re.sub(r'[?|$|€|£|!|,]', r'', Value2)
                Value2SOSSPVFloat = float(Value2)

                TotalSOSSPV_SOSFunds=float(Value1SOSFundsFloat+Value2SOSSPVFloat)
                print("TotalSOSSPV_SOSFunds is: "+str(TotalSOSSPV_SOSFunds))

                Value3 = AmountListFundLevel[x1]
                if "_" in AmountListFundLevel[x1]:
                    Value3 ="0"
                Value3 = Value3.replace(" ", "")
                Value3 = re.sub(r'[?|$|€|£|!|,]', r'', Value3)
                Value3AmountListFundLevelFloat = float(Value3)

                if Value3AmountListFundLevelFloat != TotalSOSSPV_SOSFunds:
                    print(str(Value3AmountListFundLevelFloat - TotalSOSSPV_SOSFunds))
                    print("AmountListFundLevel ( " + AmountListFundLevel[x1] + " ) and AmountListSOSFunds ( " +
                          AmountListSOSFunds[x1] + " ) with AmountListSOSSPV ( " +
                          AmountListSOSSPV[x1] + " ) **NOT** matching for year " + YearList[x1])

                    TestResult.append("Mission Control Fund ( " + AmountListFundLevel[x1] + " ) and SOS Funds ( " +
                          AmountListSOSFunds[x1] + " ) with SOS SPV ( " +
                          AmountListSOSSPV[x1] + " ) **NOT** matching for year " + YearList[x1])
                    TestResultStatus.append("Fail")

            print("-----------------------------------------------------------------------------------")
            for x11 in range(len(AmountListInvestmentLevel)):
                print("Investment level For year: " + YearList[x11])
                print("AmountListSOSInvestmenst is" + AmountListSOSInvestmenst[x11])
                print("AmountListInvestmentLevel is" + AmountListInvestmentLevel[x11])
                print()

                Value4 = AmountListSOSInvestmenst[x11]
                if "_" in AmountListSOSInvestmenst[x11]:
                    Value4 ="0"
                Value4 = Value4.replace(" ", "")
                Value4 = re.sub(r'[?|$|€|£|!|,]', r'', Value4)
                Value4SOSInvestmenstFloat = float(Value4)

                Value5 = AmountListInvestmentLevel[x11]
                if "_" in AmountListInvestmentLevel[x11]:
                    Value5 ="0"
                Value5 = Value5.replace(" ", "")
                Value5 = re.sub(r'[?|$|€|£|!|,]', r'', Value5)
                Value5InvestmentLevelFloat = float(Value5)

                if Value5InvestmentLevelFloat != Value4SOSInvestmenstFloat:
                    print(str(float(Value5InvestmentLevelFloat-Value4SOSInvestmenstFloat)))
                    print("AmountListInvestmentLevel ( " + AmountListInvestmentLevel[
                        x11] + " ) and AmountListSOSInvestments ( " + AmountListSOSInvestmenst[
                              x11] + " ) **NOT** matching for year " + YearList[x11])

                    TestResult.append("Mission Control Investment ( " + AmountListInvestmentLevel[
                        x11] + " ) and SOS Investments ( " + AmountListSOSInvestmenst[
                              x11] + " ) **NOT** matching for year " + YearList[x11])
                    TestResultStatus.append("Fail")
        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start),2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            TestResult.append(stringMainerror)
            TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


