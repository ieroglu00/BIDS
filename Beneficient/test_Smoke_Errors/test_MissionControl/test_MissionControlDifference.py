import datetime
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  TestName = "test_MissionControlDifference"
  description = "This is smoke test case to find difference values for NAV Rollover from Prior Period and NAV Roll for Period"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_MissionControl"
  global Exe
  Exe="Yes"

  ExcelFileName = "Execution"
  locx = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()
      for iat1 in range(1000):
          try:
              bool = driver.find_element_by_xpath(
                  "//div[@id='appian-working-indicator-hidden']").is_enabled()
          except Exception:
              time.sleep(1)
              break
      time.sleep(1)

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_MissionControlDifference(test_setup):
    YearCounterNumber=7
    FirstQuarter="12/31/2021"
    SHORT_TIMEOUT = 5
    LONG_TIMEOUT = 400
    LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"

    if Exe == "Yes":
        try:
            print()
            PageName = "Quarterly NAV Close"
            Ptitle1="Quarterly NAV Close - BIDS"
            driver.find_element_by_xpath("//*[@title='Quarterly NAV Close']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)

            ###############################################
            PageName = "Mission Control"
            Ptitle1 = "COR_ReportMissionControl - BIDS"
            driver.find_element_by_xpath("//*[text() = '"+PageName+"']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")

            wait = WebDriverWait(driver, LONG_TIMEOUT)
            wait.until(EC.presence_of_element_located((By.XPATH,
                "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div/span")))
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)

            #---------Setting first quarter in the quarter dropdown list----------------------
            for iat4 in range(10):
                P = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div/span").text
                if P in FirstQuarter:
                    break
                else:
                    print("Trying again as P found is "+P)
                    driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div").click()
                    ActionChains(driver).key_down(Keys.UP).perform()
                    time.sleep(2)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(2)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                        if bool1 == True:
                            ErrorFound1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                            print(ErrorFound1)
                            driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                            TestResult.append(P + " not able to open\n" + ErrorFound1)
                            TestResultStatus.append("Fail")
                            bool1 = False
                            driver.close()
                    except Exception:
                        try:
                            time.sleep(2)
                            bool2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                            if bool2 == True:
                                ErrorFound2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                print(ErrorFound2)
                                TestResult.append(P + " not able to open\n" + ErrorFound2)
                                TestResultStatus.append("Fail")
                                bool2 = False
                                driver.close()
                        except Exception:
                            pass
                        pass
                    time.sleep(1)
            #---------------------------------------------------------------------------------

            for ia in range(YearCounterNumber):
                if ia == 0:
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div/span").text
                elif ia > 0:
                    time.sleep(5)
                    driver.find_element_by_xpath(
                        "//input[@class='PickerWidget---picker_input PickerWidget---placeholder']").send_keys()
                    elements = driver.find_elements_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div")
                    for elem in elements:
                        elem.click()
                        break
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).perform()
                    time.sleep(3)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(2)
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        time.sleep(2)
                        bool1 = driver.find_element_by_xpath(
                            "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                        if bool1 == True:
                            ErrorFound1 = driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                            print(ErrorFound1)
                            driver.find_element_by_xpath(
                                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                            TestResult.append(P + " not able to open\n" + ErrorFound1)
                            TestResultStatus.append("Fail")
                            bool1 = False
                            driver.close()
                    except Exception:
                        try:
                            time.sleep(2)
                            bool2 = driver.find_element_by_xpath(
                                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                            if bool2 == True:
                                ErrorFound2 = driver.find_element_by_xpath(
                                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                                print(ErrorFound2)
                                TestResult.append(P + " not able to open\n" + ErrorFound2)
                                TestResultStatus.append("Fail")
                                bool2 = False
                                driver.close()
                        except Exception:
                            pass
                        pass
                    time.sleep(1)
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/div[2]/div/div[2]/div/div/span").text

                print("Year is "+P)
                rows = driver.find_elements_by_xpath(
                    "//div[@class='BoxLayout---box BoxLayout---margin_below_standard BoxLayout---margin_above_none'][1]/div[2]/div/div/div[4]/div")
                print("Lenght of rows " + str(len(rows)))
                if len(rows) == 0:
                    print("No Data available for quarter [ " + P + " ]")
                    TestResult.append("No Data available for quarter [ " + P + " ]")
                    TestResultStatus.append("Fail")

                for i in range(len(rows) + 1):
                    if i > 1:
                        # -----------------For NAV Rollover from Prior Period-------------------------
                        DifferenceValueString = driver.find_element_by_xpath(
                            "//div[@class='BoxLayout---box BoxLayout---margin_below_standard BoxLayout---margin_above_none'][1]/div[2]/div/div/div[4]/div[" + str(
                                i) + "]/div[2]/div/p").text
                        DifferenceValueString1=DifferenceValueString
                        print("DifferenceValueString is " + DifferenceValueString1)
                        DifferenceValueString = DifferenceValueString.replace(" ", "")
                        DifferenceValueString = re.sub(r'[?|$|.|!|,|-]', r'', DifferenceValueString)
                        DifferenceValueString = ''.join(char for char in DifferenceValueString if char.isalnum())
                        try:
                            if (int(DifferenceValueString) > 0 or int(DifferenceValueString) < 0):
                                print(
                                    "Difference found in NAV Rollover from Prior Period for quarter [ " + P + " ] and the difference amount is [ " + DifferenceValueString1+" ]")
                                TestResult.append("Difference found in NAV Rollover from Prior Period for quarter [ " + P + " ] and the difference amount is [ " + DifferenceValueString1+" ]")
                                TestResultStatus.append("Fail")
                        except Exception:
                            print("- or blank value")

                        # ---------------------------For NAV Roll for Period------------------------------
                        DifferenceValueStringOther = driver.find_element_by_xpath(
                            "//div[@class='BoxLayout---box BoxLayout---margin_below_standard BoxLayout---margin_above_none'][2]/div[2]/div/div/div[4]/div[" + str(
                                              i) + "]/div[2]/div/p").text
                        DifferenceValueStringOther1 = DifferenceValueStringOther
                        print("DifferenceValueStringOther is " + DifferenceValueStringOther1)
                        DifferenceValueStringOther = DifferenceValueStringOther.replace(" ", "")
                        DifferenceValueStringOther = re.sub(r'[?|$|.|!|,|-]', r'', DifferenceValueStringOther)
                        DifferenceValueStringOther = ''.join(char for char in DifferenceValueStringOther if char.isalnum())
                        try:
                            if (int(DifferenceValueStringOther) > 0 or int(DifferenceValueStringOther) < 0):
                                print(
                                    "Difference found in NAV Rollover for Period for quarter [ " + P + " ] and the difference amount is [ " + DifferenceValueStringOther1+" ]")
                                TestResult.append(
                                    "Difference found in NAV Rollover for Period for quarter [ " + P + " ] and the difference amount is [ " + DifferenceValueStringOther1+" ]")
                                TestResultStatus.append("Fail")
                        except Exception:
                            print("- or blank value")

                print()
        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start), 2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror = repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


