import os
import smtplib
import ssl
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import pandas as pd
import pytest

@pytest.mark.smoke
def test_ReportSendSmokeAll():
    print()
    DirectoryName = []
    PDFName1=[]
    TestName=[]
    TestDescription = []
    TestStatus = []
    SendStatus = []
    AttachmntAdded= []
    PDFpath = 'C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/'

    #-------------------To read content to send in e-Mail--------------------
    ExcelFileName = "FileName"
    loc = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/PDFFileNameData/' + ExcelFileName + '.xlsx')
    wb=openpyxl.load_workbook(loc)
    sheet = wb.active
    for i in range(1, 100):
        if sheet.cell(i, 1).value == None:
            break
        else:
            PDFName1.append(sheet.cell(i, 2).value)
            DirectoryName.append(sheet.cell(i, 3).value)
            TestName.append(sheet.cell(i, 1).value)
            TestDescription.append(sheet.cell(i, 4).value)
            TestStatus.append(sheet.cell(i, 5).value)
            SendStatus.append(sheet.cell(i, 6).value)

        B = ""
        for io in range(len(TestName)):
            try:
                B = B + "<br /><br />"+str(io+1)+") " + "".join(TestName[io])+" => "+"".join(TestDescription[io])+" => "+"".join(TestStatus[io])
            except Exception:
                print("No attachment details to add in email description")
    #print(B)

    ##############################################################
    html = '''
        <html>
            <body>
                <p>Hi Team <br />Here is the test summary report of Smoke Test 2 (To verify presence of Errors / Quick on-screen calculated numbers in all modules) <br />Below test scenarios are covered </p>
                <p></p>
                <p>'''+B+'''</p
                <p></p>
                <img src='cid:myimageid' width="500" align="center">
                <p>Please find attached PDFs of test scenarios results<br />Note: Attachments are only for FAILED test cases<br /></p>
                <p>Many Thanks <br/>Neeraj</p>
            </body>
        </html>
        '''
    ##############################################################

    def attach_file_to_email(msg, filename, extra_headers=None):
        with open(filename, "rb") as f:
            file_attachment = MIMEApplication(f.read())
        file_attachment.add_header(
            "Content-Disposition",
            f"attachment; filename= {filename}",
        )
        if extra_headers is not None:
            for name, value in extra_headers.items():
                file_attachment.add_header(name, value)
        msg.attach(file_attachment)

    email_from = 'Test Automation Team'
    #email_to =['erin.Twiss@beneficient.com','neeraj1wayitsol@gmail.com','srinath.jandhyala@beneficient.com','gregory.bybee@beneficient.com','olufunke.okeni@beneficient.com']
    email_to =['neeraj1wayitsol@gmail.com']

    SenderEmail="neeraj.kumar@bitsinglass.com"
    User = "AxBCsrijkltwxysHIJMhijodefhhij@bcd2yzA8qrs4zAB0EFG8fgh2zAB8"
    RandmStr = User[int(int(ord('b')) / 49)] + User[int(int(ord('c')) / 16.5)] + User[int(int(ord('c')) / 9.9)] + User[
        int(int(ord('b')) / 7)] + User[int((int(ord('Z')) / 5))] + \
               User[int(int(ord('X')) / 4)] + User[int(int(ord('N')) / 3)] + User[int(int(ord('Z')) / 3)] + User[
                   int(int(ord('f')) / 3)] + User[int(int(ord('r')) / 3)] + \
               User[int(int(ord('T')) / 2)] + User[int((int(ord('Z')) + 2) / 2)] + User[int((int(ord('b')) + 2) / 2)] + \
               User[int((int(ord('k')) + 2) / 2)] + User[int((int(ord('s')) + 2) / 2)]

    date_str = pd.Timestamp.today().strftime('%m-%d-%Y')
    msg = MIMEMultipart()
    msg['Subject']='[Test Suite 2 ( Smoke Errors )]-Test Automation Report-Env [Test] '+date_str
    msg['From'] = email_from
    msg['To'] = ','.join(email_to)
    msg.attach(MIMEText(html, "html"))

    #-----------------------------------------------------------------------
    try:
        attach_file_to_email(msg, 'C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/TestPieResult.png',
                         {'Content-ID': '<myimageid>'})
    except Exception:
        print("No Pie File to attach")
    # ------------------To add attachments in the report email--------------
    i=0
    for file in PDFName1:
        print()
        try:
            #print(file)
            if SendStatus[i]=="Send Only when Fail=Yes" and  TestStatus[i]=="Fail":
                    attach_file_to_email(msg,PDFpath+PDFName1[i])
                    AttachmntAdded.append("Yes")
            if SendStatus[i] == "Send Only when Fail=No":
                    print("Inside Send Only when Fail=No")
                    attach_file_to_email(msg, PDFpath + PDFName1[i])
                    AttachmntAdded.append("Yes")
        except Exception as e1:
            print("No Attachment found to Add")
            #print(e1)
        i = i + 1
    #-----------------------------------------------------------------------

    # ------------------------To attach all in e-Mail-----------------------
    email_string = msg.as_string()
    context = ssl.create_default_context()
    # -----------------------------------------------------------------------

    # ----------------------------SMTP setup--------------------------------
    server=smtplib.SMTP_SSL('smtp.gmail.com',465)
    server.login(SenderEmail,RandmStr)
    #-----------------------------------------------------------------------

    #---------------------------------Sending email-------------------------
    for io1 in range(len(AttachmntAdded)):
        if AttachmntAdded[io1] == "Yes":
            print("Inside AttachmntAdded=Yes ")
            server.sendmail(email_from, email_to, email_string)
            print("Test Report sent")
            break
    #-----------------------------------------------------------------------

    #-----------------To delete pdf report files----------------------------
    ii=0
    for ii in range(0,len(PDFName1)):
        print()
        try:
            os.remove('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/'+PDFName1[ii])
        except Exception:
            print("No Attachment found to delete")
    try:
        os.remove('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/TestPieResult.png')
    except Exception:
        print("No Attachment found to delete")
    #-----------------------------------------------------------------------
    server.quit()
