import datetime
import re
import time
from decimal import getcontext, Decimal

import openpyxl
import pyautogui
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestResultImage
  global TestDirectoryName
  global TotalBenNAV_USD_PerthisYearList
  global TotalNAVforallinvestments_PerthisYearList
  global myScreenshot
  global ct
  global ctReportHeader
  global YearCounter
  #myScreenshot = pyautogui.screenshot()

  TestName = "test_BeaconTotalBenCompStaticInvestment"
  description = "This is smoke test case to compare Total Ben NAV value for each quarter with Static Investment"
  TestResult = []
  TestResultStatus = []
  #TestResultImage = []
  TestFailStatus = []
  TotalBenNAV_USD_PerthisYearList = []
  TotalNAVforallinvestments_PerthisYearList = []
  YearCounter = []
  FailStatus="Pass"
  TestDirectoryName = "test_Beacon"
  global Exe
  Exe="Yes"

  ExcelFileName = "Execution"
  locx = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()
      for iat1 in range(1000):
          try:
              bool = driver.find_element_by_xpath(
                  "//div[@id='appian-working-indicator-hidden']").is_enabled()
          except Exception:
              time.sleep(1)
              break
      time.sleep(1)
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

  yield
  if Exe == "Yes":

      class PDF(FPDF):
          def header(self):
              self.image('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.set_text_color(0, 0, 0)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 20, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1=TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_BeaconTotalBenCompStaticInvestment(test_setup):
    YearCounterNumber = 4
    if Exe == "Yes":
        try:
            print()
            PageName = "Quarterly NAV Close"
            Ptitle1="Quarterly NAV Close - BIDS"
            driver.find_element_by_xpath("//*[@title='Quarterly NAV Close']").click()
            start = time.time()
            for iat2 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(1)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)

            PageName = "Beacon Template"
            Ptitle1 = "COR_BeaconDataTransferTemplate - BIDS"
            driver.find_element_by_xpath("//*[text() = '"+PageName+"']").click()
            start = time.time()
            for iat3 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(5)
            PageTitle1 = driver.title
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")

            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")

            wait = WebDriverWait(driver, 150)
            wait.until(EC.presence_of_element_located((By.XPATH,
                                                       "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div/span")))
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)
            for year in range(1,YearCounterNumber):
                print()
                try:
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div/span").text
                except Exception:
                    time.sleep(7)
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div/span").text
                print("Year is: " +P)
                YearCounter.append(P)
                for iat9 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath(
                            "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        time.sleep(1)
                    except Exception:
                        break
                time.sleep(5)
                TotalBenNAV_USD_PerthisYear=driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[5]/div/div/div/div[3]/div[2]/div/div/div[2]/div[1]/div[2]/div/p").text
                print("TotalBenNAV_USD_PerthisYear is "+TotalBenNAV_USD_PerthisYear)
                TotalBenNAV_USD_PerthisYearList.append(TotalBenNAV_USD_PerthisYear)
                print("Inside List is "+TotalBenNAV_USD_PerthisYearList[year-1])

                driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div").click()
                time.sleep(3)
                ActionChains(driver).key_down(Keys.DOWN).perform()
                time.sleep(3)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                for iat4 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath(
                            "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        # print("Loader is present")
                    except Exception:
                        time.sleep(1)
                        break
                time.sleep(10)



            PageName = "Static Investment Data"
            Ptitle1 = "Investment Information"
            time.sleep(3)
            driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div[3]/div/div[2]/div/div").click()
            time.sleep(3)
            ActionChains(driver).key_down(Keys.DOWN).perform()
            time.sleep(3)
            ActionChains(driver).key_down(Keys.DOWN).perform()
            time.sleep(3)
            ActionChains(driver).key_down(Keys.DOWN).perform()
            time.sleep(3)
            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
            for iat5 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                    # print("Loader is present")
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(10)
            for iat6 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/h3").is_enabled()
                    Check=driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/h3").text
                    if Check in Ptitle1:
                        break
                except Exception:
                    time.sleep(1)

            PageTitle1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/div/h3").text
            print("PageTitle1 found:"+PageTitle1)
            try:
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")

            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
                # driver.get_screenshot_as_file(
                #     r"C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/test_Beacon/" + PageName.replace(" ", "") +"_"+ ct + ".png")
                # TestResultImage.append(PageName.replace(" ", "") + "_" + ct + ".png")

            for year1 in range(1,YearCounterNumber):
                print()
                try:
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div/span").text
                except Exception:
                    time.sleep(7)
                    P = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div/span").text
                print("Year1 is: " +P)
                time.sleep(5)
                for iat8 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath(
                            "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        time.sleep(1)
                    except Exception:
                        #time.sleep(1)
                        break

                try:
                    TotalNAVforallinvestments_PerthisYear = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[5]/div/div/div/div[2]/div/div/div/div/div[2]/div[1]/div[2]/div/p").text
                except Exception:
                    time.sleep(7)
                    TotalNAVforallinvestments_PerthisYear = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[5]/div/div/div/div[2]/div/div/div/div/div[2]/div[1]/div[2]/div/p").text
                print("TotalNAVforallinvestments_PerthisYear is " + TotalNAVforallinvestments_PerthisYear)
                TotalNAVforallinvestments_PerthisYearList.append(TotalNAVforallinvestments_PerthisYear)
                print("Inside List2 is " + TotalNAVforallinvestments_PerthisYearList[year1 - 1])

                driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div").click()
                time.sleep(3)
                ActionChains(driver).key_down(Keys.DOWN).perform()
                time.sleep(3)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                for iat7 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath(
                            "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        # print("Loader is present")
                    except Exception:
                        time.sleep(1)
                        break
                time.sleep(10)

            print("len is: "+str(len(TotalNAVforallinvestments_PerthisYearList)))
            for comp in range(len(TotalNAVforallinvestments_PerthisYearList)):
                Value=TotalBenNAV_USD_PerthisYearList[comp]
                Value = Value.replace(" ", "")
                Value = re.sub(r'[?|$|€|£|!|,|-]', r'', Value)
                print("Value is :"+Value)

                Value1 = TotalNAVforallinvestments_PerthisYearList[comp]
                Value1 = Value1.replace(" ", "")
                Value1 = re.sub(r'[?|$|€|£|!|,|-]', r'', Value1)
                print("Value1 is :"+Value1)

                if (float(Value) - float(Value1)) == 0.0:
                    print("There is no difference of [ "+str(float(Value) - float(Value1))+" ] for year "+YearCounter[comp])

                else:
                    if (float(Value) - float(Value1)) >= 0.011:
                        print("There is a > difference of [ " + str(round(float(Value) - float(Value1),5)) + " ] for year " + YearCounter[comp])
                        TestResult.append("There is a difference of [ $" + str(round(float(Value) - float(Value1),5)) + " ] for year " + YearCounter[comp])
                        TestResultStatus.append("Fail")

                    if (float(Value) - float(Value1)) <= 0.011:
                        print("There is a < difference of [ " + str(round(float(Value) - float(Value1),5))  + " ] for year " + YearCounter[comp])
                print()
        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start),2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            TestResult.append(stringMainerror)
            TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


