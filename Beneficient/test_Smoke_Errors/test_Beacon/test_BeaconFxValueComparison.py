import datetime
import os
import re
import time

import fpdf
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  TestName = "test_BeaconFxValueComparison"
  description = "This is smoke test case to verify comparison of Fair Value, Initial Commitment, and Unfunded Commitment LC values with USD values considering FX rate value at Beacon Data Transfer Template grid"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_Beacon"
  global Exe
  Exe="Yes"

  ExcelFileName = "Execution"
  locx = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()
      for iat1 in range(1000):
          try:
              bool = driver.find_element_by_xpath(
                  "//div[@id='appian-working-indicator-hidden']").is_enabled()
          except Exception:
              time.sleep(1)
              break
      time.sleep(1)

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/EmailReportContent/Ben.png', 10, 8, 33)
              self.add_font('Arial', '', 'C:/Windows/Fonts/Arial.ttf', uni=True)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')


      pdf=PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.add_font('Arial', '', 'C:/Windows/Fonts/Arial.ttf', uni=True)
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 20, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1=TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      TestName = TestName.encode('latin-1', 'ignore').decode('latin-1')
      ct = ct.encode('latin-1', 'ignore').decode('latin-1')
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_BeaconFxValueCompare(test_setup):
    if Exe == "Yes":
        print()
        PageName = "Quarterly NAV Close"
        Ptitle1="Quarterly NAV Close - BIDS"
        driver.find_element_by_xpath("//*[@title='Quarterly NAV Close']").click()
        for iat2 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        PageTitle1 = driver.title
        try:
            assert Ptitle1 in PageTitle1, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " page not able to open")
            TestResultStatus.append("Fail")
        try:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_negative']").is_displayed()
            #print("Red flag present : " + str(bool))
            TestResult.append(PageName + " has a Red Flag at the top section")
            TestResultStatus.append("Fail")
        except Exception:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_positive']").is_displayed()
            #print("Green flag present : " + str(bool))
            TestResult.append(PageName + " has a Green Flag at the top section")
            TestResultStatus.append("Pass")

        PageName = "Beacon Template"
        Ptitle1 = "COR_BeaconDataTransferTemplate - BIDS"
        driver.find_element_by_xpath("//*[text() = '"+PageName+"']").click()
        for iat3 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(3)
        PageTitle1 = driver.title
        try:
            assert Ptitle1 in PageTitle1, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " page not able to open")
            TestResultStatus.append("Fail")

        for year in range(1,7):
            print()
            time.sleep(3)
            try:
                P = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div/span").text
            except Exception:
                time.sleep(10)
                P = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div/span").text

            skipyear=0
            Yearlist=['6/30/2020','9/30/2020','12/31/2020','3/31/2021']
            for j in range(len(Yearlist)):
                if Yearlist[j] in P:
                    print("Year skipped: " + Yearlist[j])
                    skipyear=1

            if skipyear==0:
                print("Year picked: " + P)
                FirstColumn = driver.find_element_by_xpath("//thead/tr[1]/th/div").text
                T_Rows = driver.find_elements_by_xpath("//tbody/tr")
                num = 2
                #len(T_Rows)
                if "Val. Err" in FirstColumn:
                    num=3
                    for ii2 in range(5):
                        #----------------To find ERR dosplayed for any Fund---------------------
                        Fund = driver.find_element_by_xpath("//tbody/tr[" + str(ii2 + 1) + "]/td[3]/p").text
                        Value1= driver.find_element_by_xpath("//tbody/tr[" + str(ii2 + 1) + "]/td[7]/div/p").text
                        Value2= driver.find_element_by_xpath("//tbody/tr[" + str(ii2 + 1) + "]/td[8]/div/p").text
                        Value3= driver.find_element_by_xpath("//tbody/tr[" + str(ii2 + 1) + "]/td[9]/div/p").text

                        Value4= driver.find_element_by_xpath("//tbody/tr[" + str(ii2 + 1) + "]/td[11]/div/p").text

                        Value5= driver.find_element_by_xpath("//tbody/tr[" + str(ii2 + 1) + "]/td[12]/div/p").text
                        Value6= driver.find_element_by_xpath(
                            "//tbody/tr[" + str(ii2 + 1) + "]/td[13]/div/p").text
                        Value7 = driver.find_element_by_xpath(
                            "//tbody/tr[" + str(ii2 + 1) + "]/td[14]/div/p").text

                        for ii4 in range(1,8):

                            if ii4==1:
                                Value=Value1
                            if ii4==2:
                                Value=Value2
                            if ii4==3:
                                Value=Value3
                            if ii4==4:
                                Value=Value4
                            if ii4==5:
                                Value=Value5
                            if ii4==6:
                                Value=Value6
                            if ii4==7:
                                Value=Value7

                            Value=Value.replace(" ", "")
                            Value = re.sub(r'[?|$|€|£|!|,]', r'', Value)
                            z = Value
                            bool1 = z.isupper() or z.islower()
                            if bool1 == True:
                                Value = re.sub(r'[a-z|A-Z]+', '', Value, re.I)
                            if ii4==1:
                                FairValue=float(Value)
                                #print("FairValue is "+str(FairValue))
                            if ii4==2:
                                InitialCommitment=float(Value)
                                #print("InitialCommitment is "+str(InitialCommitment))
                            if ii4==3:
                                UnfundedCommitment=float(Value)
                                #print("UnfundedCommitment is "+str(UnfundedCommitment))
                            if ii4==4:
                                FXRate=float(Value)
                                #print("FXRate is "+str(FXRate))
                            if ii4==5:
                                FairValueUSD=float(Value)
                                #print("FairValueUSD is "+str(FairValueUSD))
                            if ii4==6:
                                InitialCommitmentUSD=float(Value)
                                #print("InitialCommitmentUSD is "+str(InitialCommitmentUSD))
                            if ii4==7:
                                UnfundedCommitmentUSD=float(Value)
                                #print("UnfundedCommitmentUSD is "+str(UnfundedCommitmentUSD))

                        if  (round(FairValue*FXRate,2)- FairValueUSD)>0.011:
                            print()
                            print("FairValueUSD value do not match")
                            #print("FairValue*FXRate LC :"+str(round(FairValue*FXRate,2)))
                            #print("FairValueUSD :" + str(FairValueUSD))
                            #print("--------"+str(round((round(FairValue*FXRate,2)- FairValueUSD),4)))
                            Diff1=str(round((round(FairValue*FXRate,2)- FairValueUSD),4))
                            TestResult.append("Quarter[" +P+"]"+" Fair Value USD["+str(Value5)+"]do not match with Fair Value LC["+str(Value1)+"] and FX Rate["+str(Value4)+"] for Fund[" +Fund+"] Difference is "+Diff1)
                            TestResultStatus.append("Fail")


                        if  (round(InitialCommitment*FXRate,2)- InitialCommitmentUSD)>0.011:
                            print("InitialCommitmentUSD value do not match")
                            Diff2 = str(round((round(InitialCommitment * FXRate, 2) - InitialCommitmentUSD), 4))
                            TestResult.append(
                                "Quarter[" +P+"]"+" Initial Commitment USD["+str(Value6)+"]do not match with Initial Commitment LC["+str(Value2)+"] and FX Rate["+str(Value4)+"] for Fund [" + Fund + "] Difference is "+Diff2)
                            TestResultStatus.append("Fail")

                        if  (round(UnfundedCommitment*FXRate,2)- UnfundedCommitmentUSD)>0.011:
                            print("UnfundedCommitmentUSD value do not match")
                            Diff3 = str(round((round(UnfundedCommitment * FXRate, 2) - UnfundedCommitmentUSD), 4))
                            TestResult.append(
                                "Quarter[" +P+"]"+" Unfunded Commitment USD["+str(Value7)+"]do not match with Unfunded Commitment LC["+str(Value3)+"] and FX Rate["+str(Value4)+"] for Fund [" + Fund + "] Difference is "+Diff3)
                            TestResultStatus.append("Fail")

                else:
                    num = 2
                    #len(T_Rows)
                    for ii3 in range(5):
                        #----------------To find ERR dosplayed for any Fund---------------------
                        Fund = driver.find_element_by_xpath("//tbody/tr[" + str(ii3 + 1) + "]/td[2]/p").text
                        Value1= driver.find_element_by_xpath("//tbody/tr[" + str(ii3 + 1) + "]/td[6]/div/p").text
                        Value2= driver.find_element_by_xpath("//tbody/tr[" + str(ii3 + 1) + "]/td[7]/div/p").text
                        Value3= driver.find_element_by_xpath("//tbody/tr[" + str(ii3 + 1) + "]/td[8]/div/p").text

                        Value4= driver.find_element_by_xpath("//tbody/tr[" + str(ii3 + 1) + "]/td[10]/div/p").text

                        Value5= driver.find_element_by_xpath("//tbody/tr[" + str(ii3 + 1) + "]/td[11]/div/p").text
                        Value6= driver.find_element_by_xpath(
                            "//tbody/tr[" + str(ii3 + 1) + "]/td[12]/div/p").text
                        Value7 = driver.find_element_by_xpath(
                            "//tbody/tr[" + str(ii3 + 1) + "]/td[13]/div/p").text

                        for ii5 in range(1,8):

                            if ii5==1:
                                Value=Value1
                            if ii5==2:
                                Value=Value2
                            if ii5==3:
                                Value=Value3
                            if ii5==4:
                                Value=Value4
                            if ii5==5:
                                Value=Value5
                            if ii5==6:
                                Value=Value6
                            if ii5==7:
                                Value=Value7

                            Value=Value.replace(" ", "")
                            Value = re.sub(r'[?|$|€|£|!|,]', r'', Value)
                            z = Value
                            bool1 = z.isupper() or z.islower()
                            if bool1 == True:
                                Value = re.sub(r'[a-z|A-Z]+', '', Value, re.I)
                            if ii5==1:
                                FairValue=float(Value)
                                #print("FairValue is "+str(FairValue))
                            if ii5==2:
                                InitialCommitment=float(Value)
                                #print("InitialCommitment is "+str(InitialCommitment))
                            if ii5==3:
                                UnfundedCommitment=float(Value)
                                #print("UnfundedCommitment is "+str(UnfundedCommitment))
                            if ii5==4:
                                FXRate=float(Value)
                                #print("FXRate is "+str(FXRate))
                            if ii5==5:
                                FairValueUSD=float(Value)
                                #print("FairValueUSD is "+str(FairValueUSD))
                            if ii5==6:
                                InitialCommitmentUSD=float(Value)
                                #print("InitialCommitmentUSD is "+str(InitialCommitmentUSD))
                            if ii5==7:
                                UnfundedCommitmentUSD=float(Value)
                                #print("UnfundedCommitmentUSD is "+str(UnfundedCommitmentUSD))

                        # if  round(FairValue*FXRate,2)!= FairValueUSD:
                        #     print()
                        #     print("FairValueUSD value do not match")
                        #     TestResult.append("Fair Value USD value do not match with Fair Value LC and FX Rate for Fund [ " +Fund+" ] for Quarter ["+P+" ]")
                        #     TestResultStatus.append("Fail")
                        #
                        # if  round(InitialCommitment*FXRate,2)!= InitialCommitmentUSD:
                        #     print("InitialCommitmentUSD value do not match")
                        #     TestResult.append(
                        #         "Initial Commitment USD value do not match with Initial Commitment LC and FX Rate for Fund [ " + Fund + " ] for Quarter ["+P+" ]")
                        #     TestResultStatus.append("Fail")
                        #
                        # if  round(UnfundedCommitment*FXRate,2)!= UnfundedCommitmentUSD:
                        #     print("UnfundedCommitmentUSD value do not match")
                        #     TestResult.append(
                        #         "Unfunded Commitment USD value do not match with Unfunded Commitment LC and FX Rate for Fund [ " + Fund + " ] for Quarter ["+P+" ]")
                        #     TestResultStatus.append("Fail")

                        if  (round(FairValue*FXRate,2)- FairValueUSD)>0.011:
                            print()
                            print("FairValueUSD value do not match")
                            #print("FairValue*FXRate LC :"+str(round(FairValue*FXRate,2)))
                            #print("FairValueUSD :" + str(FairValueUSD))
                            #print("--------"+str(round((round(FairValue*FXRate,2)- FairValueUSD),4)))
                            Diff1=str(round((round(FairValue*FXRate,2)- FairValueUSD),4))
                            TestResult.append("Quarter[" +P+"]"+" Fair Value USD["+str(Value5)+"]do not match with Fair Value LC["+str(Value1)+"] and FX Rate["+str(Value4)+"] for Fund[" +Fund+"] Difference is "+Diff1)
                            TestResultStatus.append("Fail")


                        if  (round(InitialCommitment*FXRate,2)- InitialCommitmentUSD)>0.011:
                            print("InitialCommitmentUSD value do not match")
                            Diff2 = str(round((round(InitialCommitment * FXRate, 2) - InitialCommitmentUSD), 4))
                            TestResult.append(
                                "Quarter[" +P+"]"+" Initial Commitment USD["+str(Value6)+"]do not match with Initial Commitment LC["+str(Value2)+"] and FX Rate["+str(Value4)+"] for Fund [" + Fund + "] Difference is "+Diff2)
                            TestResultStatus.append("Fail")

                        if  (round(UnfundedCommitment*FXRate,2)- UnfundedCommitmentUSD)>0.011:
                            print("UnfundedCommitmentUSD value do not match")
                            Diff3 = str(round((round(UnfundedCommitment * FXRate, 2) - UnfundedCommitmentUSD), 4))
                            TestResult.append(
                                "Quarter[" +P+"]"+" Unfunded Commitment USD["+str(Value7)+"]do not match with Unfunded Commitment LC["+str(Value3)+"] and FX Rate["+str(Value4)+"] for Fund [" + Fund + "] Difference is "+Diff3)
                            TestResultStatus.append("Fail")

            driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[4]/div[2]/div/div[2]/div/div").click()
            time.sleep(3)
            ActionChains(driver).key_down(Keys.DOWN).perform()
            time.sleep(3)
            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
            for iat4 in range(1000):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                    #print("Loader is present")
                except Exception:
                    time.sleep(1)
                    break
            time.sleep(10)

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke_Errors/PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for ii6 in range(1, 100):
            if sheet.cell(ii6, 1).value == check:
                sheet.cell(row=ii6, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


