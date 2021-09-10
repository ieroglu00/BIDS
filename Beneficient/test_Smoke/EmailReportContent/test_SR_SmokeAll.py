import os
import smtplib
from email.message import EmailMessage
import openpyxl
import pytest

@pytest.mark.smoke
def test_ReportSendSmokeAll():
    DirectoryName = []
    PDFName1=[]
    TestName=[]
    TestDescription = []
    ExcelFileName = "FileName"
    loc = ('C:/BIDS/beneficienttest/Beneficient/PDFFileNameData/' + ExcelFileName + '.xlsx')
    wb=openpyxl.load_workbook(loc)
    sheet = wb.active
    for i in range(1, 100):
        if sheet.cell(i, 1).value == None:
            break
        else:
            PDFName1.append(sheet.cell(i, 2).value)
            DirectoryName.append(sheet.cell(i, 3).value)
            TestName.append(sheet.cell(i, 1).value)
            TestDescription.append(sheet.cell(i, 4).value)
    msg=EmailMessage()
    msg['Subject']='Test SUITE Automation Report'
    msg['From']='Neeraj'
    msg['To']='neeraj.kumar@crochetech.com,srinath.jandhyala@beneficient.com,Scott.Dacus@beneficient.com,adam.hunt@bitsinglass.com,nicholas.wurster@bitsinglass.com'

    with open('C:/BIDS/beneficienttest/Beneficient/test_Smoke/EmailReportContent/EmTemp.txt') as myfile:
        data=myfile.read()
    msg.set_content(data)
    A="Hi Team\nHere is the test summary report of a Test Suit \n\nBelow test scenarios are covered \n"
    C="\n\nPlease find attached PDFs of test scenarios results\n\n\nMany Thanks\nNeeraj"
    B = ""
    for io in range(len(TestName)):
        B = B + " \n\n"+str(io+1)+") " + "".join(TestName[io])+" => "+"".join(TestDescription[io])
    print(B)
    msg.set_content(A+B+C)
    i=0
    for file in PDFName1:
        print(file)
        with open('C:/BIDS/beneficienttest/Beneficient/test_Smoke/'+DirectoryName[i]+'/'+file, 'rb') as f:
            file_data = f.read()
            file_name = TestName[i]+".pdf"
        msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)
        i=i+1
    server=smtplib.SMTP_SSL('smtp.gmail.com',465)
    server.login("neeraj.kumar@bitsinglass.com","Motorola@408")

    server.send_message(msg)
    print("Test Report sent")
    #os.remove('C:/BIDS/beneficienttest/Beneficient/test_Smoke/'+DirectoryName+'/'+PDFName)
    server.quit()
