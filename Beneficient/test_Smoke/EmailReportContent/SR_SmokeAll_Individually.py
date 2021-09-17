import os
import smtplib
from email.message import EmailMessage
import openpyxl
import pytest

@pytest.mark.smoke
def test_ReportSendSmokeAll():
    print()
    DirectoryName = []
    PDFName1=[]
    TestName=[]
    TestDescription = []
    TestStatus = []
    SendStatus = []
    AttachmntAdded= []

    #-------------------To read content to send in e-Mail--------------------
    ExcelFileName = "FileName"
    loc = ('C:/BIDS/beneficienttest/Beneficient/PDFFileNameData/' + ExcelFileName + '.xlsx')
    wb=openpyxl.load_workbook(loc)
    sheet = wb.active
    for i in range(1, 100):
        if sheet.cell(i, 1).value == None:
            break
        else:
            PDFName1.append(sheet.cell(i, 2).value)
            DirectoryName.append(sheet.cell(i, 3).value)
            TestName.append(sheet.cell(i, 1).value)
            TestDescription.append(sheet.cell(i, 4).value)
            TestStatus.append(sheet.cell(i, 5).value)
            SendStatus.append(sheet.cell(i, 6).value)
    msg=EmailMessage()
    msg['Subject']='Test SUITE Automation Report [Smoke Test 1]'
    msg['From']='Test Automation Team'
    msg['To']='neeraj.kumar@crochetech.com'

    A="Hi Team\nHere is the test summary report of Smoke Test 1 (To verify all links, pages, green flags in a module) \n\nBelow test scenarios are covered \n"
    C="\n\nPlease find attached PDFs of test scenarios results\nNote: Attachments are only for FAILED test cases\n\n\nMany Thanks\nNeeraj"
    B = ""
    for io in range(len(TestName)):
        try:
            B = B + " \n\n"+str(io+1)+") " + "".join(TestName[io])+" => "+"".join(TestDescription[io])+" => "+"".join(TestStatus[io])
        except Exception:
            print("No attachment details to add in email description")
    #print(B)
    msg.set_content(A+B+C)
    #-----------------------------------------------------------------------

    # ------------------To add attachments in the report email--------------
    i=0
    for file in PDFName1:
        print()
        try:
            #print(file)
            if SendStatus[i]=="Send Only when Fail=Yes" and  TestStatus[i]=="Fail":
                    with open('C:/BIDS/beneficienttest/Beneficient/test_Smoke/'+DirectoryName[i]+'/'+file, 'rb') as f:
                        file_data = f.read()
                        file_name = TestName[i]+".pdf"
                    msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)
                    AttachmntAdded.append("Yes")
            if SendStatus[i] == "Send Only when Fail=No":
                    with open('C:/BIDS/beneficienttest/Beneficient/test_Smoke/'+DirectoryName[i]+'/'+file, 'rb') as f:
                        file_data = f.read()
                        file_name = TestName[i]+".pdf"
                    msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)
                    AttachmntAdded.append("Yes")
        except Exception as e1:
            print("No Attachment found to Add")
            #print(e1)
        i = i + 1
    #-----------------------------------------------------------------------

    # ----------------------------SMTP setup--------------------------------
    server=smtplib.SMTP_SSL('smtp.gmail.com',465)
    server.login("neeraj.kumar@bitsinglass.com","Motorola@408")
    #-----------------------------------------------------------------------

    #---------------------------------Sending email-------------------------
    for io1 in range(len(AttachmntAdded)):
        if AttachmntAdded[io1] == "Yes":
            server.send_message(msg)
            print("Test Report sent")
            break
    #-----------------------------------------------------------------------

    #-----------------To delete pdf report files----------------------------
    ii=0
    for ii in range(0,len(PDFName1)):
        print()
        try:
            os.remove('C:/BIDS/beneficienttest/Beneficient/test_Smoke/'+DirectoryName[ii]+'/'+PDFName1[ii])
        except Exception:
            print("No Attachment found to delete")
    #-----------------------------------------------------------------------
    server.quit()
