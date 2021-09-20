import datetime
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure

@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  TestName = "test_AllModulesVerify"
  description = "This is smoke test case to verify all Top Menu modules are opening with Green flag at the top section of each module"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "AllMainModules"
  global Exe
  Exe="Yes"

  ExcelFileName = "Execution"
  locx = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke/Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      driver.find_element_by_xpath("//input[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image('C:/BIDS/beneficienttest/Beneficient/test_Smoke/EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 20, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestResult[i1], 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = ('C:/BIDS/beneficienttest/Beneficient/PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke/Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_AllModulesVerify(test_setup):
    if Exe == "Yes":
        print()
        PageName="Funds"
        Ptitle1="Appian for The Beneficient Company (TEST)"
        PageTitle1=driver.title
        for iat1 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        try:
            assert Ptitle1 in PageTitle1, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " page not able to open")
            TestResultStatus.append("Fail")
        try:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_negative']").is_displayed()
            print("Red flag present : " + str(bool))
            TestResult.append(PageName + " has a Red Flag at the top section")
            TestResultStatus.append("Fail")
        except Exception:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_positive']").is_displayed()
            print("Green flag present : " + str(bool))
            TestResult.append(PageName + " has a Green Flag at the top section")
            TestResultStatus.append("Pass")

        PageName = "Investments"
        driver.find_element_by_xpath("//*[@title='Investments']").click()
        for iat2 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        Ptitle2 = "Investments - BIDS"
        PageTitle2 = driver.title
        try:
            assert Ptitle2 in PageTitle2, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " page not able to open")
            TestResultStatus.append("Fail")
        try:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_negative']").is_displayed()
            print("Red flag present : " + str(bool))
            TestResult.append(PageName + " has a Red Flag at the top section")
            TestResultStatus.append("Fail")
        except Exception:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_positive']").is_displayed()
            #print("Green flag present : " + str(bool))
            TestResult.append(PageName + " has a Green Flag at the top section")
            TestResultStatus.append("Pass")

        PageName = "Transactions"
        driver.find_element_by_xpath("//*[@title='Transactions']").click()
        for iat3 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        Ptitle3 = "Transactions - BIDS"
        PageTitle3 = driver.title
        try:
            assert Ptitle3 in PageTitle3, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " page not able to open")
            TestResultStatus.append("Fail")

        try:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_negative']").is_displayed()
            print("Red flag present : " + str(bool))
            TestResult.append(PageName + " has a Red Flag at the top section")
            TestResultStatus.append("Fail")
        except Exception:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_positive']").is_displayed()
            print("Green flag present : " + str(bool))
            TestResult.append(PageName + " has a Green Flag at the top section")
            TestResultStatus.append("Pass")

        PageName = "Liquid Trusts"
        driver.find_element_by_xpath("//*[@title='Liquid Trusts']").click()
        for iat4 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        Ptitle4 = "Liquid Trusts - BIDS"
        PageTitle4 = driver.title
        try:
            assert Ptitle4 in PageTitle4, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " page not able to open")
            TestResultStatus.append("Fail")

        try:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_negative']").is_displayed()
            print("Red flag present : " + str(bool))
            TestResult.append(PageName + " has a Red Flag at the top section")
            TestResultStatus.append("Fail")
        except Exception:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_positive']").is_displayed()
            print("Green flag present : " + str(bool))
            TestResult.append(PageName + " has a Green Flag at the top section")
            TestResultStatus.append("Pass")

        PageName = "Quarterly NAV Close"
        driver.find_element_by_xpath("//*[@title='Quarterly NAV Close']").click()
        for iat5 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)
        Ptitle5 = "Quarterly NAV Close - BIDS"
        PageTitle5 = driver.title
        try:
            assert Ptitle5 in PageTitle5, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            TestResult.append(PageName + " page not able to open")
            TestResultStatus.append("Fail")

        try:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_negative']").is_displayed()
            print("Red flag present : " + str(bool))
            TestResult.append(PageName + " has a Red Flag at the top section")
            TestResultStatus.append("Fail")
        except Exception:
            bool=driver.find_element_by_xpath("//span[@class='IconWidget---large IconWidget---color_positive']").is_displayed()
            print("Green flag present : " + str(bool))
            TestResult.append(PageName + " has a Green Flag at the top section")
            TestResultStatus.append("Pass")
    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = ('C:/BIDS/beneficienttest/Beneficient/PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


