import datetime
import time

import openpyxl
import xlrd
from fpdf import FPDF
from openpyxl import Workbook
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import pytest
from selenium import webdriver
import allure

@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  TestName = "test_AllModulesVerify"
  description = "This is smoke test case to verify all Top Menu modules are opening"
  TestResult = []
  TestResultStatus = []

  driver=webdriver.Chrome(executable_path="C:\Laptop Data\Work\Python\chromedriver_win32 (1)\chromedriver")
  driver.implicitly_wait(10)
  driver.maximize_window()
  driver.get("https://beneficienttest.appiancloud.com/suite/")
  enter_username("neeraj.kumar")
  enter_password("Crochet@786")
  driver.find_element_by_xpath("//input[@type='submit']").click()

  yield
  ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
  ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

  class PDF(FPDF):
      def header(self):
          self.image('ben.png', 10, 8, 33)
          self.set_font('Arial', 'B', 15)
          self.cell(73)
          self.cell(35, 10, ' Test Report ', 1, 1, 'B')
          self.set_font('Arial', 'I', 10)
          self.cell(150)
          self.cell(30, 10, ctReportHeader, 0, 0, 'C')
          self.ln(20)

      def footer(self):
          self.set_y(-15)
          self.set_font('Arial', 'I', 8)
          self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

  pdf = PDF()
  pdf.alias_nb_pages()
  pdf.add_page()
  pdf.set_font('Times', '', 12)
  pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
  pdf.cell(0, 10, "Description:  "+description, 0, 1)

  for i in range(len(TestResult)):
     pdf.set_fill_color(255, 255, 255)
     if (TestResultStatus[i] == "Fail"):
         print("Fill Red color")
         pdf.set_fill_color(255, 0, 0)
     pdf.cell(0, 20,str(i+1)+")  "+TestResult[i], 0, 1,fill=True)
  pdf.output(TestName+"_" + ct + ".pdf", 'F')

  ExcelFileName = "neeraj"
  loc = ('C:/Users/Neeraj/PycharmProjects/beneficienttest/Beneficient/' + ExcelFileName + '.xlsx')
  wb = openpyxl.load_workbook(loc)
  sheet = wb.active
  sheet.cell(row=1, column=1).value = TestName+"_" + ct + ".pdf"
  wb.save(loc)

  driver.quit()

@pytest.mark.smoke
def test_AllModulesVerify(test_setup):
    print()
    PageName="Funds"
    Ptitle1="Funds - BIDS"
    PageTitle1=driver.title
    for iat1 in range(1000):
        try:
            bool = driver.find_element_by_xpath(
                "//div[@id='appian-working-indicator-hidden']").is_enabled()
        except Exception:
            time.sleep(1)
            break
    time.sleep(2)
    try:
        assert Ptitle1 in PageTitle1, PageName + " not able to open"
        TestResult.append(PageName + " page Opened successfully")
        TestResultStatus.append("Pass")
    except Exception:
        TestResult.append(PageName + " page not able to open")
        TestResultStatus.append("Fail")

    # PageName = "Investments"
    # driver.find_element_by_xpath("//*[@title='Investments']").click()
    # for iat2 in range(1000):
    #     try:
    #         bool = driver.find_element_by_xpath(
    #             "//div[@id='appian-working-indicator-hidden']").is_enabled()
    #     except Exception:
    #         time.sleep(1)
    #         break
    # time.sleep(2)
    # Ptitle2 = "Investments - BIDSs"
    # PageTitle2 = driver.title
    # try:
    #     assert Ptitle2 in PageTitle2, PageName + " not able to open"
    #     TestResult.append(PageName + " page Opened successfully")
    #     TestResultStatus.append("Pass")
    # except Exception:
    #     TestResult.append(PageName + " page not able to open")
    #     TestResultStatus.append("Fail")
    #
    # PageName = "Transactions"
    # driver.find_element_by_xpath("//*[@title='Transactions']").click()
    # for iat3 in range(1000):
    #     try:
    #         bool = driver.find_element_by_xpath(
    #             "//div[@id='appian-working-indicator-hidden']").is_enabled()
    #     except Exception:
    #         time.sleep(1)
    #         break
    # time.sleep(2)
    # Ptitle3 = "Transactions - BIDS"
    # PageTitle3 = driver.title
    # try:
    #     assert Ptitle3 in PageTitle3, PageName + " not able to open"
    #     TestResult.append(PageName + " page Opened successfully")
    #     TestResultStatus.append("Pass")
    # except Exception:
    #     TestResult.append(PageName + " page not able to open")
    #     TestResultStatus.append("Fail")

    # PageName = "Liquid Trusts"
    # driver.find_element_by_xpath("//*[@title='Liquid Trusts']").click()
    # for iat4 in range(1000):
    #     try:
    #         bool = driver.find_element_by_xpath(
    #             "//div[@id='appian-working-indicator-hidden']").is_enabled()
    #     except Exception:
    #         time.sleep(1)
    #         break
    # time.sleep(2)
    # Ptitle4 = "Liquid Trusts - BIDSw"
    # PageTitle4 = driver.title
    # try:
    #     assert Ptitle4 in PageTitle4, PageName + " not able to open"
    #     TestResult.append(PageName + " page Opened successfully")
    # except Exception:
    #     TestResult.append(PageName + " page not able to open")
    #
    # PageName = "Quarterly NAV Close"
    # driver.find_element_by_xpath("//*[@title='Quarterly NAV Close']").click()
    # for iat5 in range(1000):
    #     try:
    #         bool = driver.find_element_by_xpath(
    #             "//div[@id='appian-working-indicator-hidden']").is_enabled()
    #     except Exception:
    #         time.sleep(1)
    #         break
    # time.sleep(2)
    # Ptitle5 = "Quarterly NAV Close - BIDS"
    # PageTitle5 = driver.title
    # try:
    #     assert Ptitle5 in PageTitle5, PageName + " not able to open"
    #     TestResult.append(PageName + " page Opened successfully")
    # except Exception:
    #     TestResult.append(PageName + " page not able to open")


